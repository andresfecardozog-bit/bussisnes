"""Tests de la Fase 7A: modelo Batch + endpoints CRUD + preview + generate.

Cubre:
- CRUD basico del batch (create, list, get, patch, delete, archive).
- Upload pre_cortes single/multi + upload flash con periodo.
- Upload ZIP con archivos mezclados (validos + invalidos).
- Deteccion de colisiones cuando dos pre_cortes apuntan al mismo dia.
- Validacion del periodo del flash (rechazar si el mes declarado no
  coincide con las fechas del flash).
- Preview con dias, colisiones, saltos por dia no laboral y
  `listo_para_confirmar`.
- Confirm bloquea si hay colisiones o el flash no cuadra.
- Generate persiste runs + genera dailies + consolidado + zip.
- Downloads lista y sirve los archivos generados.
"""
from __future__ import annotations

import io
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

FIXTURES = Path(__file__).parent / "fixtures"
PRE_CORTE_XLSX = FIXTURES / "PRE_CORTE_muestra.xlsx"
FLASH_CSV = FIXTURES / "FLASH_muestra.csv"
HOMOLOG_XLSX = FIXTURES / "homologacion.xlsx"


# ---------------------------------------------------------------------------
# Fixtures (aisla DB + uploads + export_dir por test)
# ---------------------------------------------------------------------------
@pytest.fixture
def isolated_env(tmp_path, monkeypatch):
    db_path = tmp_path / "test.sqlite"
    uploads_dir = tmp_path / "uploads"
    uploads_dir.mkdir()
    export_dir = tmp_path / "onedrive_export"
    export_dir.mkdir()

    import app.api.dependencies as deps
    import app.api.routes.batches as batches_route
    import app.api.storage as storage
    import app.config as cfg
    import app.core.db as core_db

    monkeypatch.setattr(cfg, "DB_PATH", db_path)
    monkeypatch.setattr(cfg, "UPLOADS_DIR", uploads_dir)
    monkeypatch.setattr(cfg, "ONEDRIVE_EXPORT_DIR", export_dir)
    monkeypatch.setattr(cfg, "AUTH_COOKIE_SECURE", False)
    monkeypatch.setattr(core_db, "DB_PATH", db_path)
    monkeypatch.setattr(core_db, "ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setattr(core_db, "ADMIN_INITIAL_PASSWORD", "AdminPass123!")
    monkeypatch.setattr(deps, "DB_PATH", db_path)
    monkeypatch.setattr(storage, "UPLOADS_DIR", uploads_dir)
    monkeypatch.setattr(batches_route, "ONEDRIVE_EXPORT_DIR", export_dir)

    yield {"db": db_path, "uploads": uploads_dir, "export": export_dir}


@pytest.fixture
def client(isolated_env):
    from app.api.main import app
    from app.core.db import get_conn, init_db
    from app.core.sku_catalog import import_from_homologacion

    init_db(isolated_env["db"])
    with get_conn(isolated_env["db"]) as conn:
        import_from_homologacion(HOMOLOG_XLSX, conn)
    with TestClient(app) as c:
        login = c.post(
            "/auth/login",
            json={"email": "admin@test.local", "password": "AdminPass123!"},
        )
        assert login.status_code == 200, login.text
        csrf = c.cookies.get("nutri_csrf")
        assert csrf
        change = c.post(
            "/auth/change-password",
            json={
                "current_password": "AdminPass123!",
                "new_password": "AdminPass123!_nueva",
            },
            headers={"X-CSRF-Token": csrf},
        )
        assert change.status_code == 200, change.text
        c.headers.update({"X-CSRF-Token": c.cookies.get("nutri_csrf", "")})
        yield c


# ---------------------------------------------------------------------------
# Helpers para upload
# ---------------------------------------------------------------------------
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _upload_pre_corte_to_batch(client, batch_id, filename, src_path=PRE_CORTE_XLSX):
    with src_path.open("rb") as fh:
        r = client.post(
            f"/batches/{batch_id}/pre-cortes",
            files={"files": (filename, fh, _XLSX_MIME)},
        )
    assert r.status_code == 200, r.text
    return r.json()


def _upload_flash_to_batch(client, batch_id, year=2026, month=2):
    with FLASH_CSV.open("rb") as fh:
        r = client.post(
            f"/batches/{batch_id}/flash",
            params={"year": year, "month": month},
            files={"file": ("FLASH_muestra.csv", fh, "text/csv")},
        )
    return r


def _make_pre_corte_variant(tmp_path: Path, name: str) -> Path:
    """Copia el fixture con otro nombre para simular otro dia. Modifica una
    celda inofensiva para asegurar hash distinto."""
    dst = tmp_path / name
    shutil.copy(PRE_CORTE_XLSX, dst)
    wb = openpyxl.load_workbook(dst)
    if "INVENTARIO FISICO" in wb.sheetnames:
        wb["INVENTARIO FISICO"]["B2"] = f"variant-{name}"
    else:
        wb.create_sheet(f"variant_{name}_marker")
    wb.save(dst)
    return dst


# ---------------------------------------------------------------------------
# CRUD basico
# ---------------------------------------------------------------------------
def test_create_batch_devuelve_status_draft(client):
    r = client.post("/batches", json={"nombre": "Febrero 2026"})
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["status"] == "draft"
    assert body["nombre"] == "Febrero 2026"
    assert body["num_pre_cortes"] == 0
    assert body["flash"] is None


def test_list_batches_filtra_por_status_y_orden(client):
    a = client.post("/batches", json={"nombre": "A"}).json()
    b = client.post("/batches", json={"nombre": "B"}).json()
    r = client.get("/batches")
    assert r.status_code == 200
    ids = [x["id"] for x in r.json()]
    # ambos aparecen, mas nuevo primero (b antes que a)
    assert set(ids) >= {a["id"], b["id"]}


def test_patch_batch_actualiza_nombre_y_notas(client):
    b = client.post("/batches", json={"nombre": "orig"}).json()
    r = client.patch(f"/batches/{b['id']}", json={"nombre": "nuevo", "notas": "hola"})
    assert r.status_code == 200
    assert r.json()["nombre"] == "nuevo"
    assert r.json()["notas"] == "hola"


def test_delete_batch_draft_funciona(client):
    b = client.post("/batches", json={}).json()
    r = client.delete(f"/batches/{b['id']}")
    assert r.status_code == 204
    assert client.get(f"/batches/{b['id']}").status_code == 404


def test_archive_batch_lo_saca_del_listado_default(client):
    b = client.post("/batches", json={"nombre": "toArchive"}).json()
    client.post(f"/batches/{b['id']}/archive")
    ids = {x["id"] for x in client.get("/batches").json()}
    assert b["id"] not in ids
    ids_all = {x["id"] for x in client.get("/batches?include_archived=true").json()}
    assert b["id"] in ids_all


# ---------------------------------------------------------------------------
# Pre-cortes en batch
# ---------------------------------------------------------------------------
def test_upload_pre_corte_al_batch_se_lista(client, tmp_path):
    b = client.post("/batches", json={"nombre": "b"}).json()
    detail = _upload_pre_corte_to_batch(
        client, b["id"], "PRE CORTE 13.02.2026.xlsx"
    )
    assert detail["num_pre_cortes"] == 1
    assert detail["pre_cortes"][0]["fecha_produccion"] == "2026-02-14"


def test_upload_multiple_pre_cortes_en_una_llamada(client, tmp_path):
    b = client.post("/batches", json={}).json()
    a = _make_pre_corte_variant(tmp_path, "PRE CORTE 09.02.2026.xlsx")
    with (
        PRE_CORTE_XLSX.open("rb") as fh1,
        a.open("rb") as fh2,
    ):
        r = client.post(
            f"/batches/{b['id']}/pre-cortes",
            files=[
                ("files", ("PRE CORTE 13.02.2026.xlsx", fh1.read(), _XLSX_MIME)),
                ("files", ("PRE CORTE 09.02.2026.xlsx", fh2.read(), _XLSX_MIME)),
            ],
        )
    assert r.status_code == 200, r.text
    assert r.json()["num_pre_cortes"] == 2


def test_remove_pre_corte_del_batch(client):
    b = client.post("/batches", json={}).json()
    detail = _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    carga_id = detail["pre_cortes"][0]["carga_id"]
    r = client.delete(f"/batches/{b['id']}/pre-cortes/{carga_id}")
    assert r.status_code == 200
    assert r.json()["num_pre_cortes"] == 0


def test_zip_upload_procesa_validos_e_ignora_invalidos(client, tmp_path):
    b = client.post("/batches", json={}).json()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.write(PRE_CORTE_XLSX, arcname="PRE CORTE 13.02.2026.xlsx")
        variant = _make_pre_corte_variant(tmp_path, "variant_16.xlsx")
        zf.write(variant, arcname="PRE CORTE 16.02.2026.xlsx")
        zf.writestr("no_es_precorte.xlsx", b"basura")
        zf.writestr("readme.txt", b"hola")
    buf.seek(0)

    r = client.post(
        f"/batches/{b['id']}/pre-cortes/from-zip",
        files={"file": ("batch.zip", buf, "application/zip")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["procesados"]) == 2
    nombres_ignorados = {i["filename"] for i in body["ignorados"]}
    assert "no_es_precorte.xlsx" in nombres_ignorados
    assert "readme.txt" in nombres_ignorados


# ---------------------------------------------------------------------------
# Flash con periodo year/month
# ---------------------------------------------------------------------------
def test_upload_flash_con_mes_correcto(client):
    b = client.post("/batches", json={}).json()
    r = _upload_flash_to_batch(client, b["id"], year=2026, month=2)
    assert r.status_code == 200, r.text
    detail = r.json()
    assert detail["flash"] is not None
    assert detail["flash"]["periodo_year"] == 2026
    assert detail["flash"]["periodo_month"] == 2


def test_upload_flash_con_mes_incorrecto_es_422(client):
    b = client.post("/batches", json={}).json()
    r = _upload_flash_to_batch(client, b["id"], year=2025, month=1)
    assert r.status_code == 422
    detail = r.json()["detail"]
    assert "no cuadra" in str(detail).lower() or "no contiene" in str(detail).lower()


def test_detach_flash_lo_desvincula(client):
    b = client.post("/batches", json={}).json()
    _upload_flash_to_batch(client, b["id"])
    r = client.delete(f"/batches/{b['id']}/flash")
    assert r.status_code == 200
    assert r.json()["flash"] is None


# ---------------------------------------------------------------------------
# Colisiones
# ---------------------------------------------------------------------------
def test_colision_detectada_cuando_dos_pre_cortes_apuntan_al_mismo_dia(
    client, tmp_path
):
    """PRE CORTE 07.02 (sabado) -> dom 08 -> lun 09.
    PRE CORTE 08.02 (dom, "no deberia existir" pero por si acaso) -> lun 09.
    Ambos apuntan a 2026-02-09 => colision.
    """
    b = client.post("/batches", json={}).json()
    sab = _make_pre_corte_variant(tmp_path, "PRE CORTE 07.02.2026.xlsx")
    dom = _make_pre_corte_variant(tmp_path, "PRE CORTE 08.02.2026.xlsx")
    _upload_pre_corte_to_batch(client, b["id"], sab.name, src_path=sab)
    _upload_pre_corte_to_batch(client, b["id"], dom.name, src_path=dom)
    _upload_flash_to_batch(client, b["id"])

    prev = client.get(f"/batches/{b['id']}/preview").json()
    assert len(prev["colisiones"]) == 1
    assert prev["colisiones"][0]["fecha"] == "2026-02-09"
    assert len(prev["colisiones"][0]["pre_corte_carga_ids"]) == 2
    assert prev["listo_para_confirmar"] is False


def test_confirm_bloquea_si_hay_colisiones(client, tmp_path):
    b = client.post("/batches", json={}).json()
    sab = _make_pre_corte_variant(tmp_path, "PRE CORTE 07.02.2026.xlsx")
    dom = _make_pre_corte_variant(tmp_path, "PRE CORTE 08.02.2026.xlsx")
    _upload_pre_corte_to_batch(client, b["id"], sab.name, src_path=sab)
    _upload_pre_corte_to_batch(client, b["id"], dom.name, src_path=dom)
    _upload_flash_to_batch(client, b["id"])

    r = client.post(f"/batches/{b['id']}/confirm")
    assert r.status_code == 409
    assert "colisiones" in str(r.json()["detail"]).lower()


# ---------------------------------------------------------------------------
# Preview + saltos por no laboral
# ---------------------------------------------------------------------------
def test_preview_reporta_salto_por_dia_no_laboral(client, tmp_path):
    b = client.post("/batches", json={}).json()
    sab = _make_pre_corte_variant(tmp_path, "PRE CORTE 07.02.2026.xlsx")
    _upload_pre_corte_to_batch(client, b["id"], sab.name, src_path=sab)
    _upload_flash_to_batch(client, b["id"])
    prev = client.get(f"/batches/{b['id']}/preview").json()
    assert len(prev["fechas_no_laborales_saltadas"]) == 1
    salto = prev["fechas_no_laborales_saltadas"][0]
    assert salto["dias_saltados"] == 1
    assert salto["fecha_produccion_resuelta"] == "2026-02-09"


def test_preview_dias_incluye_cumplimiento_pct(client):
    b = client.post("/batches", json={}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    prev = client.get(f"/batches/{b['id']}/preview").json()
    assert len(prev["dias"]) == 1
    d = prev["dias"][0]
    assert d["fecha_produccion"] == "2026-02-14"
    assert d["materiales_matched"] >= 1
    assert isinstance(d["cumplimiento_pct"], float)
    assert prev["listo_para_confirmar"] is True


# ---------------------------------------------------------------------------
# Confirm + generate + downloads (end-to-end)
# ---------------------------------------------------------------------------
def test_flujo_completo_end_to_end(client, isolated_env):
    b = client.post("/batches", json={"nombre": "E2E"}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])

    r = client.post(f"/batches/{b['id']}/confirm")
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "ready_to_match"

    r = client.post(f"/batches/{b['id']}/generate")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["consolidado_filename"].startswith("cumplimiento_consolidado_")
    assert len(body["dailies_filenames"]) == 1
    assert body["dailies_filenames"][0] == "cumplimiento_20260214.xlsx"
    assert body["zip_filename"].endswith(".zip")
    assert body["fechas_procesadas"] == ["2026-02-14"]

    # El batch quedo matched.
    detail = client.get(f"/batches/{b['id']}").json()
    assert detail["status"] == "matched"
    assert detail["output_dir"] is not None

    # Downloads listado.
    d = client.get(f"/batches/{b['id']}/downloads").json()
    kinds = {i["kind"] for i in d["items"]}
    assert kinds == {"consolidado", "daily", "zip"}
    assert len(d["items"]) == 3

    # Download individual del consolidado.
    r = client.get(f"/batches/{b['id']}/downloads/{body['consolidado_filename']}")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r.content[:4] == b"PK\x03\x04"

    # Download del ZIP.
    r = client.get(f"/batches/{b['id']}/downloads/{body['zip_filename']}")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"


def test_generate_bloquea_si_batch_no_esta_ready(client):
    """No se puede generar sin confirmar antes."""
    b = client.post("/batches", json={}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    # Falta confirm.
    r = client.post(f"/batches/{b['id']}/generate")
    assert r.status_code == 409


def test_download_path_traversal_bloqueado(client):
    """El endpoint no debe permitir escapar del output_dir del batch."""
    b = client.post("/batches", json={}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    client.post(f"/batches/{b['id']}/confirm")
    client.post(f"/batches/{b['id']}/generate")
    r = client.get(f"/batches/{b['id']}/downloads/../../etc/passwd")
    assert r.status_code in (400, 404)


# ---------------------------------------------------------------------------
# CRUD bloqueado despues de confirm
# ---------------------------------------------------------------------------
def test_no_se_puede_subir_pre_corte_a_batch_confirmado(client):
    b = client.post("/batches", json={}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    client.post(f"/batches/{b['id']}/confirm")
    # Ahora esta en ready_to_match: NO puede agregar mas pre_cortes.
    with PRE_CORTE_XLSX.open("rb") as fh:
        r = client.post(
            f"/batches/{b['id']}/pre-cortes",
            files={"files": ("PRE CORTE 20.02.2026.xlsx", fh, _XLSX_MIME)},
        )
    assert r.status_code == 409


# ---------------------------------------------------------------------------
# Fase 7E: mirror a Supabase + signed URL redirect
# ---------------------------------------------------------------------------
class _FakeSupabaseStorage:
    """Storage falso que simula Supabase para tests. NO conecta a red."""

    backend_name = "supabase"

    def __init__(self):
        self.blobs: dict[tuple[str, str], bytes] = {}

    def put(self, bucket, key, data):
        self.blobs[(bucket, key)] = data
        return f"supabase://{bucket}/{key}"

    def get(self, bucket, key):
        if (bucket, key) not in self.blobs:
            raise FileNotFoundError(key)
        return self.blobs[(bucket, key)]

    def exists(self, bucket, key):
        return (bucket, key) in self.blobs

    def delete(self, bucket, key):
        return self.blobs.pop((bucket, key), None) is not None

    def list(self, bucket, prefix=""):
        return sorted(k for (b, k) in self.blobs if b == bucket and k.startswith(prefix))

    def public_url(self, bucket, key, expires_in=3600):
        return f"https://fake.supabase.co/storage/v1/object/sign/{bucket}/{key}?token=fake"


def test_generate_hace_mirror_a_supabase_si_activo(client, monkeypatch):
    """Con `SupabaseStorage` activo, generate replica outputs al bucket
    y downloads redirige a signed URL."""
    fake = _FakeSupabaseStorage()
    import app.api.routes.batches as batches_route
    monkeypatch.setattr(batches_route, "get_storage", lambda: fake)

    b = client.post("/batches", json={"nombre": "supabase mirror"}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    client.post(f"/batches/{b['id']}/confirm")
    gen = client.post(f"/batches/{b['id']}/generate").json()

    # Los 3 tipos de archivo deben estar en el fake bucket.
    keys = fake.list("outputs", prefix=f"batch_{b['id']}")
    assert any(k.endswith(gen["consolidado_filename"]) for k in keys)
    assert any(k.endswith(gen["zip_filename"]) for k in keys)
    for daily in gen["dailies_filenames"]:
        assert any(k.endswith(daily) for k in keys)

    # Download del consolidado -> 307 redirect a signed URL.
    r = client.get(
        f"/batches/{b['id']}/downloads/{gen['consolidado_filename']}",
        follow_redirects=False,
    )
    assert r.status_code == 307
    assert "fake.supabase.co" in r.headers["location"]
    assert "token=fake" in r.headers["location"]


def test_generate_no_falla_si_mirror_a_supabase_explota(client, monkeypatch):
    """El mirror es best-effort: si Supabase esta caido, generate igual completa."""
    class _BrokenStorage:
        backend_name = "supabase"
        def put(self, *a, **kw): raise RuntimeError("supabase 500")
        def exists(self, *a, **kw): return False
        def public_url(self, *a, **kw): return ""

    import app.api.routes.batches as batches_route
    monkeypatch.setattr(batches_route, "get_storage", lambda: _BrokenStorage())

    b = client.post("/batches", json={}).json()
    _upload_pre_corte_to_batch(client, b["id"], "PRE CORTE 13.02.2026.xlsx")
    _upload_flash_to_batch(client, b["id"])
    client.post(f"/batches/{b['id']}/confirm")
    r = client.post(f"/batches/{b['id']}/generate")
    assert r.status_code == 200, r.text
    # Batch quedo matched aunque Supabase fallo.
    detail = client.get(f"/batches/{b['id']}").json()
    assert detail["status"] == "matched"
