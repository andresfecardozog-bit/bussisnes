"""Tests del renderer Excel declarativo (Fase 4, app/platform/render_excel.py).

Cubre:
- El .xlsx abre con openpyxl y tiene exactamente las hojas del spec.
- Portada: header 'Indicador' + KPI de cumplimiento con semaforo aplicado.
- Hojas tabla son Excel Tables nativas; porcentajes como fraccion 0-1.
- Las keys internas `_k{i}` se renombran a las join keys del profile.
- spec.columns filtra/ordena e ignora columnas inexistentes.
- group_by aplica la agrupacion visual (valores repetidos ocultos).
- Contrato extendido CEN vs SAP: hojas breakdown, tablas de nivel de
  servicio en la portada, hojas del data_model como ListObjects con nombre
  exacto, y end-to-end con el borrador CEN + fixtures reales.
- Puritanismo: render_excel.py no importa styles de openpyxl directamente.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app.platform.engine import GenericMatchResult, run_profile
from app.platform.profile import MatchProfile
from app.platform.render_excel import render_excel

PROFILES_DIR = Path(__file__).resolve().parents[1] / "profiles"
FIXTURES = Path(__file__).parent / "fixtures"
CEN_JUNIO = FIXTURES / "cen" / "cen_junio_muestra.xlsx"
SAP_MUESTRA = FIXTURES / "cen" / "sap_junio_muestra.xlsx"


# ---------------------------------------------------------------------------
# Fixtures sinteticos (sin DB, sin archivos fuente)
# ---------------------------------------------------------------------------

def _mini_profile(sheets: list[dict] | None = None) -> MatchProfile:
    """Profile minimo valido con report.excel configurable."""
    if sheets is None:
        sheets = [
            {"name": "Portada", "kind": "portada"},
            {"name": "Resumen", "kind": "kpi_resumen", "source": "kpis"},
            {"name": "Detalle", "kind": "tabla", "source": "matched"},
            {"name": "No_Cruzados", "kind": "tabla", "source": "no_cruzados"},
        ]
    return MatchProfile.model_validate({
        "profile_id": "mini_test",
        "version": 1,
        "descripcion": "Cruce sintetico de prueba plan vs real",
        "left": {
            "role": "plan",
            "label": "Plan",
            "loader": {
                "type": "tabular",
                "columns": [{"name": "material", "source": "MATERIAL"}],
            },
        },
        "right": {
            "role": "real",
            "label": "Real",
            "loader": {
                "type": "tabular",
                "columns": [{"name": "material", "source": "MATERIAL"}],
            },
        },
        "join": {"keys": [{"left": "material", "right": "material"}]},
        "kpis": [
            {
                "id": "cumplimiento_global_pct",
                "label": "Cumplimiento global (%)",
                "op": "ratio_pct_of_sums",
                "numerator": "real",
                "denominator": "plan",
                "semaforo": {
                    "verde_min": 95.0, "verde_max": 105.0,
                    "amarillo_min": 85.0, "amarillo_max": 115.0,
                },
            },
            {"id": "plan_total", "label": "Plan total", "op": "sum", "numerator": "plan"},
            {"id": "filas", "label": "Filas cruzadas", "op": "count"},
        ],
        "report": {"excel": {"filename_prefix": "mini", "sheets": sheets}},
    })


def _mini_result() -> GenericMatchResult:
    matched = pd.DataFrame({
        "_k0": ["30018", "30049", "30055"],
        "grupo": ["A", "A", "B"],
        "plan": [100.0, 200.0, 300.0],
        "real": [95.0, 210.0, 290.0],
        "cumplimiento_pct": [95.0, 105.0, 96.67],
    })
    solo_left = pd.DataFrame({"_k0": ["30099"], "plan": [50.0]})
    solo_right = pd.DataFrame(columns=["_k0", "real"])
    no_cruzados = pd.DataFrame({
        "origen": ["plan"],
        "key": ["30099"],
        "motivo": ["sin contraparte en la fuente derecha"],
    })
    return GenericMatchResult(
        profile_id="mini_test",
        profile_version=1,
        parameters={},
        matched=matched,
        solo_left=solo_left,
        solo_right=solo_right,
        no_cruzados=no_cruzados,
        kpis={"cumplimiento_global_pct": 99.17, "plan_total": 600.0, "filas": 3},
    )


def _pre_corte_profile() -> MatchProfile:
    raw = (PROFILES_DIR / "pre_corte_v1.json").read_text(encoding="utf-8")
    return MatchProfile.from_json(raw)


def _pre_corte_result() -> GenericMatchResult:
    """Resultado sintetico con las columnas que produce el profile real."""
    matched = pd.DataFrame({
        "_k0": [30018, 30049],
        "material_left": [30018, 30049],
        "material_right": [30018, 30049],
        "referencia": ["HUEVO AA X30", "HUEVO AAA X30"],
        "notificado_unidades": [1000.0, 2000.0],
        "producir_unidades": [1100.0, 2100.0],
        "real_unidades_flash": [980.0, 2050.0],
        "delta_unidades": [-20.0, 50.0],
        "cumplimiento_pct": [98.0, 102.5],
    })
    no_cruzados = pd.DataFrame({
        "origen": ["pre_corte", "flash"],
        "key": ["30055", "30077"],
        "motivo": [
            "material notificado sin venta en FLASH",
            "material vendido sin notificacion previa",
        ],
    })
    return GenericMatchResult(
        profile_id="pre_corte_v1",
        profile_version=1,
        parameters={"fecha_produccion": "2026-02-14"},
        matched=matched,
        solo_left=pd.DataFrame({"_k0": [30055], "notificado_unidades": [500.0]}),
        solo_right=pd.DataFrame({"_k0": [30077], "real_unidades_flash": [300.0]}),
        no_cruzados=no_cruzados,
        kpis={
            "cumplimiento_global_pct": 101.0,
            "plan_total_unidades": 3000.0,
            "real_total_unidades": 3030.0,
            "materiales_cruzados": 2,
        },
    )


# ---------------------------------------------------------------------------
# Estructura del archivo
# ---------------------------------------------------------------------------

def test_render_excel_genera_hojas_del_spec(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "mini.xlsx")
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == ["Portada", "Resumen", "Detalle", "No_Cruzados"]


def test_render_excel_profile_sin_report_excel_falla(tmp_path):
    profile = _mini_profile()
    profile = profile.model_copy(update={"report": None})
    with pytest.raises(ValueError, match="report.excel"):
        render_excel(profile, _mini_result(), tmp_path / "x.xlsx")


def test_render_excel_pre_corte_v1_hojas_del_profile_real(tmp_path):
    dest = render_excel(
        _pre_corte_profile(), _pre_corte_result(), tmp_path / "pc.xlsx"
    )
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == ["Portada", "Resumen", "Detalle_Material", "No_Cruzados"]


# ---------------------------------------------------------------------------
# Portada
# ---------------------------------------------------------------------------

def test_portada_tiene_tabla_kpi_con_cumplimiento(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "p.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]

    assert "sintetico" in str(ws["A4"].value), "titulo = profile.descripcion"

    kpi_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Indicador":
            kpi_row = r
            break
    assert kpi_row is not None, "la portada debe tener el header 'Indicador'"
    labels = {
        str(ws.cell(row=r, column=1).value)
        for r in range(kpi_row + 1, ws.max_row + 1)
    }
    assert any("Cumplimiento" in l for l in labels)
    assert any("Plan total" in l for l in labels)


def test_portada_kpi_con_semaforo_del_profile(tmp_path):
    """cumplimiento_global_pct=99.17 esta en rango verde (95-105)."""
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "p.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").startswith("Cumplimiento"):
            cell = ws.cell(row=r, column=2)
            assert cell.value == pytest.approx(0.9917), "pct como fraccion"
            assert cell.style == "st_kpi_value_pct_good"
            return
    pytest.fail("no se encontro la fila del KPI de cumplimiento en la portada")


def test_portada_kpi_semaforo_rojo_si_fuera_de_rango(tmp_path):
    result = _mini_result()
    result.kpis["cumplimiento_global_pct"] = 60.0
    dest = render_excel(_mini_profile(), result, tmp_path / "p.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").startswith("Cumplimiento"):
            assert ws.cell(row=r, column=2).style == "st_kpi_value_pct_bad"
            return
    pytest.fail("no se encontro la fila del KPI de cumplimiento")


def test_kpi_none_se_omite_sin_explotar(tmp_path):
    result = _mini_result()
    result.kpis["cumplimiento_global_pct"] = None
    dest = render_excel(_mini_profile(), result, tmp_path / "p.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    labels = {
        str(ws.cell(row=r, column=1).value)
        for r in range(1, ws.max_row + 1)
    }
    assert not any("Cumplimiento global" in l for l in labels)
    assert any("Plan total" in l for l in labels)


# ---------------------------------------------------------------------------
# Hojas tabla
# ---------------------------------------------------------------------------

def test_hojas_tabla_son_excel_tables_nativas(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    for name in ("Detalle", "No_Cruzados"):
        assert len(wb[name].tables) >= 1, f"'{name}' debe tener una Excel Table"


def test_key_interna_renombrada_a_join_key(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Detalle"]
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    assert "Material" in headers, f"_k0 debe renombrarse a 'material': {headers}"
    assert not any(str(h).startswith("_k") for h in headers)


def test_spec_columns_filtra_y_ordena_ignorando_inexistentes(tmp_path):
    sheets = [
        {
            "name": "Detalle",
            "kind": "tabla",
            "source": "matched",
            "columns": ["real", "material", "no_existe", "plan"],
        },
    ]
    dest = render_excel(_mini_profile(sheets), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Detalle"]
    headers = [
        ws.cell(row=4, column=c).value
        for c in range(1, 4)
    ]
    assert headers == ["Real", "Material", "Plan"]


def test_group_by_agrupa_visualmente(tmp_path):
    sheets = [
        {
            "name": "Detalle",
            "kind": "tabla",
            "source": "matched",
            "group_by": "grupo",
            "columns": ["grupo", "material", "plan"],
        },
    ]
    dest = render_excel(_mini_profile(sheets), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Detalle"]
    # grupo A ocupa filas 5-6: el valor solo debe verse en la fila 5.
    assert ws.cell(row=5, column=1).value == "A"
    assert ws.cell(row=6, column=1).value is None
    assert ws.cell(row=7, column=1).value == "B"


def test_group_by_columna_inexistente_no_falla(tmp_path):
    sheets = [
        {"name": "Detalle", "kind": "tabla", "source": "matched",
         "group_by": "no_existe"},
    ]
    dest = render_excel(_mini_profile(sheets), _mini_result(), tmp_path / "t.xlsx")
    assert openpyxl.load_workbook(dest)["Detalle"].max_row >= 5


def test_pct_guardado_como_fraccion_con_semaforo(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Detalle"]
    headers = {
        ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 1)
    }
    col = headers["Cumplimiento pct"]
    cell = ws.cell(row=5, column=col)
    assert cell.number_format == "0.00%"
    assert cell.value == pytest.approx(0.95), "95.0 del motor -> 0.95 en Excel"
    # Semaforo (formato condicional) presente en la hoja.
    assert list(ws.conditional_formatting._cf_rules), (
        "columna de cumplimiento pct debe llevar semaforo"
    )


def test_solo_left_y_solo_right_como_fuentes(tmp_path):
    sheets = [
        {"name": "Solo_Plan", "kind": "tabla", "source": "solo_left"},
        {"name": "Solo_Real", "kind": "tabla", "source": "solo_right"},
    ]
    dest = render_excel(_mini_profile(sheets), _mini_result(), tmp_path / "t.xlsx")
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == ["Solo_Plan", "Solo_Real"]
    assert len(wb["Solo_Plan"].tables) == 1
    # solo_right esta vacio: la tabla existe igual (0 filas de datos).
    assert len(wb["Solo_Real"].tables) == 1


# ---------------------------------------------------------------------------
# Contrato extendido: breakdowns + nivel de servicio + data_model
# ---------------------------------------------------------------------------

def _profile_extendido() -> MatchProfile:
    """Mini-profile con breakdown, service_level y data_model declarados."""
    base = _mini_profile()
    raw = base.model_dump()
    raw["service_level"] = {
        "plan_column": "plan",
        "real_column": "real",
        "pedido_key": "material",
    }
    raw["breakdowns"] = [
        {
            "id": "por_grupo",
            "label": "Cumplimiento por grupo",
            "dimensions": ["grupo"],
            "metrics": [
                {"id": "plan_total", "op": "sum", "column": "plan"},
                {
                    "id": "cumplimiento_pct",
                    "op": "ratio_pct_of_sums",
                    "numerator": "real",
                    "denominator": "plan",
                },
            ],
        }
    ]
    raw["data_model"] = {
        "fact_name": "FactMini",
        "dimensions": [{"name": "DimGrupo", "key": "grupo"}],
    }
    raw["report"]["excel"]["sheets"] = [
        {"name": "Portada", "kind": "portada"},
        {"name": "Por_Grupo", "kind": "breakdown", "breakdown_id": "por_grupo"},
        {"name": "Detalle", "kind": "tabla", "source": "matched"},
    ]
    return MatchProfile.model_validate(raw)


def _result_extendido() -> GenericMatchResult:
    result = _mini_result()
    result.breakdowns = {
        "por_grupo": pd.DataFrame({
            "grupo": ["A", "B"],
            "plan_total": [300.0, 300.0],
            "cumplimiento_pct": [101.67, 96.67],
        })
    }
    result.service_level = {
        "nivel": "linea",
        "total_lineas_pedido": 4,
        "total_unidades_plan": 650.0,
        "clases": {
            "completo": {
                "lineas": 2, "unidades_plan": 300.0, "unidades_real": 305.0,
                "pct_lineas": 50.0, "pct_unidades_plan": 46.15,
            },
            "parcial": {
                "lineas": 1, "unidades_plan": 300.0, "unidades_real": 290.0,
                "pct_lineas": 25.0, "pct_unidades_plan": 46.15,
            },
            "no_entregado": {
                "lineas": 1, "unidades_plan": 50.0, "unidades_real": 0.0,
                "pct_lineas": 25.0, "pct_unidades_plan": 7.69,
            },
            "sin_pedido": {
                "lineas": 0, "unidades_plan": 0.0, "unidades_real": 0.0,
            },
        },
        "pedidos": {
            "total": 4,
            "clases": {
                "completo": {"pedidos": 2, "pct": 50.0},
                "parcial": {"pedidos": 1, "pct": 25.0},
                "no_entregado": {"pedidos": 1, "pct": 25.0},
            },
        },
    }
    result.kpis["service_level"] = result.service_level
    return result


def test_hoja_breakdown_presente_con_datos(tmp_path):
    dest = render_excel(_profile_extendido(), _result_extendido(), tmp_path / "b.xlsx")
    wb = openpyxl.load_workbook(dest)
    assert "Por_Grupo" in wb.sheetnames
    ws = wb["Por_Grupo"]
    assert len(ws.tables) == 1
    headers = [ws.cell(row=4, column=c).value for c in range(1, 4)]
    assert headers == ["Grupo", "Plan total", "Cumplimiento pct"]
    assert ws.cell(row=5, column=1).value == "A"
    # pct del breakdown como fraccion + semaforo condicional
    assert ws.cell(row=5, column=3).value == pytest.approx(1.0167)
    assert list(ws.conditional_formatting._cf_rules)


def test_hoja_breakdown_sin_datos_en_result_falla_explicito(tmp_path):
    result = _result_extendido()
    result.breakdowns = {}
    with pytest.raises(ValueError, match="por_grupo"):
        render_excel(_profile_extendido(), result, tmp_path / "b.xlsx")


def test_portada_tiene_tabla_nivel_de_servicio(tmp_path):
    dest = render_excel(_profile_extendido(), _result_extendido(), tmp_path / "s.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    textos = {
        str(ws.cell(row=r, column=1).value)
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
    }
    assert "Nivel de servicio" in textos
    for clase in ("Completo", "Parcial", "No entregado", "Sin pedido"):
        assert clase in textos, f"falta la clase '{clase}' en la portada"
    assert "Pedidos" in textos
    assert "Total" in textos
    # ambas tablas son Excel Tables nativas
    assert "NivelServicio" in ws.tables
    assert "PedidosNivelServicio" in ws.tables


def test_portada_nivel_servicio_columnas_y_fracciones(tmp_path):
    dest = render_excel(_profile_extendido(), _result_extendido(), tmp_path / "s.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Clase":
            header_row = r
            break
    assert header_row is not None
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 7)]
    assert headers == [
        "Clase", "Lineas", "Unidades plan", "Unidades real",
        "% lineas", "% unidades",
    ]
    fila_completo = header_row + 1
    assert ws.cell(row=fila_completo, column=1).value == "Completo"
    assert ws.cell(row=fila_completo, column=5).value == pytest.approx(0.50)
    assert ws.cell(row=fila_completo, column=5).number_format == "0.00%"


def test_data_model_hojas_como_listobjects_con_nombre_exacto(tmp_path):
    dest = render_excel(_profile_extendido(), _result_extendido(), tmp_path / "d.xlsx")
    wb = openpyxl.load_workbook(dest)
    # las hojas del modelo van al final, despues de las del spec
    assert wb.sheetnames[:3] == ["Portada", "Por_Grupo", "Detalle"]
    assert set(wb.sheetnames[3:]) == {"DimGrupo", "FactMini"}

    for tabla in ("FactMini", "DimGrupo"):
        ws = wb[tabla]
        assert tabla in ws.tables, (
            f"el ListObject de '{tabla}' debe llamarse exactamente '{tabla}' "
            f"(Power Query importa por nombre); encontrados: {list(ws.tables)}"
        )
        # sin decoracion: el header es la fila 1
        assert ws.cell(row=1, column=1).value is not None


def test_data_model_fact_completo_sin_truncar(tmp_path):
    profile = _profile_extendido()
    result = _result_extendido()
    dest = render_excel(profile, result, tmp_path / "d.xlsx")
    wb = openpyxl.load_workbook(dest)
    ws = wb["FactMini"]
    total = len(result.matched) + len(result.solo_left) + len(result.solo_right)
    # fila 1 = headers, luego una fila por linea del outer join
    assert ws.max_row == 1 + total
    headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
    assert "fact_id" in headers
    assert "estado_cruce" in headers
    assert "dim_grupo_id" in headers


def test_profile_sin_data_model_no_agrega_hojas_extra(tmp_path):
    dest = render_excel(_mini_profile(), _mini_result(), tmp_path / "n.xlsx")
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == ["Portada", "Resumen", "Detalle", "No_Cruzados"]


# ---------------------------------------------------------------------------
# End-to-end: borrador CEN vs SAP con fixtures reales de junio
# ---------------------------------------------------------------------------

def test_cen_vs_sap_end_to_end_excel(tmp_path):
    raw = (PROFILES_DIR / "cen_vs_sap_v1_borrador.json").read_text(encoding="utf-8")
    profile = MatchProfile.from_json(raw)
    result = run_profile(profile, left_path=CEN_JUNIO, right_path=SAP_MUESTRA)

    dest = render_excel(profile, result, tmp_path / "cen.xlsx")
    wb = openpyxl.load_workbook(dest)

    hojas_spec = [s.name for s in profile.report.excel.sheets]
    assert wb.sheetnames[:len(hojas_spec)] == hojas_spec
    assert set(wb.sheetnames[len(hojas_spec):]) == {
        "DimCliente", "DimMaterial", "DimDistrito", "FactNivelServicio",
    }

    # breakdown con datos reales
    ws_mat = wb["Por_Material"]
    assert len(ws_mat.tables) == 1
    assert ws_mat.max_row > 5, "Por_Material debe traer filas del fixture"

    # data model: ListObject con nombre exacto y todas las filas del cruce
    ws_fact = wb["FactNivelServicio"]
    assert "FactNivelServicio" in ws_fact.tables
    total = sum(result.summary()[k] for k in ("matched", "solo_left", "solo_right"))
    assert ws_fact.max_row == 1 + total

    # portada con nivel de servicio real
    ws_p = wb["Portada"]
    textos = {
        str(ws_p.cell(row=r, column=1).value)
        for r in range(1, ws_p.max_row + 1)
        if ws_p.cell(row=r, column=1).value is not None
    }
    assert "Nivel de servicio" in textos
    assert "Pedidos" in textos


# ---------------------------------------------------------------------------
# Puritanismo: la paleta vive solo en excel_style.py
# ---------------------------------------------------------------------------

def test_render_excel_no_importa_styles_directamente():
    src = (
        Path(__file__).parent.parent / "app" / "platform" / "render_excel.py"
    ).read_text(encoding="utf-8")
    prohibidas = ("Font(", "PatternFill(", "Border(", "Alignment(", "Side(")
    for pat in prohibidas:
        assert pat not in src, (
            f"'{pat}' encontrado en render_excel.py. Toda la estilizacion "
            f"debe delegarse en app/core/excel_style.py"
        )
    assert "from openpyxl.styles" not in src, (
        "render_excel.py no debe importar de openpyxl.styles"
    )
