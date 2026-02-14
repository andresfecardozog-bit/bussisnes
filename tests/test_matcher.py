"""Tests para app.core.matcher (cruce por MATERIAL SAP)."""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.core.aggregator import aggregate_flash
from app.core.loaders import load_flash, load_pre_corte
from app.core.matcher import match_by_material, run_batch_pipeline, run_full_pipeline

FIXTURES = Path(__file__).parent / "fixtures"
PRE_CORTE_XLSX = FIXTURES / "PRE_CORTE_muestra.xlsx"
FLASH_MINI = FIXTURES / "FLASH_muestra.csv"
FECHA_PROD = date(2026, 2, 14)

MATERIALES_FLASH_EN_FIXTURE = {30040, 30046, 30018, 40789}


def _make_variant(tmp_path: Path, src: Path, tag: str) -> Path:
    dst = tmp_path / f"variant_{tag}.xlsx"
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    if "INVENTARIO FISICO" in wb.sheetnames:
        wb["INVENTARIO FISICO"]["B2"] = f"variant-{tag}"
    else:
        wb.create_sheet(f"variant_{tag}_marker")
    wb.save(dst)
    return dst


@pytest.fixture
def rebuild_pre_corte(isolated_db_with_catalog):
    """Fixture que fuerza recargar el pre_corte para tener el catalogo poblado."""
    return isolated_db_with_catalog


def _resultado(_):
    pre, _pm = load_pre_corte(PRE_CORTE_XLSX)
    flash, _fm = load_flash(FLASH_MINI)
    agg = aggregate_flash(flash, FECHA_PROD)
    return match_by_material(pre, agg, FECHA_PROD), pre


def test_matched_incluye_materiales_del_flash(rebuild_pre_corte):
    r, _ = _resultado(rebuild_pre_corte)
    matched_set = set(int(m) for m in r.matched["material"])
    assert matched_set == MATERIALES_FLASH_EN_FIXTURE


def test_matched_cumplimiento_material_30040(rebuild_pre_corte):
    r, _ = _resultado(rebuild_pre_corte)
    fila = r.matched[r.matched["material"] == 30040].iloc[0]
    assert fila["notificado_unidades"] == pytest.approx(246_420.0)
    assert fila["real_unidades_flash"] == pytest.approx(246_000.0)
    assert fila["delta_unidades"] == pytest.approx(-420.0)
    assert fila["cumplimiento_pct"] == pytest.approx(99.83, abs=0.05)


def test_solo_pre_corte_incluye_los_que_no_facturaron(rebuild_pre_corte):
    r, _ = _resultado(rebuild_pre_corte)
    solo_pre_set = set(int(m) for m in r.solo_pre_corte["material"])
    assert MATERIALES_FLASH_EN_FIXTURE.isdisjoint(solo_pre_set)
    assert len(r.solo_pre_corte) == 18 - len(MATERIALES_FLASH_EN_FIXTURE)


def test_solo_flash_incluye_material_sin_notificar(rebuild_pre_corte):
    r, _ = _resultado(rebuild_pre_corte)
    assert 99999 in set(int(m) for m in r.solo_flash["material"])


def test_no_cruzados_tiene_motivos_correctos(rebuild_pre_corte):
    r, _ = _resultado(rebuild_pre_corte)
    origenes = set(r.no_cruzados["origen"])
    assert "pre_corte" in origenes
    assert "flash" in origenes


def test_run_full_pipeline_extrae_fecha_de_filename(rebuild_pre_corte, tmp_path):
    dst = tmp_path / "PRE CORTE 13.02.2026.xlsx"
    shutil.copy(PRE_CORTE_XLSX, dst)
    r = run_full_pipeline(dst, FLASH_MINI)
    assert r.fecha_produccion == FECHA_PROD
    assert r.summary()["matched"] == len(MATERIALES_FLASH_EN_FIXTURE)


def test_conservacion_de_filas(rebuild_pre_corte):
    r, pre = _resultado(rebuild_pre_corte)
    assert len(pre) == len(r.matched) + len(r.solo_pre_corte)


def test_batch_pipeline_procesa_multiples_pre_corte(rebuild_pre_corte, tmp_path):
    """N archivos PRE CORTE del mes vs UN flash mensual."""
    a = tmp_path / "PRE CORTE 13.02.2026.xlsx"
    b = tmp_path / "PRE CORTE 14.02.2026.xlsx"
    shutil.copy(PRE_CORTE_XLSX, a)
    _make_variant(tmp_path, PRE_CORTE_XLSX, tag="dia2").rename(b)

    batch = run_batch_pipeline(
        pre_corte_paths=[a, b],
        flash_path=FLASH_MINI,
    )

    assert len(batch["runs"]) == 2
    assert batch["flash_meta"]["tipo"] == "flash"

    dia1 = batch["runs"][0]
    dia2 = batch["runs"][1]
    assert dia1["fecha_produccion"] == FECHA_PROD
    # PRE CORTE 14.02.2026 (sabado) -> 15.02.2026 (domingo, no laboral)
    # -> skip al lunes 16.02.2026 por la regla de la Fase 6.6.
    assert dia2["fecha_produccion"] == date(2026, 2, 16)


def test_batch_pipeline_flash_hash_estable(rebuild_pre_corte, tmp_path):
    a = tmp_path / "PRE CORTE 13.02.2026.xlsx"
    b = tmp_path / "PRE CORTE 14.02.2026.xlsx"
    shutil.copy(PRE_CORTE_XLSX, a)
    _make_variant(tmp_path, PRE_CORTE_XLSX, tag="dia2").rename(b)
    batch = run_batch_pipeline(
        pre_corte_paths=[a, b],
        flash_path=FLASH_MINI,
    )
    flash_hash = batch["flash_meta"]["hash_sha256"]
    assert flash_hash and len(flash_hash) == 64
