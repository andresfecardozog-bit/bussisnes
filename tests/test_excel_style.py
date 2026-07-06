"""Tests unitarios del modulo de estilo `app.core.excel_style`.

Enfoque: verificar que las primitivas (NamedStyle, Excel Tables,
semaforo) se comportan segun contrato. Sin dependencias a la DB ni al
pipeline; solo openpyxl.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl.workbook import Workbook

from app.core import excel_style as xs
from app.core.excel_style import ColumnSpec


# ---------- register_styles ----------

def test_register_styles_agrega_todos_los_nombres():
    wb = Workbook()
    xs.register_styles(wb)
    for name in xs._STYLE_NAMES:
        assert name in wb.named_styles, f"falta NamedStyle: {name}"


def test_register_styles_es_idempotente():
    wb = Workbook()
    xs.register_styles(wb)
    n1 = len(wb.named_styles)
    xs.register_styles(wb)
    xs.register_styles(wb)
    n2 = len(wb.named_styles)
    assert n1 == n2, "registrar dos veces no debe duplicar estilos"


# ---------- write_dataframe_as_table ----------

def test_write_dataframe_as_table_crea_excel_table_nativa():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    df = pd.DataFrame({
        "fecha": pd.to_datetime(["2026-02-14", "2026-02-15"]).date,
        "plan": [100, 200],
        "cumpl": [0.95, 1.20],
    })
    cols = (
        ColumnSpec("Fecha", "fecha", "date"),
        ColumnSpec("Plan", "plan", "int"),
        ColumnSpec("Cumpl", "cumpl", "pct"),
    )
    table = xs.write_dataframe_as_table(ws, df, cols, "TestTable",
                                        start_row=1)
    assert table.displayName == "TestTable"
    assert "TestTable" in ws.tables
    assert ws.tables["TestTable"].ref == table.ref
    assert table.tableStyleInfo.name == xs.DEFAULT_TABLE_STYLE
    assert table.tableStyleInfo.showRowStripes is True


def test_write_dataframe_as_table_valores_como_tipo_nativo():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    df = pd.DataFrame({
        "plan": [100, 200],
        "cumpl": [0.85, 1.05],
    })
    cols = (
        ColumnSpec("Plan", "plan", "int"),
        ColumnSpec("Cumpl", "cumpl", "pct"),
    )
    xs.write_dataframe_as_table(ws, df, cols, "T1", start_row=1)
    assert isinstance(ws["A2"].value, int)
    assert isinstance(ws["B2"].value, float)
    assert ws["B2"].value == 0.85
    assert ws["B2"].number_format == "0.00%"


def test_write_dataframe_as_table_dataframe_vacio_no_explota():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    df = pd.DataFrame(columns=["plan"])
    cols = (ColumnSpec("Plan", "plan", "int"),)
    xs.write_dataframe_as_table(ws, df, cols, "Vacio", start_row=1)
    assert "Vacio" in ws.tables


def test_write_dataframe_as_table_sanitiza_nombre_invalido():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    df = pd.DataFrame({"a": [1]})
    cols = (ColumnSpec("A", "a", "int"),)
    t = xs.write_dataframe_as_table(ws, df, cols, "1nombre invalido", 1)
    assert t.displayName[0].isalpha(), "nombre de Table debe empezar por letra"
    assert " " not in t.displayName
    assert t.displayName in ws.tables


# ---------- apply_traffic_light ----------

def test_apply_traffic_light_tres_reglas_discretas():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    ws["A1"] = 0.95
    ws["A2"] = 0.80
    ws["A3"] = 1.30
    xs.apply_traffic_light(ws, "A", from_row=1, to_row=3)

    rangos = list(ws.conditional_formatting._cf_rules.keys())
    assert len(rangos) == 1, "todas las reglas deben estar en el mismo rango"
    assert "A1:A3" in str(rangos[0].sqref)

    reglas = list(ws.conditional_formatting._cf_rules[rangos[0]])
    assert len(reglas) == 3
    formulas = [r.formula for r in reglas]
    assert ["0.95", "1.05"] in formulas
    assert ["0.85", "1.15"] in formulas
    operators = [r.operator for r in reglas]
    assert "notBetween" in operators
    assert operators.count("between") == 2


def test_apply_traffic_light_rango_vacio_es_noop():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    xs.apply_traffic_light(ws, "A", from_row=5, to_row=4)
    assert not ws.conditional_formatting._cf_rules


# ---------- add_title_block ----------

def test_add_title_block_arma_titulo_y_subtitulo():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    start = xs.add_title_block(ws, "TITULO", "subtitulo", span_cols=5)
    assert ws["A1"].value == "TITULO"
    assert ws["A2"].value == "subtitulo"
    assert start == 4
    assert ws["A1"].style == "st_title"
    assert ws["A2"].style == "st_subtitle"


# ---------- add_total_row ----------

def test_add_total_row_aplica_estilos_de_total():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    cols = (
        ColumnSpec("Fecha", "fecha", "date"),
        ColumnSpec("Plan", "plan", "int"),
        ColumnSpec("Cumpl", "cumpl", "pct"),
    )
    xs.add_total_row(
        ws, cols,
        values={"fecha": None, "plan": 300, "cumpl": 0.95},
        row_num=10,
        label_col_key="fecha",
        label_text="TOTAL",
    )
    assert ws["A10"].value == "TOTAL"
    assert ws["A10"].style == "st_total_label"
    assert ws["B10"].value == 300
    assert ws["B10"].style == "st_total_int"
    assert ws["C10"].value == 0.95
    assert ws["C10"].style == "st_total_pct"


# ---------- write_kpi_table ----------

@pytest.mark.parametrize("valor,expected_suffix", [
    (0.95, "good"),
    (1.00, "good"),
    (1.05, "good"),
    (0.90, "warn"),
    (1.10, "warn"),
    (0.70, "bad"),
    (1.30, "bad"),
])
def test_write_kpi_table_aplica_semaforo_en_valor_pct(valor, expected_suffix):
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    kpis = (
        xs.KpiRow("Cumplimiento", valor, "pct", semaforo=True),
    )
    xs.write_kpi_table(ws, kpis, start_row=1)
    assert ws["A1"].value == "Indicador"
    assert ws["B1"].value == "Valor"
    assert ws["A1"].style == "st_kpi_header"
    assert ws["A2"].value == "Cumplimiento"
    assert ws["B2"].value == pytest.approx(valor)
    assert ws["B2"].style == f"st_kpi_value_pct_{expected_suffix}"
    assert ws["B2"].number_format == "0.00%"


def test_write_kpi_table_int_usa_estilo_body_int():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    kpis = (
        xs.KpiRow("Plan total", 299416, "int"),
        xs.KpiRow("Real total", 833236, "int"),
    )
    next_row = xs.write_kpi_table(ws, kpis, start_row=5)
    assert ws["B6"].value == 299416
    assert ws["B6"].style == "st_kpi_value_int"
    assert ws["B6"].number_format == "#,##0"
    assert next_row == 5 + len(kpis) + 1


# ---------- insert_logo ----------

def test_insert_logo_devuelve_none_si_no_existe(tmp_path):
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    assert xs.insert_logo(ws, tmp_path / "no-existe.jpg") is None


def test_insert_logo_agrega_imagen_si_existe():
    logo = Path(__file__).resolve().parents[1] / "resources" / "image_720508810_0.jpg"
    if not logo.exists():
        pytest.skip("logo corporativo no disponible")
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    img = xs.insert_logo(ws, logo, anchor="A1", max_height_px=50)
    assert img is not None
    assert img.height <= 50


# ---------- autofit_columns ----------

def test_autofit_columns_respeta_max_width():
    wb = Workbook()
    xs.register_styles(wb)
    ws = wb.active
    ws["A1"] = "hola"
    ws["A2"] = "x" * 500
    xs.autofit_columns(ws, columns=[1], max_width=25)
    assert ws.column_dimensions["A"].width == 25
