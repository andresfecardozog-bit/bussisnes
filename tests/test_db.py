"""Tests para app.core.db (schema, idempotencia, batch, rollback)."""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.core.aggregator import aggregate_flash
from app.core.db import (
    already_loaded,
    create_run,
    get_conn,
    get_or_insert_carga,
    get_run,
    init_db,
    list_cargas,
    list_recent_runs,
    persist_run,
    update_run,
)
from app.core.loaders import load_flash, load_pre_corte
from app.core.matcher import match_by_material

FIXTURES = Path(__file__).parent / "fixtures"
PRE_CORTE_XLSX = FIXTURES / "PRE_CORTE_muestra.xlsx"
FLASH_MINI = FIXTURES / "FLASH_muestra.csv"


def _make_variant(tmp_path: Path, src: Path, tag: str) -> Path:
    """Copia el xlsx y modifica una celda inofensiva para generar hash distinto."""
    dst = tmp_path / f"variant_{tag}.xlsx"
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    if "INVENTARIO FISICO" in wb.sheetnames:
        wb["INVENTARIO FISICO"]["B2"] = f"variant-{tag}"
    else:
        wb.create_sheet(f"variant_{tag}_marker")
    wb.save(dst)
    return dst


def _run_pipeline_and_persist(db_path, pre_path, flash_path, fecha_prod: date):
    pre_df, pre_meta = load_pre_corte(pre_path)
    flash_df, flash_meta = load_flash(flash_path)
    agg = aggregate_flash(flash_df, fecha_prod)
    result = match_by_material(pre_df, agg, fecha_prod)
    with get_conn(db_path) as conn:
        return persist_run(
            conn,
            pre_corte_meta=pre_meta,
            pre_corte_df=pre_df,
            flash_meta=flash_meta,
            flash_agregado_df=agg,
            match_result=result,
            fecha_archivo=fecha_prod,
        )


def test_init_db_crea_todas_las_tablas(isolated_db_with_catalog):
    with get_conn(isolated_db_with_catalog) as conn:
        rows = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
    names = [r["name"] for r in rows if not r["name"].startswith("sqlite_")]
    assert set(names) == {
        "cargas",
        "cruce",
        "flash_agregado",
        "no_cruzados",
        "pre_corte",
        "runs",
        "sku_catalog",
        "batches",
        "batch_pre_cortes",
    }


def test_init_db_es_idempotente(isolated_db_with_catalog):
    init_db(isolated_db_with_catalog)
    init_db(isolated_db_with_catalog)


def test_already_loaded_false_para_hash_nuevo(isolated_db_with_catalog):
    with get_conn(isolated_db_with_catalog) as conn:
        assert already_loaded(conn, "pre_corte", "hash_inexistente") is None


def test_get_or_insert_carga_no_duplica(isolated_db_with_catalog):
    _, pre_meta = load_pre_corte(PRE_CORTE_XLSX)
    with get_conn(isolated_db_with_catalog) as conn:
        id1, es_nueva1 = get_or_insert_carga(
            conn, pre_meta,
            fecha_archivo=date(2026, 2, 13), fecha_produccion=date(2026, 2, 14),
        )
        assert es_nueva1 is True
        id2, es_nueva2 = get_or_insert_carga(
            conn, pre_meta,
            fecha_archivo=date(2026, 2, 13), fecha_produccion=date(2026, 2, 14),
        )
        assert es_nueva2 is False
        assert id1 == id2


def test_persist_run_escribe_cruce_y_no_cruzados(isolated_db_with_catalog):
    summary = _run_pipeline_and_persist(
        isolated_db_with_catalog, PRE_CORTE_XLSX, FLASH_MINI, date(2026, 2, 14),
    )
    assert summary["cruce_filas_insertadas"] >= 3
    assert summary["no_cruzados_filas_insertadas"] >= 1
    assert summary["pre_es_nueva"] is True
    assert summary["flash_es_nueva"] is True

    with get_conn(isolated_db_with_catalog) as conn:
        cnt = conn.execute("SELECT COUNT(*) AS n FROM cruce").fetchone()["n"]
        assert cnt >= 3


def test_persist_run_es_idempotente_por_par(isolated_db_with_catalog):
    """Reprocesar el mismo par (pre_corte, flash) NO duplica filas en cruce."""
    _run_pipeline_and_persist(
        isolated_db_with_catalog, PRE_CORTE_XLSX, FLASH_MINI, date(2026, 2, 14),
    )
    second = _run_pipeline_and_persist(
        isolated_db_with_catalog, PRE_CORTE_XLSX, FLASH_MINI, date(2026, 2, 14),
    )

    assert second["pre_es_nueva"] is False
    assert second["flash_es_nueva"] is False
    assert second["cruce_filas_insertadas"] == 0
    assert second["no_cruzados_filas_insertadas"] == 0

    with get_conn(isolated_db_with_catalog) as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM cargas").fetchone()["n"] == 2


def test_batch_pipeline_dos_pre_cortes_vs_un_flash(isolated_db_with_catalog, tmp_path):
    """Escenario real: dos PRE CORTE de dias distintos contra el mismo FLASH."""
    variante = _make_variant(tmp_path, PRE_CORTE_XLSX, tag="dia2")

    _run_pipeline_and_persist(
        isolated_db_with_catalog, PRE_CORTE_XLSX, FLASH_MINI, date(2026, 2, 14),
    )
    _run_pipeline_and_persist(
        isolated_db_with_catalog, variante, FLASH_MINI, date(2026, 2, 15),
    )

    with get_conn(isolated_db_with_catalog) as conn:
        cargas = list_cargas(conn)
        tipos = [c["tipo"] for c in cargas]
        assert tipos.count("pre_corte") == 2
        assert tipos.count("flash") == 1

        fechas = conn.execute(
            "SELECT DISTINCT fecha_produccion FROM cruce ORDER BY fecha_produccion"
        ).fetchall()
        fechas_iso = [f["fecha_produccion"] for f in fechas]
        assert "2026-02-14" in fechas_iso


def test_persist_run_rollback_si_falla_en_medio(isolated_db_with_catalog, monkeypatch):
    from app.core import db as db_mod

    original = db_mod.persist_cruce

    def broken(*a, **kw):
        raise RuntimeError("simulacion de fallo en persist_cruce")

    monkeypatch.setattr(db_mod, "persist_cruce", broken)
    with pytest.raises(RuntimeError):
        _run_pipeline_and_persist(
            isolated_db_with_catalog, PRE_CORTE_XLSX, FLASH_MINI, date(2026, 2, 14),
        )
    monkeypatch.setattr(db_mod, "persist_cruce", original)

    with get_conn(isolated_db_with_catalog) as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM cruce").fetchone()["n"] == 0
        assert conn.execute("SELECT COUNT(*) AS n FROM cargas").fetchone()["n"] == 0


def test_runs_ciclo_completo(isolated_db_with_catalog):
    with get_conn(isolated_db_with_catalog) as conn:
        run_id = create_run(
            conn,
            fecha_produccion=date(2026, 3, 2),
            current_step="load_pre_corte",
        )
        assert len(run_id) == 32

        update_run(conn, run_id, current_step="matching", status="running")
        update_run(
            conn,
            run_id,
            status="completed",
            current_step="done",
            summary={"matched": 3, "solo_pre": 1},
            ended=True,
        )
        r = get_run(conn, run_id)
        assert r["status"] == "completed"
        assert r["summary"] == {"matched": 3, "solo_pre": 1}
        assert r["ended_at"] is not None


def test_list_recent_runs_ordena_desc(isolated_db_with_catalog):
    with get_conn(isolated_db_with_catalog) as conn:
        r1 = create_run(conn, current_step="a")
        r2 = create_run(conn, current_step="b")
        recent = list_recent_runs(conn, limit=5)
        ids = [r["id"] for r in recent]
        assert ids[0] == r2
        assert ids[1] == r1
