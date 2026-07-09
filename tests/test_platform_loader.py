"""Tests del ConfigurableLoader contra las trampas reales del caso CEN/SAP."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from app.platform.loader import detect_real_format, load_source, registered_loader_names
from app.platform.profile import RegisteredLoaderSpec, TabularLoaderSpec

FIXTURES = Path(__file__).parent / "fixtures"
CEN_XLSX = FIXTURES / "cen" / "Acumulado CEN P7 2026.xlsx"
SAP_MUESTRA = FIXTURES / "cen" / "sap_junio_muestra.xlsx"


def _cen_spec(**overrides) -> TabularLoaderSpec:
    base = dict(
        sheet=None,
        header_row=1,
        columns=[
            {"name": "numero_orden", "source": "Numero de la Orden de compra", "dtype": "str"},
            {"name": "codigo_item", "source": "Codigo item proveedor", "dtype": "str", "required": False},
            {"name": "cantidad_total", "source": "Cantidad Total", "dtype": "float_clean"},
        ],
        drop_rows_where_null=["numero_orden"],
    )
    base.update(overrides)
    return TabularLoaderSpec(**base)


def _sap_spec() -> TabularLoaderSpec:
    return TabularLoaderSpec(
        sheet=None,
        header_row=None,
        columns=[
            {"name": "numero_orden", "source": 56, "dtype": "str"},
            {"name": "codigo_item", "source": 40, "dtype": "str", "required": False},
            {"name": "cantidad", "source": 42, "dtype": "float_clean"},
        ],
        drop_rows_where_null=["numero_orden"],
    )


def test_cen_carga_con_autodeteccion_de_hoja():
    df, meta = load_source(CEN_XLSX, _cen_spec())
    assert len(df) > 1000
    assert meta["formato_real"] == "xlsx"
    assert df["numero_orden"].notna().all()
    # grano: fila = linea de producto -> ordenes repetidas
    assert df["numero_orden"].duplicated().any()


def test_cen_hoja_pivote_no_confunde_la_autodeteccion(tmp_path):
    """Simula el CEN P2: una hoja pivote sin los headers + la hoja real."""
    wb = openpyxl.Workbook()
    pivote = wb.active
    pivote.title = "Hoja3"
    pivote.append(["Razon Social Empresa Compradora", "(Todas)"])
    pivote.append(["algo", 1])
    data = wb.create_sheet("Hoja1")
    data.append(["Numero de la Orden de compra", "Codigo item proveedor", "Cantidad Total"])
    data.append(["004-001", "30049", 10])
    path = tmp_path / "cen_p2_simulado.xlsx"
    wb.save(path)

    df, _ = load_source(path, _cen_spec())
    assert len(df) == 1
    assert df.iloc[0]["numero_orden"] == "004-001"


def test_sap_extension_mentirosa_y_sin_headers(tmp_path):
    """El SAP real llega como .XLS pero es xlsx: el loader mira la firma."""
    disfrazado = tmp_path / "junio.XLS"
    disfrazado.write_bytes(SAP_MUESTRA.read_bytes())
    assert detect_real_format(disfrazado) == "xlsx"

    df, meta = load_source(disfrazado, _sap_spec())
    # el fixture tiene 400 filas con numero de pedido en col 56; las 200
    # sin pedido se descartan por drop_rows_where_null pero contabilizadas.
    # Ojo (hallazgo con datos reales): col 56 mezcla ordenes CEN
    # ('004-0019349') con pedidos de otros origenes ('261357', ...); el
    # filtro fino es tema de la entrevista de Fase 3.
    assert len(df) == 400
    assert meta["descartes"]["null_numero_orden"] == 200
    assert df["numero_orden"].notna().all()


def test_filas_fantasma_recortadas(tmp_path):
    """Simula CEN P4: dimension declarada mucho mayor que las filas reales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Numero de la Orden de compra", "Codigo item proveedor", "Cantidad Total"])
    ws.append(["004-002", "30018", 5])
    # celda lejana vacia infla la dimension declarada
    ws.cell(row=50000, column=1, value=None)
    path = tmp_path / "fantasma.xlsx"
    wb.save(path)

    df, meta = load_source(path, _cen_spec())
    assert len(df) == 1


def test_contabilidad_de_descartes():
    df, meta = load_source(SAP_MUESTRA, _sap_spec())
    total = meta["num_filas_procesadas"] + sum(meta["descartes"].values())
    assert total == meta["num_filas_original"]


def test_columna_requerida_faltante_falla_explicito(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Otra Columna", "Cantidad Total"])
    ws.append(["x", 1])
    path = tmp_path / "sin_columna.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="obligatorias|headers requeridos"):
        load_source(path, _cen_spec(sheet=0))


def test_loaders_registrados_existen():
    names = registered_loader_names()
    assert "pre_corte_resumen" in names
    assert "flash_sap" in names


def test_loader_registrado_inexistente():
    with pytest.raises(ValueError, match="no existe"):
        load_source(SAP_MUESTRA, RegisteredLoaderSpec(name="no_existo"))
