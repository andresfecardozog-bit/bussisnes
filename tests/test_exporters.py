"""Tests del export final de la Fase 6.

Cubre:
- Estructura del `.xlsx`: 5 hojas con nombres exactos.
- Cada hoja de datos es una Excel Table nativa.
- Cumplimiento se guarda como fraccion (0.85), no como 85.
- Semaforo aplicado a la columna `cumplimiento_pct`.
- Totales del Resumen coinciden con la suma de `pre_corte`.
- Las 6 categorias fijas (A/AA/AAA/AAAA/B/C) aparecen aunque no haya
  actividad en alguna.
- Contrato de "paleta centralizada": `exporters.py` no importa `Font`,
  `PatternFill`, `Border`, `Alignment`, `Side` directamente.
- Endpoint `GET /kpis/excel` valida el rango y devuelve el archivo.
"""
from __future__ import annotations

import re
import shutil
import sqlite3
from datetime import date
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.core.aggregator import aggregate_flash
from app.core.date_extractor import extract_file_date, extract_production_date
from app.core.db import get_conn, persist_run
from app.core.exporters import (
    _resumen_por_categoria,
    _resumen_por_semana,
    _resumen_totales,
    export_batch_completo,
    export_cumplimiento_diario,
    export_cumplimiento_xlsx,
    suggested_export_filename,
)
from app.core.loaders import load_flash, load_pre_corte
from app.core.matcher import match_by_material

FIXTURES = Path(__file__).parent / "fixtures"
PRE_CORTE_XLSX = FIXTURES / "PRE_CORTE_muestra.xlsx"
FLASH_CSV = FIXTURES / "FLASH_muestra.csv"
HOMOLOG_XLSX = FIXTURES / "homologacion.xlsx"
FECHA_PROD = date(2026, 2, 14)


# ---------------------------------------------------------------------------
# Fixture: DB aislada + pipeline corrido = cruce y no_cruzados poblados
# ---------------------------------------------------------------------------
@pytest.fixture
def db_con_cruce(isolated_db_with_catalog, tmp_path) -> Path:
    """Corre el pipeline sobre el fixture y persiste. Retorna el path a la DB."""
    db_path = isolated_db_with_catalog
    pre_named = tmp_path / "PRE CORTE 13.02.2026.xlsx"
    shutil.copy(PRE_CORTE_XLSX, pre_named)

    pre_df, pre_meta = load_pre_corte(pre_named)
    flash_df, flash_meta = load_flash(FLASH_CSV)
    fecha_arch = extract_file_date(pre_named)
    fecha_prod = extract_production_date(pre_named)
    assert fecha_prod == FECHA_PROD
    agg = aggregate_flash(flash_df, fecha_prod)
    result = match_by_material(pre_df, agg, fecha_prod)
    with get_conn(db_path) as conn:
        persist_run(
            conn,
            pre_corte_meta=pre_meta,
            pre_corte_df=pre_df,
            flash_meta=flash_meta,
            flash_agregado_df=agg,
            match_result=result,
            fecha_archivo=fecha_arch,
        )
    return db_path


@pytest.fixture
def db_con_cruce_multi_fecha(db_con_cruce) -> Path:
    """Extiende `db_con_cruce` con cruce para 4 fechas laborales adicionales
    (09, 10, 11, 12 feb 2026), cada una con cumplimiento distinto para
    triggerar los 3 estados del semaforo.

    Usa cargas fake nuevas + FKs desactivadas para no depender de crear
    archivos reales (solo necesitamos filas en `cruce` para probar el export).
    """
    db_path = db_con_cruce
    fechas_extra = [
        (date(2026, 2, 9),  1.00),   # verde
        (date(2026, 2, 10), 0.92),   # amarillo
        (date(2026, 2, 11), 0.78),   # rojo bajo
        (date(2026, 2, 12), 1.10),   # amarillo alto
    ]
    with get_conn(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = OFF;")
        base = conn.execute(
            "SELECT material, referencia, nomb_material_flash, "
            "notificado_unidades, producir_unidades "
            "FROM cruce WHERE fecha_produccion = ?",
            (FECHA_PROD.isoformat(),),
        ).fetchall()
        assert base, "fixture base debe tener cruce para 14/02/2026"
        for i, (f, factor) in enumerate(fechas_extra):
            fake_pre_id = 9000 + i
            fake_flash_id = 9500 + i
            for r in base:
                notif = float(r["notificado_unidades"])
                real = notif * factor
                delta = real - notif
                cumpl = factor * 100.0
                conn.execute(
                    """
                    INSERT INTO cruce (
                        pre_corte_carga_id, flash_carga_id, fecha_produccion,
                        material, referencia, nomb_material_flash,
                        notificado_unidades, producir_unidades,
                        real_unidades_flash, delta_unidades, cumplimiento_pct,
                        match_bool
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                    """,
                    (
                        fake_pre_id, fake_flash_id, f.isoformat(),
                        r["material"], r["referencia"], r["nomb_material_flash"],
                        notif, r["producir_unidades"], real, delta, cumpl,
                    ),
                )
            conn.execute(
                """INSERT INTO no_cruzados
                   (pre_corte_carga_id, flash_carga_id, fecha_produccion,
                    origen, material, referencia_o_nombre, valor, motivo)
                   VALUES (?, ?, ?, 'flash', 99999, 'FUGA DEMO', 500,
                           'material sin plan')""",
                (fake_pre_id, fake_flash_id, f.isoformat()),
            )
        conn.commit()
        conn.execute("PRAGMA foreign_keys = ON;")
    return db_path


# ---------------------------------------------------------------------------
# Helpers puros
# ---------------------------------------------------------------------------
def test_suggested_export_filename_dia_unico():
    """Rango de 1 dia usa el naming corto tipo 'cumplimiento_YYYYMMDD.xlsx'."""
    n = suggested_export_filename(date(2026, 2, 14), date(2026, 2, 14))
    assert n == "cumplimiento_20260214.xlsx"


def test_suggested_export_filename_rango_es_consolidado():
    n = suggested_export_filename(date(2026, 2, 14), date(2026, 2, 28))
    assert n == "cumplimiento_consolidado_20260214_20260228.xlsx"


def test_resumen_totales_cumplimiento_es_fraccion():
    import pandas as pd
    cruce = pd.DataFrame({
        "fecha_produccion": [date(2026, 2, 14)] * 2,
        "plan": [1000, 2000],
        "real": [1050, 1900],
    })
    r = _resumen_totales(cruce)
    assert list(r.columns) == [
        "fecha_produccion", "plan_total", "real_total",
        "delta_total", "cumplimiento_pct",
    ]
    assert r["plan_total"].iloc[0] == 3000
    assert r["real_total"].iloc[0] == 2950
    assert r["cumplimiento_pct"].iloc[0] == pytest.approx(2950 / 3000)


def test_resumen_por_categoria_rellena_categorias_faltantes():
    import pandas as pd
    cruce = pd.DataFrame({
        "fecha_produccion": [date(2026, 2, 14)] * 2,
        "tipo": ["A", "B"],
        "plan": [100, 200],
        "real": [90, 250],
    })
    r = _resumen_por_categoria(cruce)
    tipos = set(r["tipo"].unique())
    for t in ("A", "AA", "AAA", "AAAA", "B", "C"):
        assert t in tipos, f"tipo {t} deberia estar aunque tenga 0"


# ---------------------------------------------------------------------------
# Export end-to-end
# ---------------------------------------------------------------------------
def test_export_genera_xlsx_con_5_hojas(db_con_cruce, tmp_path):
    dest = tmp_path / "cumplimiento_test.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == [
        "Portada", "Resumen", "Por_Categoria",
        "Detalle_Material", "No_Cruzados",
    ]


def test_export_cada_hoja_de_datos_es_excel_table(db_con_cruce, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    for name in ("Resumen", "Por_Categoria", "Detalle_Material", "No_Cruzados"):
        ws = wb[name]
        assert len(ws.tables) >= 1, f"'{name}' debe tener una Excel Table"


def test_export_cumplimiento_guardado_como_fraccion(db_con_cruce, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Resumen"]
    fila_datos = ws[5]
    cumpl_cell = fila_datos[4]
    assert cumpl_cell.number_format == "0.00%"
    assert isinstance(cumpl_cell.value, float)
    assert cumpl_cell.value < 20.0, (
        "cumplimiento debe estar guardado como fraccion (0.85), "
        f"no como porcentaje entero: valor={cumpl_cell.value}"
    )


def test_export_semaforo_aplicado_al_cumplimiento(db_con_cruce, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    for name in ("Resumen", "Por_Categoria", "Detalle_Material"):
        ws = wb[name]
        rangos = list(ws.conditional_formatting._cf_rules.keys())
        assert rangos, f"'{name}' debe tener conditional_formatting"
        reglas = list(ws.conditional_formatting._cf_rules[rangos[0]])
        assert len(reglas) == 3, f"'{name}' debe tener 3 reglas de semaforo"


def test_export_no_cruzados_incluye_flash_y_pre_corte(db_con_cruce, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["No_Cruzados"]
    origenes = {
        ws.cell(row=r, column=2).value
        for r in range(5, ws.max_row + 1)
        if ws.cell(row=r, column=2).value not in (None, "")
    }
    assert "flash" in origenes


def test_export_rango_sin_datos_no_falla(isolated_db_with_catalog, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(
        date(2020, 1, 1), date(2020, 1, 2), dest,
        db_path=isolated_db_with_catalog,
    )
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    assert wb.sheetnames == [
        "Portada", "Resumen", "Por_Categoria",
        "Detalle_Material", "No_Cruzados",
    ]


def test_export_categorias_completas_en_por_categoria(db_con_cruce, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Por_Categoria"]
    tipos = {
        ws.cell(row=r, column=2).value
        for r in range(5, ws.max_row + 1)
        if ws.cell(row=r, column=2).value is not None
    }
    for t in ("A", "AA", "AAA", "AAAA", "B", "C"):
        assert t in tipos, f"tipo {t} debe aparecer en Por_Categoria"


def test_export_portada_tiene_tabla_kpi(db_con_cruce, tmp_path):
    """La portada es una tabla relacional label|valor, sin literatura."""
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]

    assert "Cumplimiento" in str(ws["A4"].value)

    kpi_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Indicador":
            kpi_row = r
            break
    assert kpi_row is not None, "la portada debe tener el header 'Indicador'"
    assert ws.cell(row=kpi_row, column=2).value == "Valor"

    labels = {
        str(ws.cell(row=r, column=1).value)
        for r in range(kpi_row + 1, ws.max_row + 1)
    }
    assert any("Cumplimiento" in l for l in labels)
    assert any("Plan total" in l for l in labels)
    assert any("Real total" in l for l in labels)


# ---------------------------------------------------------------------------
# Fase 6.5: agrupacion de fechas + Por_Semana + batch multi-fecha + zip
# ---------------------------------------------------------------------------
def test_resumen_por_semana_agrupa_por_semana_iso():
    import pandas as pd
    cruce = pd.DataFrame({
        "fecha_produccion": [
            date(2026, 2, 9), date(2026, 2, 10),  # semana 7 (S07)
            date(2026, 2, 16),                     # semana 8 (S08)
        ],
        "plan": [1000, 1000, 1000],
        "real": [900, 1000, 1100],
    })
    r = _resumen_por_semana(cruce)
    assert len(r) == 2
    labels = list(r["semana_label"])
    assert labels[0].startswith("S07")
    assert labels[1].startswith("S08")


def test_export_multi_fecha_incluye_hoja_por_semana(db_con_cruce_multi_fecha, tmp_path):
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(date(2026, 2, 9), date(2026, 2, 14), dest,
                             db_path=db_con_cruce_multi_fecha)
    wb = openpyxl.load_workbook(dest)
    assert "Por_Semana" in wb.sheetnames
    # Debe aparecer antes de Por_Categoria en orden de hojas
    idx_semana = wb.sheetnames.index("Por_Semana")
    idx_categoria = wb.sheetnames.index("Por_Categoria")
    assert idx_semana < idx_categoria


def test_export_dia_unico_no_agrega_hoja_por_semana(db_con_cruce, tmp_path):
    """Con 1 sola fecha, la hoja Por_Semana no aporta y se omite."""
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    assert "Por_Semana" not in wb.sheetnames


def test_export_multi_fecha_agrupa_fechas_en_detalle(db_con_cruce_multi_fecha, tmp_path):
    """En Detalle_Material con multiples fechas, la columna fecha muestra
    su valor solo en la primera fila del grupo y queda vacia en las
    siguientes filas del mismo dia. Nunca dos fechas iguales consecutivas
    con ambas visibles.
    """
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(date(2026, 2, 9), date(2026, 2, 14), dest,
                             db_path=db_con_cruce_multi_fecha)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Detalle_Material"]
    valores_fecha = []
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            valores_fecha.append(v)
    # Todas las fechas visibles deben ser unicas (una por dia).
    assert len(valores_fecha) == len(set(valores_fecha)), (
        "Fechas duplicadas visibles indican que la agrupacion no se aplico"
    )
    assert len(valores_fecha) >= 2


def test_export_batch_completo_genera_dailies_consolidado_y_zip(
    db_con_cruce_multi_fecha, tmp_path
):
    out = export_batch_completo(
        date(2026, 2, 9), date(2026, 2, 14), tmp_path,
        db_path=db_con_cruce_multi_fecha,
    )
    assert out["consolidado"].exists()
    assert out["consolidado"].name == "cumplimiento_consolidado_20260209_20260214.xlsx"
    assert out["zip"].exists()
    assert out["zip"].name == "cumplimiento_batch_20260209_20260214.zip"

    assert len(out["dailies"]) == 5  # 09, 10, 11, 12, 14 (13 no tiene datos)
    nombres = {d.name for d in out["dailies"]}
    assert "cumplimiento_20260209.xlsx" in nombres
    assert "cumplimiento_20260214.xlsx" in nombres

    # El 13/02 estaba en el rango pero sin datos, se reporta explicito.
    assert date(2026, 2, 13) in out["fechas_sin_datos_en_rango"]

    # El zip debe contener consolidado + N dailies (todos con nombres
    # relativos, sin paths absolutos).
    import zipfile as _zf
    with _zf.ZipFile(out["zip"]) as zf:
        contenido = set(zf.namelist())
    assert "cumplimiento_consolidado_20260209_20260214.xlsx" in contenido
    assert "cumplimiento_20260214.xlsx" in contenido
    assert len(contenido) == 1 + len(out["dailies"])


def test_export_batch_completo_daily_no_lleva_hoja_por_semana(
    db_con_cruce_multi_fecha, tmp_path
):
    out = export_batch_completo(
        date(2026, 2, 9), date(2026, 2, 14), tmp_path,
        db_path=db_con_cruce_multi_fecha,
    )
    for daily in out["dailies"]:
        wb = openpyxl.load_workbook(daily)
        assert "Por_Semana" not in wb.sheetnames


def test_export_cumplimiento_diario_es_dia_unico(db_con_cruce, tmp_path):
    dest = tmp_path / "solo_dia.xlsx"
    export_cumplimiento_diario(FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Resumen"]
    fechas = {
        ws.cell(row=r, column=1).value
        for r in range(5, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
        and str(ws.cell(row=r, column=1).value) != "TOTAL"
    }
    assert len(fechas) == 1


def test_export_portada_sin_instrucciones_ni_leyendas(db_con_cruce, tmp_path):
    """El archivo es corporativo: no lleva 'como leer', ni leyenda semaforo."""
    dest = tmp_path / "e.xlsx"
    export_cumplimiento_xlsx(FECHA_PROD, FECHA_PROD, dest, db_path=db_con_cruce)
    wb = openpyxl.load_workbook(dest)
    ws = wb["Portada"]
    prohibidas = ("como leer", "como se lee", "leyenda", "semaforo:", "verde 95")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            texto = str(v).lower()
            for pat in prohibidas:
                assert pat not in texto, (
                    f"la portada contiene literatura prohibida '{pat}' en "
                    f"{ws.cell(row=r, column=c).coordinate}: {v!r}"
                )


# ---------------------------------------------------------------------------
# Contrato: exporters.py no puede importar styles directamente
# ---------------------------------------------------------------------------
def test_exporters_no_importa_styles_directamente():
    """La regla de oro: solo `excel_style.py` toca `openpyxl.styles`.

    Si este test falla, alguien anadio formato inline en `exporters.py`.
    Se debe mover a `excel_style.py` como una funcion o NamedStyle nuevo.
    """
    src = (Path(__file__).parent.parent / "app" / "core" / "exporters.py").read_text(
        encoding="utf-8"
    )
    prohibidas = ("Font(", "PatternFill(", "Border(", "Alignment(", "Side(")
    for pat in prohibidas:
        assert pat not in src, (
            f"'{pat}' encontrado en exporters.py. Toda la estilizacion debe "
            f"delegarse en app/core/excel_style.py"
        )
    assert "from openpyxl.styles" not in src, (
        "exporters.py no debe importar de openpyxl.styles"
    )


# ---------------------------------------------------------------------------
# API endpoint
# ---------------------------------------------------------------------------
@pytest.fixture
def api_client_con_cruce(db_con_cruce, tmp_path, monkeypatch):
    """TestClient con la DB ya con cruce persistido."""
    uploads_dir = tmp_path / "uploads"
    uploads_dir.mkdir()
    export_dir = tmp_path / "onedrive_export"
    export_dir.mkdir()

    import app.api.dependencies as deps
    import app.api.routes.exports as exports_route
    import app.api.storage as storage
    import app.config as cfg
    import app.core.db as core_db

    monkeypatch.setattr(cfg, "DB_PATH", db_con_cruce)
    monkeypatch.setattr(core_db, "DB_PATH", db_con_cruce)
    monkeypatch.setattr(deps, "DB_PATH", db_con_cruce)
    monkeypatch.setattr(storage, "UPLOADS_DIR", uploads_dir)
    monkeypatch.setattr(exports_route, "ONEDRIVE_EXPORT_DIR", export_dir)

    from app.api.main import app
    with TestClient(app) as c:
        yield c


def test_endpoint_kpis_excel_devuelve_xlsx(api_client_con_cruce):
    r = api_client_con_cruce.get(
        "/kpis/excel",
        params={"desde": "2026-02-14", "hasta": "2026-02-14"},
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disp = r.headers.get("content-disposition", "")
    assert "cumplimiento_20260214.xlsx" in disp
    assert len(r.content) > 1000, "el archivo debe pesar mas de 1KB"
    assert r.content[:4] == b"PK\x03\x04", "los .xlsx son ZIPs; magic PK.."


def test_endpoint_kpis_excel_rango_invertido_es_422(api_client_con_cruce):
    r = api_client_con_cruce.get(
        "/kpis/excel",
        params={"desde": "2026-02-20", "hasta": "2026-02-14"},
    )
    assert r.status_code == 422
    assert "hasta" in r.json()["detail"].lower()
