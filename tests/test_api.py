"""Tests de la API FastAPI usando TestClient.

Cubre:
- Happy path completo: upload pre_corte (.xlsx) + flash (.csv) -> start-batch ->
  pipeline atomico -> approve -> persist -> verificar SQLite.
- Idempotencia: re-subir mismos archivos devuelve `ya_existia=True`.
- Errores 404 (run inexistente), 415 (extension no permitida) y 422 (nombre
  invalido).
- Endpoints del catalogo SKU.
"""
from __future__ import annotations

import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

FIXTURES = Path(__file__).parent / "fixtures"
PRE_CORTE_XLSX = FIXTURES / "PRE_CORTE_muestra.xlsx"
HOMOLOG_XLSX = FIXTURES / "homologacion.xlsx"
FLASH_CSV = FIXTURES / "FLASH_muestra.csv"


@pytest.fixture
def isolated_env(tmp_path, monkeypatch):
    """Aisla DB y UPLOADS por test para no contaminar el historico real."""
    db_path = tmp_path / "test.sqlite"
    uploads_dir = tmp_path / "uploads"
    uploads_dir.mkdir()

    import app.api.dependencies as deps
    import app.api.storage as storage
    import app.config as cfg
    import app.core.db as core_db

    monkeypatch.setattr(cfg, "DB_PATH", db_path)
    monkeypatch.setattr(cfg, "UPLOADS_DIR", uploads_dir)
    monkeypatch.setattr(core_db, "DB_PATH", db_path)
    monkeypatch.setattr(deps, "DB_PATH", db_path)
    monkeypatch.setattr(storage, "UPLOADS_DIR", uploads_dir)

    yield {"db_path": db_path, "uploads_dir": uploads_dir}


@pytest.fixture
def client(isolated_env):
    from app.api.main import app
    from app.core.db import get_conn, init_db
    from app.core.sku_catalog import import_from_homologacion

    init_db(isolated_env["db_path"])
    with get_conn(isolated_env["db_path"]) as conn:
        import_from_homologacion(HOMOLOG_XLSX, conn)

    with TestClient(app) as c:
        yield c


def _upload_pre(client: TestClient, filename: str, src: Path = PRE_CORTE_XLSX) -> dict:
    with src.open("rb") as fh:
        r = client.post(
            "/files/upload-pre-corte",
            files={"file": (
                filename, fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
        )
    assert r.status_code == 200, r.text
    return r.json()


def _upload_flash(client: TestClient, filename: str = "FLASH_muestra.csv") -> dict:
    src = FIXTURES / filename
    with src.open("rb") as fh:
        r = client.post(
            "/files/upload-flash",
            files={"file": (filename, fh, "text/csv")},
        )
    assert r.status_code == 200, r.text
    return r.json()


def _make_variant(tmp_path: Path, src: Path, tag: str) -> Path:
    """Copia el fixture y modifica una celda inofensiva para generar hash distinto.

    Se toca INVENTARIO FISICO (celdas planas, no formulas) asi el RESUMEN queda
    identico y los totales del test no cambian, pero el hash SHA256 si.
    """
    dst = tmp_path / f"variant_{tag}.xlsx"
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    if "INVENTARIO FISICO" in wb.sheetnames:
        ws = wb["INVENTARIO FISICO"]
        ws["B2"] = f"variant-{tag}"
    else:
        wb.create_sheet(f"variant_{tag}_marker")
    wb.save(dst)
    return dst


def test_health(client):
    r = client.get("/health")
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "ok"
    assert "version" in body


def test_upload_pre_corte_valida_fecha_y_cobertura(client):
    body = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx")
    assert body["tipo"] == "pre_corte"
    assert body["fecha_archivo"] == "2026-02-13"
    assert body["fecha_produccion"] == "2026-02-14"
    assert body["num_filas_original"] == 18
    assert body["notificacion_presente"] is True
    assert body["catalog_coverage_pct"] == 100.0
    assert body["num_filas_sin_sap"] == 0
    assert body["ya_existia"] is False
    assert body["dias_saltados"] == 0
    assert body["motivos_saltados"] == []


def test_upload_pre_corte_sabado_reporta_skip_a_lunes(client, tmp_path):
    """Sabado 07/02/2026 -> domingo 08 (no laboral) -> lunes 09 con dias_saltados=1."""
    import shutil

    dst = tmp_path / "PRE CORTE 07.02.2026.xlsx"
    shutil.copy(PRE_CORTE_XLSX, dst)
    body = _upload_pre(client, "PRE CORTE 07.02.2026.xlsx", src=dst)
    assert body["fecha_archivo"] == "2026-02-07"
    assert body["fecha_produccion"] == "2026-02-09"
    assert body["dias_saltados"] == 1
    assert len(body["motivos_saltados"]) == 1
    assert "Domingo" in body["motivos_saltados"][0]


def test_get_calendario_no_laborales_2026(client):
    r = client.get("/calendario/no-laborales", params={"year": 2026})
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == 2026
    assert body["cobertura_desde"].startswith("2024")
    assert body["cobertura_hasta"].startswith("2030")
    fechas = {f["fecha"] for f in body["festivos"]}
    assert "2026-01-01" in fechas
    assert "2026-04-02" in fechas  # Jueves Santo
    assert "2026-12-25" in fechas
    assert "2026-07-13" not in fechas  # Chiquinquira excluido


def test_get_calendario_no_laborales_fuera_de_rango_es_404(client):
    r = client.get("/calendario/no-laborales", params={"year": 2019})
    assert r.status_code == 404


def test_upload_pre_corte_idempotente(client):
    a = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx")
    b = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx")
    assert a["carga_id"] == b["carga_id"]
    assert b["ya_existia"] is True


def test_upload_pre_corte_rechaza_csv(client):
    with (FIXTURES / "FLASH_muestra.csv").open("rb") as fh:
        r = client.post(
            "/files/upload-pre-corte",
            files={"file": ("PRE CORTE 13.02.2026.csv", fh, "text/csv")},
        )
    assert r.status_code == 415
    assert "extension" in r.text.lower()


def test_upload_flash(client):
    body = _upload_flash(client)
    assert body["tipo"] == "flash"
    assert body["num_filas_original"] == 6
    assert body["hash_sha256"]


def test_start_batch_valida_tipos(client):
    pre = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx")
    flash = _upload_flash(client)

    r_bad = client.post(
        "/runs/start-batch",
        json={"pre_corte_carga_ids": [pre["carga_id"]], "flash_carga_id": pre["carga_id"]},
    )
    assert r_bad.status_code == 400

    r_ok = client.post(
        "/runs/start-batch",
        json={"pre_corte_carga_ids": [pre["carga_id"]], "flash_carga_id": flash["carga_id"]},
    )
    assert r_ok.status_code == 200
    body = r_ok.json()
    assert body["total_sub_runs"] == 1
    assert len(body["sub_run_ids"]) == 1


def test_pipeline_completo_happy_path(client, isolated_env):
    pre = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx")
    flash = _upload_flash(client)

    r_batch = client.post(
        "/runs/start-batch",
        json={"pre_corte_carga_ids": [pre["carga_id"]], "flash_carga_id": flash["carga_id"]},
    )
    sub_id = r_batch.json()["sub_run_ids"][0]

    r_date = client.post(f"/pipeline/{sub_id}/extract-date")
    assert r_date.status_code == 200
    assert r_date.json()["fecha_produccion"] == "2026-02-14"

    r_lp = client.post(f"/pipeline/{sub_id}/load-pre-corte")
    assert r_lp.status_code == 200
    assert r_lp.json()["num_filas_procesadas"] == 18

    r_lf = client.post(f"/pipeline/{sub_id}/load-flash")
    assert r_lf.status_code == 200

    r_agg = client.post(f"/pipeline/{sub_id}/aggregate")
    assert r_agg.status_code == 200
    agg = r_agg.json()
    assert agg["fecha_produccion"] == "2026-02-14"
    assert agg["materiales_flash"] >= 3

    r_match = client.post(f"/pipeline/{sub_id}/match")
    assert r_match.status_code == 200
    m = r_match.json()
    assert m["matched"] >= 3
    assert m["solo_flash"] >= 1

    r_val = client.post(f"/pipeline/{sub_id}/validate")
    assert r_val.status_code == 200
    val = r_val.json()
    for v in val["validaciones"]:
        assert isinstance(v["ok"], bool)

    r_kpi = client.post(f"/pipeline/{sub_id}/kpis-preview")
    assert r_kpi.status_code == 200
    kpi = r_kpi.json()
    assert len(kpi["cards"]) >= 3

    r_get = client.get(f"/runs/{sub_id}")
    assert r_get.json()["status"] == "awaiting_approval"

    r_reject_pre_approve = client.post(f"/pipeline/{sub_id}/persist")
    assert r_reject_pre_approve.status_code == 409

    r_appr = client.post(
        f"/runs/{sub_id}/approve",
        json={"aprobado_por": "test", "comentarios": "ok"},
    )
    assert r_appr.status_code == 200
    assert r_appr.json()["status"] == "approved"

    r_persist = client.post(f"/pipeline/{sub_id}/persist")
    assert r_persist.status_code == 200
    p = r_persist.json()
    assert p["cruce_filas_insertadas"] >= 3

    r_persist2 = client.post(f"/pipeline/{sub_id}/persist")
    assert r_persist2.status_code == 409

    conn = sqlite3.connect(isolated_env["db_path"])
    conn.row_factory = sqlite3.Row
    n_cruce = conn.execute("SELECT COUNT(*) AS n FROM cruce").fetchone()["n"]
    n_runs_completed = conn.execute(
        "SELECT COUNT(*) AS n FROM runs WHERE status = 'completed'"
    ).fetchone()["n"]
    conn.close()
    assert n_cruce >= 3
    assert n_runs_completed == 1


def test_pipeline_batch_dos_precortes_un_flash(client, isolated_env, tmp_path):
    v2 = _make_variant(tmp_path, PRE_CORTE_XLSX, tag="dia2")
    pre1 = _upload_pre(client, "PRE CORTE 13.02.2026.xlsx", src=PRE_CORTE_XLSX)
    pre2 = _upload_pre(client, "PRE CORTE 14.02.2026.xlsx", src=v2)
    assert pre1["carga_id"] != pre2["carga_id"]
    flash = _upload_flash(client)

    r_batch = client.post(
        "/runs/start-batch",
        json={
            "pre_corte_carga_ids": [pre1["carga_id"], pre2["carga_id"]],
            "flash_carga_id": flash["carga_id"],
        },
    )
    assert r_batch.status_code == 200
    body = r_batch.json()
    assert body["total_sub_runs"] == 2

    master_id = body["master_run_id"]
    r_subs = client.get(f"/runs/{master_id}/sub-runs")
    assert r_subs.status_code == 200
    assert len(r_subs.json()) == 2


def test_get_run_inexistente(client):
    r = client.get("/runs/nonexistent")
    assert r.status_code == 404


def test_upload_pre_corte_nombre_sin_fecha_falla(client):
    with PRE_CORTE_XLSX.open("rb") as fh:
        r = client.post(
            "/files/upload-pre-corte",
            files={"file": (
                "archivo_sin_fecha.xlsx", fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
        )
    assert r.status_code == 422


# ---------- Endpoints del catalogo SKU ----------

def test_get_catalog_lista_entradas(client):
    r = client.get("/catalog?limit=10")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] > 0
    assert "homologacion" in body["por_fuente"]
    assert len(body["entradas"]) <= 10


def test_post_manual_mapping_registra_y_prioriza(client):
    payload = {
        "referencia": "NUEVA_MARCA",
        "tipo": "AA",
        "formato": "AMARRADO",
        "unidades_por_empaque": 30,
        "material_sap": 55555,
        "nombre_notificacion": "test manual",
    }
    r = client.post("/catalog/manual-mapping", json=payload)
    assert r.status_code == 200
    body = r.json()
    assert body["es_nueva"] is True
    assert body["material_sap"] == 55555

    r2 = client.get("/catalog?referencia=NUEVA_MARCA")
    assert r2.status_code == 200
    entradas = r2.json()["entradas"]
    assert len(entradas) == 1
    assert entradas[0]["fuente"] == "manual"


def test_post_import_homologacion_endpoint(client):
    with HOMOLOG_XLSX.open("rb") as fh:
        r = client.post(
            "/catalog/import-homologacion",
            files={"file": (
                "homologacion.xlsx", fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
        )
    assert r.status_code == 200
    body = r.json()
    assert body["import_stats"]["leidas"] > 0
    assert body["catalog_stats"]["total_entradas"] > 0
