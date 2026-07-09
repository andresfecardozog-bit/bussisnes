"""ConfigurableLoader: carga fuentes segun el loader spec del MatchProfile.

Deterministico, sin LLM. Maneja las trampas reales documentadas en
AGENTS.md (caso CEN/SAP):

- Extension mentirosa: detecta el formato real por firma binaria (PK zip =
  xlsx moderno) y no por extension. Los .XLS del SAP son xlsx renombrados.
- Sin encabezados: `header_row=None` -> columnas posicionales 1-based.
- Auto-deteccion de hoja: si `sheet=None`, elige la primera hoja cuyo
  header contiene todas las columnas requeridas (evita hojas pivote).
- Filas fantasma: descarta filas completamente vacias (CEN P4 declara
  1M de filas).

Retorna siempre `(DataFrame, meta)` con contabilidad de filas para los
validadores cero-perdida: num_filas_original, num_filas_procesadas,
num_filas_descartadas + motivo.
"""
from __future__ import annotations

import shutil
import tempfile
from pathlib import Path
from typing import Any, Callable

import openpyxl
import pandas as pd

from app.core.loaders import _clean_numeric_series, hash_file
from app.platform.profile import (
    ColumnDtype,
    ColumnSpec,
    LoaderSpec,
    RegisteredLoaderSpec,
    TabularLoaderSpec,
)

_XLSX_MAGIC = b"PK\x03\x04"
_XLS_LEGACY_MAGIC = b"\xd0\xcf\x11\xe0"

# Registry de loaders custom en codigo (caso RESUMEN del PRE CORTE).
# name -> callable(path) -> (DataFrame, meta)
_REGISTERED: dict[str, Callable[[Path], tuple[pd.DataFrame, dict[str, Any]]]] = {}


def register_loader(
    name: str,
) -> Callable[[Callable[[Path], tuple[pd.DataFrame, dict[str, Any]]]], Callable]:
    def deco(fn: Callable[[Path], tuple[pd.DataFrame, dict[str, Any]]]) -> Callable:
        _REGISTERED[name] = fn
        return fn

    return deco


def registered_loader_names() -> list[str]:
    return sorted(_REGISTERED)


def detect_real_format(path: Path) -> str:
    """'xlsx' | 'xls_legacy' | 'csv' segun la firma binaria, no la extension."""
    with path.open("rb") as fh:
        head = fh.read(4)
    if head.startswith(_XLSX_MAGIC):
        return "xlsx"
    if head.startswith(_XLS_LEGACY_MAGIC):
        return "xls_legacy"
    return "csv"


def _open_workbook(path: Path) -> openpyxl.Workbook:
    """Abre un xlsx aunque la extension mienta (copia temporal .xlsx)."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    tmp = Path(tempfile.gettempdir()) / f"nutri_reext_{path.stem}.xlsx"
    shutil.copy(path, tmp)
    return openpyxl.load_workbook(tmp, data_only=True, read_only=True)


def _sheet_rows(ws: Any) -> list[tuple[Any, ...]]:
    """Materializa filas recortando la cola de filas completamente vacias."""
    rows = list(ws.iter_rows(values_only=True))
    while rows and all(v is None or (isinstance(v, str) and not v.strip()) for v in rows[-1]):
        rows.pop()
    return rows


def _headers_match(header_row: tuple[Any, ...], spec: TabularLoaderSpec) -> bool:
    headers = {str(v).strip() for v in header_row if v is not None}
    wanted = {str(c.source) for c in spec.columns if c.required and isinstance(c.source, str)}
    return wanted.issubset(headers) and bool(wanted)


def _pick_sheet(wb: openpyxl.Workbook, spec: TabularLoaderSpec) -> Any:
    if isinstance(spec.sheet, int):
        return wb[wb.sheetnames[spec.sheet]]
    if isinstance(spec.sheet, str):
        if spec.sheet not in wb.sheetnames:
            raise ValueError(
                f"Hoja '{spec.sheet}' no existe. Hojas: {wb.sheetnames}"
            )
        return wb[spec.sheet]
    # Auto-deteccion: primera hoja cuyo header row contiene las requeridas.
    if spec.header_row is None:
        return wb[wb.sheetnames[0]]
    for name in wb.sheetnames:
        ws = wb[name]
        for i, row in enumerate(ws.iter_rows(max_row=spec.header_row, values_only=True)):
            if i + 1 == spec.header_row and _headers_match(row, spec):
                return ws
    raise ValueError(
        f"Ninguna hoja contiene los headers requeridos "
        f"{[c.source for c in spec.columns if c.required]}. Hojas: {wb.sheetnames}"
    )


def _int_like_to_str(v: Any) -> Any:
    """Celdas numericas de Excel llegan como float: '30018.0' != '30018'.
    Los codigos/keys enteros se renderizan sin el '.0'."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int) and not isinstance(v, bool):
        return str(v)
    return v


def _coerce_column(series: pd.Series, col: ColumnSpec) -> pd.Series:
    if col.dtype == ColumnDtype.STR:
        out = series.map(_int_like_to_str).astype("string").str.strip()
        return out
    if col.dtype == ColumnDtype.INT:
        return pd.to_numeric(series, errors="coerce").astype("Int64")
    if col.dtype == ColumnDtype.FLOAT:
        return pd.to_numeric(series, errors="coerce")
    if col.dtype == ColumnDtype.FLOAT_CLEAN:
        return _clean_numeric_series(series)
    if col.dtype == ColumnDtype.DATE:
        if col.date_format:
            dt = pd.to_datetime(series, format=col.date_format, errors="coerce")
        else:
            dt = pd.to_datetime(series, errors="coerce")
        return dt.dt.date
    raise ValueError(f"dtype no soportado: {col.dtype}")


def _load_tabular(path: Path, spec: TabularLoaderSpec) -> tuple[pd.DataFrame, dict[str, Any]]:
    real_format = detect_real_format(path)
    if real_format == "csv":
        raw = pd.read_csv(path, dtype=object, header=None)
        rows = [tuple(r) for r in raw.itertuples(index=False, name=None)]
    elif real_format == "xlsx":
        wb = _open_workbook(path)
        try:
            ws = _pick_sheet(wb, spec)
            rows = _sheet_rows(ws)
        finally:
            wb.close()
    else:
        raise ValueError(
            f"{path.name}: formato xls legacy (BIFF) no soportado; convertir a xlsx"
        )

    if spec.header_row is not None:
        if len(rows) < spec.header_row:
            raise ValueError(f"{path.name}: no hay fila de header {spec.header_row}")
        header = [
            str(v).strip() if v is not None else f"_col_{i+1}"
            for i, v in enumerate(rows[spec.header_row - 1])
        ]
        data_rows = rows[spec.header_row:]
        df_raw = pd.DataFrame(data_rows, columns=header)
        # Headers duplicados en el archivo: pandas los tolera, nosotros
        # tomamos la primera ocurrencia al seleccionar por nombre.
        df_raw = df_raw.loc[:, ~df_raw.columns.duplicated()]
    else:
        data_rows = rows
        # Si la primera fila esta completamente vacia (caso SAP), se
        # descarta pero se contabiliza.
        df_raw = pd.DataFrame(data_rows)
        df_raw.columns = [i + 1 for i in range(df_raw.shape[1])]

    num_filas_original = len(df_raw)

    # Descartar filas completamente vacias (fantasma / separadores).
    non_empty_mask = df_raw.notna().any(axis=1)
    df_raw = df_raw[non_empty_mask]
    filas_vacias = int(num_filas_original - len(df_raw))

    out = pd.DataFrame(index=df_raw.index)
    missing_required: list[str] = []
    for col in spec.columns:
        key: Any = col.source
        if isinstance(key, str):
            if key not in df_raw.columns:
                if col.required:
                    missing_required.append(str(key))
                    continue
                out[col.name] = pd.NA
                continue
            series = df_raw[key]
        else:
            if key not in df_raw.columns:
                if col.required:
                    missing_required.append(f"col posicional {key}")
                    continue
                out[col.name] = pd.NA
                continue
            series = df_raw[key]
        out[col.name] = _coerce_column(series, col)

    if missing_required:
        raise ValueError(
            f"{path.name}: faltan columnas obligatorias: {missing_required}. "
            f"Columnas presentes: {list(df_raw.columns)[:40]}"
        )

    descartes: dict[str, int] = {"filas_vacias": filas_vacias}
    for col_name in spec.drop_rows_where_null:
        if col_name not in out.columns:
            raise ValueError(
                f"drop_rows_where_null referencia columna inexistente: {col_name}"
            )
        before = len(out)
        out = out[out[col_name].notna()]
        descartes[f"null_{col_name}"] = before - len(out)

    meta: dict[str, Any] = {
        "filename": path.name,
        "path": str(path),
        "hash_sha256": hash_file(path),
        "formato_real": real_format,
        "num_filas_original": num_filas_original,
        "num_filas_procesadas": len(out),
        "descartes": descartes,
    }
    return out.reset_index(drop=True), meta


def load_source(path: str | Path, spec: LoaderSpec) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Punto de entrada unico: despacha segun el tipo de loader spec."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    if isinstance(spec, RegisteredLoaderSpec):
        if spec.name not in _REGISTERED:
            raise ValueError(
                f"Loader registrado '{spec.name}' no existe. "
                f"Disponibles: {registered_loader_names()}"
            )
        return _REGISTERED[spec.name](path)
    return _load_tabular(path, spec)


# ---------------------------------------------------------------------------
# Loaders registrados: reusan los parsers legados ya validados por tests.
# ---------------------------------------------------------------------------

@register_loader("pre_corte_resumen")
def _load_pre_corte_registrado(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    from app.core.loaders import load_pre_corte

    return load_pre_corte(path)


@register_loader("flash_sap")
def _load_flash_registrado(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    from app.core.loaders import load_flash

    return load_flash(path)
