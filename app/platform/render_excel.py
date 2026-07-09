"""Renderer Excel declarativo de la plataforma (Fase 4).

Genera un .xlsx corporativo a partir del `report.excel` de un MatchProfile
y del `GenericMatchResult` que produjo el motor. Es la generalizacion del
exporter legado (`app/core/exporters.py`): las hojas ya no estan
hardcodeadas, las declara el profile (`ExcelReportSpec`).

Soporta el contrato extendido del caso CEN vs SAP:
- Hojas `kind="breakdown"`: desgloses dimensionales de result.breakdowns.
- Nivel de servicio: tablas "Nivel de servicio" (lineas) y "Pedidos" en la
  portada cuando el profile declara service_level.
- data_model: al final se agrega una hoja por tabla del modelo exportable
  (build_data_model). Son la "base de datos en Excel": tablas planas con
  ids, escritas como Excel Table nativa cuyo ListObject name es EXACTO al
  nombre de la tabla (FactNivelServicio, DimCliente...) para que Power
  Query las importe por nombre. Sin titulos ni decoracion, sin truncar.

Regla heredada e innegociable: este modulo NO importa `Font`, `PatternFill`,
`Border`, `Alignment` ni `Side` de openpyxl. Toda la estilizacion se delega
en `app.core.excel_style` (paleta corporativa centralizada). Un test de
puritanismo lo verifica.

Convenciones de datos:
- El motor guarda porcentajes como numero 0-100 (ej. cumplimiento 98.21).
  En hojas de reporte se dividen entre 100 para el formato "0.00%"; en las
  hojas del data_model se conservan crudos (es una base de datos).
- Las columnas internas `_k{i}` del matched se renombran al nombre de la
  join key izquierda (`profile.join.keys[i].left`) para legibilidad.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core import excel_style as xs
from app.core.excel_style import ColumnSpec, KpiRow
from app.platform.data_model import build_data_model
from app.platform.engine import GenericMatchResult
from app.platform.profile import ExcelSheetSpec, KpiSpec, MatchProfile, SemaforoSpec

LOGO_PATH = Path(__file__).resolve().parents[2] / "resources" / "image_720508810_0.jpg"

# Columnas cuyo nombre delata un porcentaje 0-100 del motor.
_PCT_TOKEN = "pct"

_SERVICE_LEVEL_LABELS = {
    "completo": "Completo",
    "parcial": "Parcial",
    "no_entregado": "No entregado",
    "sin_pedido": "Sin pedido",
}


# ---------------------------------------------------------------------------
# Helpers de datos
# ---------------------------------------------------------------------------

def _rename_key_columns(df: pd.DataFrame, profile: MatchProfile) -> pd.DataFrame:
    """`_k0`, `_k1`... -> nombre de la join key izquierda del profile.

    Si el nombre destino ya existe como columna (key presente en una sola
    fuente, sin sufijo del merge), se usa `<nombre>_key` para no pisarla.
    """
    renames: dict[str, str] = {}
    for i, key in enumerate(profile.join.keys):
        kcol = f"_k{i}"
        if kcol not in df.columns:
            continue
        target = key.left if key.left not in df.columns else f"{key.left}_key"
        renames[kcol] = target
    return df.rename(columns=renames) if renames else df


def _is_pct_column(name: str) -> bool:
    return _PCT_TOKEN in name.lower()


def _column_kind(df: pd.DataFrame, name: str) -> str:
    serie = df[name]
    if _is_pct_column(name) and pd.api.types.is_numeric_dtype(serie):
        return "pct"
    if pd.api.types.is_datetime64_any_dtype(serie):
        return "date"
    if pd.api.types.is_numeric_dtype(serie):
        return "int"
    return "text"


def _semaforo_column(name: str) -> bool:
    """Columnas porcentuales que llevan semaforo: cumplimiento_pct y todo
    lo que termine en _pct (metricas ratio de breakdowns)."""
    lname = name.lower()
    return lname == "cumplimiento_pct" or lname.endswith("_pct")


def _prepare_table_df(
    df: pd.DataFrame, spec_columns: list[str] | None, profile: MatchProfile
) -> tuple[pd.DataFrame, list[ColumnSpec]]:
    """Renombra keys internas, filtra/ordena columnas segun el spec y
    convierte porcentajes 0-100 a fraccion para el formato "0.00%"."""
    df = _rename_key_columns(df.copy(), profile)

    if spec_columns:
        visibles = [c for c in spec_columns if c in df.columns]
    else:
        visibles = [c for c in df.columns if not c.startswith("_")]

    df = df[visibles].copy()
    columns: list[ColumnSpec] = []
    for name in visibles:
        kind = _column_kind(df, name)
        if kind == "pct":
            df[name] = pd.to_numeric(df[name], errors="coerce") / 100.0
        columns.append(ColumnSpec(name.replace("_", " ").capitalize(), name, kind))
    return df, columns


def _source_dataframe(result: GenericMatchResult, source: str) -> pd.DataFrame:
    frames = {
        "matched": result.matched,
        "no_cruzados": result.no_cruzados,
        "solo_left": result.solo_left,
        "solo_right": result.solo_right,
    }
    if source not in frames:
        raise ValueError(f"Fuente de hoja no soportada: '{source}'")
    return frames[source]


# ---------------------------------------------------------------------------
# KPIs -> filas de la tabla de indicadores
# ---------------------------------------------------------------------------

def _semaforo_rangos(spec: SemaforoSpec) -> dict:
    """SemaforoSpec del profile (0-100) -> rangos en fraccion para excel_style."""
    return {
        "verde_min": spec.verde_min / 100.0,
        "verde_max": None if spec.verde_max is None else spec.verde_max / 100.0,
        "amarillo_min": spec.amarillo_min / 100.0,
        "amarillo_max": None if spec.amarillo_max is None else spec.amarillo_max / 100.0,
    }


def _kpi_rows(profile: MatchProfile, result: GenericMatchResult) -> list[KpiRow]:
    rows: list[KpiRow] = []
    for kpi in profile.kpis:
        valor = result.kpis.get(kpi.id)
        if valor is None:
            continue
        rows.append(_kpi_row(kpi, valor))
    return rows


def _kpi_row(kpi: KpiSpec, valor: float) -> KpiRow:
    if kpi.op == "ratio_pct_of_sums":
        return KpiRow(
            kpi.label,
            float(valor) / 100.0,
            "pct",
            semaforo=kpi.semaforo is not None,
            rangos=_semaforo_rangos(kpi.semaforo) if kpi.semaforo else None,
        )
    return KpiRow(kpi.label, float(valor), "int")


# ---------------------------------------------------------------------------
# Nivel de servicio -> tablas de la portada
# ---------------------------------------------------------------------------

def _service_level_frames(
    service_level: dict,
) -> tuple[pd.DataFrame, pd.DataFrame | None]:
    """Bloque service_level del motor -> (tabla lineas, tabla pedidos|None).

    Porcentajes ya convertidos a fraccion para el formato "0.00%".
    """
    clases = service_level.get("clases", {})
    filas = []
    for clase in ("completo", "parcial", "no_entregado", "sin_pedido"):
        datos = clases.get(clase)
        if datos is None:
            continue
        pct_lineas = datos.get("pct_lineas")
        pct_unidades = datos.get("pct_unidades_plan")
        filas.append({
            "clase": _SERVICE_LEVEL_LABELS[clase],
            "lineas": datos.get("lineas", 0),
            "unidades_plan": datos.get("unidades_plan", 0.0),
            "unidades_real": datos.get("unidades_real", 0.0),
            "pct_lineas": None if pct_lineas is None else pct_lineas / 100.0,
            "pct_unidades": None if pct_unidades is None else pct_unidades / 100.0,
        })
    lineas_df = pd.DataFrame(filas)

    pedidos_df = None
    pedidos = service_level.get("pedidos")
    if pedidos:
        filas_p = [{
            "clase": "Total",
            "pedidos": pedidos.get("total", 0),
            "pct": 1.0 if pedidos.get("total") else None,
        }]
        for clase in ("completo", "parcial", "no_entregado"):
            datos = pedidos.get("clases", {}).get(clase)
            if datos is None:
                continue
            pct = datos.get("pct")
            filas_p.append({
                "clase": _SERVICE_LEVEL_LABELS[clase],
                "pedidos": datos.get("pedidos", 0),
                "pct": None if pct is None else pct / 100.0,
            })
        pedidos_df = pd.DataFrame(filas_p)
    return lineas_df, pedidos_df


def _write_service_level_block(
    ws: Worksheet, service_level: dict, start_row: int
) -> int:
    """Escribe las tablas 'Nivel de servicio' y 'Pedidos'. Retorna la fila
    siguiente al bloque."""
    lineas_df, pedidos_df = _service_level_frames(service_level)

    row = start_row
    ws.cell(row=row, column=1, value="Nivel de servicio").style = "st_subtitle"
    row += 1
    columnas_lineas = [
        ColumnSpec("Clase", "clase", "text"),
        ColumnSpec("Lineas", "lineas", "int"),
        ColumnSpec("Unidades plan", "unidades_plan", "int"),
        ColumnSpec("Unidades real", "unidades_real", "int"),
        ColumnSpec("% lineas", "pct_lineas", "pct"),
        ColumnSpec("% unidades", "pct_unidades", "pct"),
    ]
    xs.write_dataframe_as_table(
        ws, lineas_df, columnas_lineas, "NivelServicio", start_row=row
    )
    row += len(lineas_df) + 2

    if pedidos_df is not None:
        ws.cell(row=row, column=1, value="Pedidos").style = "st_subtitle"
        row += 1
        columnas_pedidos = [
            ColumnSpec("Clase", "clase", "text"),
            ColumnSpec("Pedidos", "pedidos", "int"),
            ColumnSpec("% pedidos", "pct", "pct"),
        ]
        xs.write_dataframe_as_table(
            ws, pedidos_df, columnas_pedidos, "PedidosNivelServicio", start_row=row
        )
        row += len(pedidos_df) + 2
    return row


# ---------------------------------------------------------------------------
# Hojas
# ---------------------------------------------------------------------------

def _write_portada(
    wb: Workbook, spec: ExcelSheetSpec, profile: MatchProfile, result: GenericMatchResult
) -> None:
    ws = wb.create_sheet(spec.name)
    xs.insert_logo(ws, LOGO_PATH, anchor="A1", max_height_px=60)
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 22

    titulo = profile.descripcion or profile.profile_id
    xs.add_title_block(
        ws,
        titulo,
        f"Perfil {profile.profile_id} v{profile.version}"
        f"      Generado: {datetime.now():%d/%m/%Y %H:%M}",
        span_cols=4,
        start_row=4,
    )

    kpis = _kpi_rows(profile, result)
    resumen = result.summary()
    kpis.append(KpiRow("Filas no cruzadas", resumen["no_cruzados"], "int"))
    next_row = xs.write_kpi_table(ws, kpis, start_row=8, start_col=1)

    if profile.service_level and result.service_level:
        _write_service_level_block(ws, result.service_level, next_row + 1)

    xs.autofit_columns(ws, columns=range(3, 7), max_width=18)
    xs.set_page_setup(ws, freeze_from_row=8)


def _write_kpi_resumen(
    wb: Workbook, spec: ExcelSheetSpec, profile: MatchProfile, result: GenericMatchResult
) -> None:
    ws = wb.create_sheet(spec.name)
    next_row = xs.add_title_block(
        ws,
        "Indicadores del cruce",
        f"Perfil {profile.profile_id} v{profile.version}",
        span_cols=2,
    )
    xs.write_kpi_table(ws, _kpi_rows(profile, result), start_row=next_row)
    xs.set_page_setup(ws, freeze_from_row=next_row + 1)


def _write_data_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    columns: list[ColumnSpec],
    table_name: str,
    start_row: int,
    group_key: str | None = None,
) -> None:
    """Tabla + semaforo en columnas porcentuales de cumplimiento + ajustes."""
    xs.write_dataframe_as_table(
        ws, df, columns, table_name, start_row=start_row, group_by_key=group_key
    )
    n_rows = max(len(df), 1)
    for j, col in enumerate(columns):
        if col.kind == "pct" and _semaforo_column(col.key):
            letter = get_column_letter(1 + j)
            xs.apply_traffic_light(
                ws, letter, from_row=start_row + 1, to_row=start_row + n_rows
            )
    xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def _write_tabla(
    wb: Workbook, spec: ExcelSheetSpec, profile: MatchProfile, result: GenericMatchResult
) -> None:
    source = spec.source or "matched"
    df, columns = _prepare_table_df(
        _source_dataframe(result, source), spec.columns, profile
    )

    ws = wb.create_sheet(spec.name)
    xs.add_title_block(
        ws,
        spec.name.replace("_", " "),
        f"Fuente: {source}  -  Perfil {profile.profile_id} v{profile.version}",
        span_cols=max(len(columns), 2),
    )
    group_key = spec.group_by if spec.group_by in df.columns else None
    _write_data_sheet(ws, df, columns, spec.name, start_row=4, group_key=group_key)


def _write_breakdown(
    wb: Workbook, spec: ExcelSheetSpec, profile: MatchProfile, result: GenericMatchResult
) -> None:
    if spec.breakdown_id not in result.breakdowns:
        raise ValueError(
            f"Hoja '{spec.name}': el resultado no trae el breakdown "
            f"'{spec.breakdown_id}' (breakdowns disponibles: "
            f"{sorted(result.breakdowns)})"
        )
    bd_spec = next(
        (b for b in profile.breakdowns if b.id == spec.breakdown_id), None
    )
    label = bd_spec.label if bd_spec else spec.name.replace("_", " ")

    df, columns = _prepare_table_df(
        result.breakdowns[spec.breakdown_id], spec.columns, profile
    )
    ws = wb.create_sheet(spec.name)
    xs.add_title_block(
        ws,
        label,
        f"Desglose '{spec.breakdown_id}'  -  Perfil {profile.profile_id} "
        f"v{profile.version}",
        span_cols=max(len(columns), 2),
    )
    group_key = spec.group_by if spec.group_by in df.columns else None
    _write_data_sheet(ws, df, columns, spec.name, start_row=4, group_key=group_key)


def _data_model_column_kind(df: pd.DataFrame, name: str) -> str:
    """Tipos para las hojas 'base de datos': sin conversion de porcentajes
    (valores crudos del motor) y floats con decimales preservados."""
    serie = df[name]
    if pd.api.types.is_datetime64_any_dtype(serie):
        return "date"
    if name == "fact_id" or name.endswith("_id"):
        return "int"
    if pd.api.types.is_integer_dtype(serie):
        return "int"
    if pd.api.types.is_numeric_dtype(serie):
        return "num"
    return "text"


def _write_data_model_sheets(
    wb: Workbook, profile: MatchProfile, result: GenericMatchResult
) -> None:
    """Una hoja por tabla del modelo, sin decoracion: header en fila 1,
    ListObject con el nombre EXACTO de la tabla (Power Query importa por
    nombre). Datos completos, sin truncar."""
    tablas = build_data_model(profile, result)
    for nombre, df in tablas.items():
        ws = wb.create_sheet(nombre[:31])
        columns = [
            ColumnSpec(col, col, _data_model_column_kind(df, col))
            for col in df.columns
        ]
        xs.write_dataframe_as_table(ws, df, columns, nombre, start_row=1)
        xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
        xs.set_page_setup(ws, freeze_from_row=2)


# ---------------------------------------------------------------------------
# Entrada publica
# ---------------------------------------------------------------------------

def render_excel(profile: MatchProfile, result: GenericMatchResult, dest: Path) -> Path:
    """Genera el .xlsx declarado en `profile.report.excel`.

    Cada `ExcelSheetSpec` produce una hoja en el orden declarado:
    - portada: logo + titulo (descripcion del profile) + tabla KPI +
      tablas de nivel de servicio si el profile lo declara.
    - kpi_resumen: tabla de indicadores (label | valor) con semaforo.
    - tabla: DataFrame de la particion `source` como Excel Table nativa.
    - breakdown: desglose dimensional de result.breakdowns.

    Si el profile declara data_model, al final se agrega una hoja por
    tabla del modelo exportable (fact + dimensiones) como Excel Table con
    nombre estable.
    """
    if profile.report is None or profile.report.excel is None:
        raise ValueError(
            f"El profile '{profile.profile_id}' no declara report.excel"
        )

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    xs.register_styles(wb)

    for sheet in profile.report.excel.sheets:
        if sheet.kind == "portada":
            _write_portada(wb, sheet, profile, result)
        elif sheet.kind == "kpi_resumen":
            _write_kpi_resumen(wb, sheet, profile, result)
        elif sheet.kind == "tabla":
            _write_tabla(wb, sheet, profile, result)
        elif sheet.kind == "breakdown":
            _write_breakdown(wb, sheet, profile, result)
        else:
            raise ValueError(f"Tipo de hoja no soportado: '{sheet.kind}'")

    if profile.data_model is not None:
        _write_data_model_sheets(wb, profile, result)

    wb.save(dest)
    return dest
