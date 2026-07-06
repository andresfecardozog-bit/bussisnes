"""Capa de estilo unica del workbook Excel (paleta corporativa NutriAvicola).

**Este es el UNICO modulo que puede llamar a `openpyxl.styles.Font`,
`PatternFill`, `Border`, `Alignment`, `Side`.** Cualquier otro modulo (en
particular `app.core.exporters`) debe consumir las funciones publicas que
aca se exportan. Un test en `tests/test_exporters.py` verifica que
`exporters.py` no importe styles de openpyxl directamente.

Diseno:

- **Paleta corporativa NutriAvicola**: navy oscuro para header y texto,
  naranja para acentos y totales, blanco para fondo, grises para grid.
- Bordes visibles en cada celda (color `#8C8C8C`, no gris casi invisible).
- Excel Table nativa con `TableStyleMedium9` (bordes marcados).
- KPI table (label|valor) con bordes fuertes, semaforeado.
- Portada minimalista: logo + titulo + tabla de indicadores. Sin
  instrucciones, sin leyendas: es un archivo corporativo, no un tutorial.

Cambiar la paleta corporativa = editar una constante hex en este archivo.
Todas las hojas heredan el cambio automaticamente.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Paleta corporativa NutriAvicola (basada en el logo `resources/`).
# Cambiar aca => se propaga a todas las hojas.
# ---------------------------------------------------------------------------
BRAND_NAVY = "0F2E4C"         # azul-negro del texto del logo (headers)
BRAND_NAVY_DEEP = "081A2C"    # variante mas oscura (bordes de header)
BRAND_ORANGE = "E87722"       # naranja del NA del logo (acentos, totales)
BRAND_ORANGE_LIGHT = "FCE1C7" # tint claro (fondo fila total)
BRAND_ORANGE_DEEP = "B85F1A"  # sombra del naranja (borde fila total)
HEADER_FG = "FFFFFF"
GRID_LINE = "8C8C8C"          # bordes de celda visibles (gris medio)
GRID_LINE_LIGHT = "BFBFBF"
TEXT_MUTED = "595959"

# Semaforo (fills fuertes tipo Excel condicional clasico)
GOOD = "63BE7B"
GOOD_TEXT = "078B10"
WARN = "FFEB84"
WARN_TEXT = "D2A000"
BAD = "F8696B"
BAD_TEXT = "E90561"

# Rangos del semaforo de cumplimiento (fracciones, no porcentajes).
SEMAFORO = {
    "verde_min": 0.95,
    "verde_max": 1.05,
    "amarillo_min": 0.85,
    "amarillo_max": 1.15,
}

# Excel Table style nativa con bordes marcados. TableStyleMedium9 = azul con
# bordes oscuros, primera fila resaltada.
DEFAULT_TABLE_STYLE = "TableStyleMedium9"

# Sentinel para "no hay valor previo aun" (distinto de None valido).
_UNSET: object = object()


# ---------------------------------------------------------------------------
# Registro de NamedStyle
# ---------------------------------------------------------------------------
_STYLE_NAMES = (
    "st_title",
    "st_subtitle",
    "st_header",
    "st_kpi_header",
    "st_body_int",
    "st_body_pct",
    "st_body_text",
    "st_body_date",
    "st_total_int",
    "st_total_pct",
    "st_total_label",
    "st_kpi_label",
    "st_kpi_value_int",
    "st_kpi_value_pct_good",
    "st_kpi_value_pct_warn",
    "st_kpi_value_pct_bad",
)


def _side(style: str = "thin", color: str = GRID_LINE) -> Side:
    return Side(style=style, color=color)


def _border_all(style: str = "thin", color: str = GRID_LINE) -> Border:
    s = _side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)


def _make_named_styles() -> list[NamedStyle]:
    body_border = _border_all("thin", GRID_LINE)
    header_border = Border(
        left=_side("thin", BRAND_NAVY_DEEP),
        right=_side("thin", BRAND_NAVY_DEEP),
        top=_side("medium", BRAND_NAVY_DEEP),
        bottom=_side("medium", BRAND_NAVY_DEEP),
    )
    total_border = Border(
        left=_side("thin", BRAND_ORANGE_DEEP),
        right=_side("thin", BRAND_ORANGE_DEEP),
        top=_side("medium", BRAND_ORANGE_DEEP),
        bottom=_side("medium", BRAND_ORANGE_DEEP),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    styles: list[NamedStyle] = []

    st_title = NamedStyle(name="st_title")
    st_title.font = Font(name="Calibri", size=18, bold=True, color=BRAND_NAVY)
    st_title.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(st_title)

    st_subtitle = NamedStyle(name="st_subtitle")
    st_subtitle.font = Font(name="Calibri", size=11, italic=True, color=TEXT_MUTED)
    st_subtitle.alignment = Alignment(horizontal="left", vertical="center")
    styles.append(st_subtitle)

    st_header = NamedStyle(name="st_header")
    st_header.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    st_header.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    st_header.alignment = center
    st_header.border = header_border
    styles.append(st_header)

    st_kpi_header = NamedStyle(name="st_kpi_header")
    st_kpi_header.font = Font(name="Calibri", size=12, bold=True, color=HEADER_FG)
    st_kpi_header.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    st_kpi_header.alignment = center
    st_kpi_header.border = header_border
    styles.append(st_kpi_header)

    st_body_int = NamedStyle(name="st_body_int")
    st_body_int.font = Font(name="Calibri", size=11, color=BRAND_NAVY_DEEP)
    st_body_int.number_format = "#,##0"
    st_body_int.alignment = right
    st_body_int.border = body_border
    styles.append(st_body_int)

    st_body_pct = NamedStyle(name="st_body_pct")
    st_body_pct.font = Font(name="Calibri", size=11, color=BRAND_NAVY_DEEP)
    st_body_pct.number_format = "0.00%"
    st_body_pct.alignment = right
    st_body_pct.border = body_border
    styles.append(st_body_pct)

    st_body_text = NamedStyle(name="st_body_text")
    st_body_text.font = Font(name="Calibri", size=11, color=BRAND_NAVY_DEEP)
    st_body_text.alignment = left
    st_body_text.border = body_border
    styles.append(st_body_text)

    st_body_date = NamedStyle(name="st_body_date")
    st_body_date.font = Font(name="Calibri", size=11, color=BRAND_NAVY_DEEP)
    st_body_date.number_format = "yyyy-mm-dd"
    st_body_date.alignment = center
    st_body_date.border = body_border
    styles.append(st_body_date)

    st_total_int = NamedStyle(name="st_total_int")
    st_total_int.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY_DEEP)
    st_total_int.fill = PatternFill("solid", fgColor=BRAND_ORANGE_LIGHT)
    st_total_int.number_format = "#,##0"
    st_total_int.alignment = right
    st_total_int.border = total_border
    styles.append(st_total_int)

    st_total_pct = NamedStyle(name="st_total_pct")
    st_total_pct.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY_DEEP)
    st_total_pct.fill = PatternFill("solid", fgColor=BRAND_ORANGE_LIGHT)
    st_total_pct.number_format = "0.00%"
    st_total_pct.alignment = right
    st_total_pct.border = total_border
    styles.append(st_total_pct)

    st_total_label = NamedStyle(name="st_total_label")
    st_total_label.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY_DEEP)
    st_total_label.fill = PatternFill("solid", fgColor=BRAND_ORANGE_LIGHT)
    st_total_label.alignment = left
    st_total_label.border = total_border
    styles.append(st_total_label)

    st_kpi_label = NamedStyle(name="st_kpi_label")
    st_kpi_label.font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY_DEEP)
    st_kpi_label.fill = PatternFill("solid", fgColor="F2F2F2")
    st_kpi_label.alignment = left
    st_kpi_label.border = body_border
    styles.append(st_kpi_label)

    st_kpi_value_int = NamedStyle(name="st_kpi_value_int")
    st_kpi_value_int.font = Font(name="Calibri", size=12, bold=True, color=BRAND_NAVY_DEEP)
    st_kpi_value_int.number_format = "#,##0"
    st_kpi_value_int.alignment = right
    st_kpi_value_int.border = body_border
    styles.append(st_kpi_value_int)

    for tier, bg, fg in (
        ("good", GOOD, GOOD_TEXT),
        ("warn", WARN, WARN_TEXT),
        ("bad", BAD, BAD_TEXT),
    ):
        st = NamedStyle(name=f"st_kpi_value_pct_{tier}")
        st.font = Font(name="Calibri", size=14, bold=True, color=fg)
        st.fill = PatternFill("solid", fgColor=bg)
        st.number_format = "0.00%"
        st.alignment = Alignment(horizontal="center", vertical="center")
        st.border = _border_all("medium", BRAND_NAVY_DEEP)
        styles.append(st)

    return styles


def register_styles(wb: Workbook) -> None:
    """Registra los `NamedStyle` una sola vez en el workbook. Idempotente.

    En openpyxl >= 3.0, `wb.named_styles` retorna una lista de nombres
    (strings), no de objetos. Por eso comparamos por string.
    """
    existing = set(wb.named_styles)
    for style in _make_named_styles():
        if style.name not in existing:
            wb.add_named_style(style)


# ---------------------------------------------------------------------------
# Bloques de titulo y paginacion
# ---------------------------------------------------------------------------
def add_title_block(
    ws: Worksheet,
    title: str,
    subtitle: str,
    span_cols: int = 8,
    start_row: int = 1,
) -> int:
    """Escribe titulo y subtitulo con merge y estilos. Deja fila vacia debajo.

    Retorna el numero de fila donde debe empezar la tabla siguiente.
    """
    end_col_letter = get_column_letter(span_cols)

    ws.row_dimensions[start_row].height = 30
    ws.cell(row=start_row, column=1, value=title).style = "st_title"
    ws.merge_cells(f"A{start_row}:{end_col_letter}{start_row}")

    ws.row_dimensions[start_row + 1].height = 18
    ws.cell(row=start_row + 1, column=1, value=subtitle).style = "st_subtitle"
    ws.merge_cells(f"A{start_row + 1}:{end_col_letter}{start_row + 1}")

    ws.row_dimensions[start_row + 2].height = 8
    return start_row + 3


def set_page_setup(ws: Worksheet, freeze_from_row: int = 5) -> None:
    """Landscape, fit-to-page, gridlines ocultas (los bordes son propios),
    zoom 90, freeze panes.
    """
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.freeze_panes = f"A{freeze_from_row}"
    ws.print_options.gridLines = False


# ---------------------------------------------------------------------------
# Escritura de DataFrames como Excel Table nativa
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ColumnSpec:
    """Descriptor de columna para elegir el NamedStyle correcto."""

    name: str        # nombre visible en el header
    key: str         # clave en el DataFrame
    kind: str = "text"   # int | pct | text | date

    @property
    def body_style(self) -> str:
        return {
            "int": "st_body_int",
            "pct": "st_body_pct",
            "date": "st_body_date",
        }.get(self.kind, "st_body_text")

    @property
    def total_style(self) -> str:
        return {
            "int": "st_total_int",
            "pct": "st_total_pct",
        }.get(self.kind, "st_total_label")


def _sanitize_table_name(name: str) -> str:
    """Excel Table names: alfanum + underscore, deben empezar con letra."""
    cleaned = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = "T_" + cleaned
    return cleaned


def write_dataframe_as_table(
    ws: Worksheet,
    df: pd.DataFrame,
    columns: Sequence[ColumnSpec],
    table_name: str,
    start_row: int,
    start_col: int = 1,
    table_style: str = DEFAULT_TABLE_STYLE,
    group_by_key: str | None = None,
) -> Table:
    """Escribe headers + valores del DataFrame y crea una Excel Table.

    La Excel Table nativa (ListObject) le da al usuario:
    - Filtros y ordenamiento automatico.
    - Banded rows que sobreviven al filtrado.
    - Referencias por nombre en formulas (`=SUMA(Detalle_Material[real])`).
    - Expansion automatica al agregar filas.

    **Ademas** aplicamos bordes explicitos en cada celda del rango de datos
    para que los limites de tabla sean 100% visibles aunque el usuario
    cambie el TableStyle. Los porcentajes se guardan como fraccion (0.85),
    no como 85; el formato `"0.00%"` los renderiza como "85.00%".

    **`group_by_key`** (opcional): agrupa visualmente filas consecutivas con
    el mismo valor en esa columna:
    - La celda del grupo muestra el valor solo en la PRIMERA fila del grupo
      y queda vacia en las siguientes (efecto merge sin romper el
      ListObject ni los filtros de Excel).
    - En el cambio de grupo se aplica un borde medium naranja arriba de
      TODAS las celdas de la fila para separar secciones visualmente.
    - Sin efecto si `df` tiene <= 1 fila o si el key no aparece en columns.
    """
    n_rows = len(df)
    n_cols = len(columns)
    end_col_letter = get_column_letter(start_col + n_cols - 1)

    for j, col in enumerate(columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=col.name)
        cell.style = "st_header"
    ws.row_dimensions[start_row].height = 28

    group_col_idx = None
    if group_by_key:
        for j, col in enumerate(columns):
            if col.key == group_by_key:
                group_col_idx = j
                break

    separator_border = Border(
        left=_side("thin", GRID_LINE),
        right=_side("thin", GRID_LINE),
        top=_side("medium", BRAND_ORANGE_DEEP),
        bottom=_side("thin", GRID_LINE),
    )

    previous_group_value: object = _UNSET
    for i in range(n_rows):
        row_num = start_row + 1 + i
        row = df.iloc[i]
        current_group_value = row.get(group_by_key) if group_col_idx is not None else None
        is_new_group = (
            group_col_idx is not None
            and (previous_group_value is _UNSET or current_group_value != previous_group_value)
        )
        for j, col in enumerate(columns):
            value = row.get(col.key)
            if pd.isna(value):
                value = None
            elif col.kind == "int":
                value = int(round(float(value)))
            elif col.kind == "pct":
                value = float(value)
            elif col.kind == "date":
                if isinstance(value, str):
                    value = pd.to_datetime(value).date()
                elif hasattr(value, "date"):
                    try:
                        value = value.date()
                    except TypeError:
                        pass
            # Modo grouping: en la columna del grupo, mostrar valor solo la
            # primera fila del grupo.
            if group_col_idx is not None and j == group_col_idx and not is_new_group:
                value = None
            cell = ws.cell(row=row_num, column=start_col + j, value=value)
            cell.style = col.body_style
            if is_new_group and i > 0:
                cell.border = separator_border
        if group_col_idx is not None:
            previous_group_value = current_group_value
        ws.row_dimensions[row_num].height = 20

    if n_rows == 0:
        ws.cell(row=start_row + 1, column=start_col, value="(sin datos)").style = "st_body_text"
        end_row = start_row + 1
    else:
        end_row = start_row + n_rows

    ref = (
        f"{get_column_letter(start_col)}{start_row}:"
        f"{end_col_letter}{end_row}"
    )
    table = Table(displayName=_sanitize_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return table


# ---------------------------------------------------------------------------
# Fila TOTAL (opcional, va DEBAJO de la Excel Table, no dentro)
# ---------------------------------------------------------------------------
def add_total_row(
    ws: Worksheet,
    columns: Sequence[ColumnSpec],
    values: dict[str, float | int | str | None],
    row_num: int,
    start_col: int = 1,
    label_col_key: str | None = None,
    label_text: str = "TOTAL",
) -> None:
    """Escribe una fila TOTAL con estilos `st_total_*`.

    La fila NO forma parte de la Excel Table (queda debajo). Esto es a
    proposito: si formara parte, el filtrado la ocultaria.
    """
    label_key = label_col_key or columns[0].key
    for j, col in enumerate(columns):
        if col.key == label_key:
            ws.cell(row=row_num, column=start_col + j, value=label_text).style = "st_total_label"
        else:
            v = values.get(col.key)
            if v is not None and col.kind == "int":
                v = int(round(float(v)))
            elif v is not None and col.kind == "pct":
                v = float(v)
            ws.cell(row=row_num, column=start_col + j, value=v).style = col.total_style
    ws.row_dimensions[row_num].height = 24


# ---------------------------------------------------------------------------
# Semaforo condicional (verde/amarillo/rojo) sobre columna de cumplimiento_pct
# ---------------------------------------------------------------------------
def apply_traffic_light(
    ws: Worksheet,
    col_letter: str,
    from_row: int,
    to_row: int,
) -> None:
    """Aplica 3 reglas de formato condicional sobre `col_letter[from_row:to_row]`.

    Las celdas se guardan como fraccion (0.95 = 95%). Los umbrales usan
    `SEMAFORO` del propio modulo.

    - verde: `[verde_min, verde_max]` => fondo `GOOD`, texto `GOOD_TEXT`
    - amarillo: `[amarillo_min, verde_min)` o `(verde_max, amarillo_max]`
      => fondo `WARN`, texto `WARN_TEXT`
    - rojo: `< amarillo_min` o `> amarillo_max`
      => fondo `BAD`, texto `BAD_TEXT`
    """
    if to_row < from_row:
        return
    rng = f"{col_letter}{from_row}:{col_letter}{to_row}"

    green_fill = PatternFill("solid", fgColor=GOOD)
    green_font = Font(color=GOOD_TEXT, bold=True)
    warn_fill = PatternFill("solid", fgColor=WARN)
    warn_font = Font(color=WARN_TEXT, bold=True)
    bad_fill = PatternFill("solid", fgColor=BAD)
    bad_font = Font(color=BAD_TEXT, bold=True)

    v_min = str(SEMAFORO["verde_min"])
    v_max = str(SEMAFORO["verde_max"])
    a_min = str(SEMAFORO["amarillo_min"])
    a_max = str(SEMAFORO["amarillo_max"])

    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="between", formula=[v_min, v_max],
                   stopIfTrue=True, fill=green_fill, font=green_font),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="between", formula=[a_min, a_max],
                   stopIfTrue=True, fill=warn_fill, font=warn_font),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="notBetween", formula=[a_min, a_max],
                   stopIfTrue=True, fill=bad_fill, font=bad_font),
    )


# ---------------------------------------------------------------------------
# KPI table (label | valor) — reemplaza el card floating
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class KpiRow:
    label: str
    valor: float | int
    kind: str = "int"  # int | pct
    semaforo: bool = False  # si True y kind==pct, aplica color por rango


def _semaforo_tier(valor: float) -> str:
    if SEMAFORO["verde_min"] <= valor <= SEMAFORO["verde_max"]:
        return "good"
    if SEMAFORO["amarillo_min"] <= valor <= SEMAFORO["amarillo_max"]:
        return "warn"
    return "bad"


def write_kpi_table(
    ws: Worksheet,
    kpis: Sequence[KpiRow],
    start_row: int,
    start_col: int = 1,
    label_col_width: int = 42,
    value_col_width: int = 20,
) -> int:
    """Escribe una tabla vertical de indicadores (label | valor) con bordes.

    Cada indicador es una fila. El header ("Indicador | Valor") va en la
    primera fila. Retorna el numero de fila siguiente al final de la tabla.
    """
    left_col = get_column_letter(start_col)
    right_col = get_column_letter(start_col + 1)

    ws.column_dimensions[left_col].width = label_col_width
    ws.column_dimensions[right_col].width = value_col_width

    ws.cell(row=start_row, column=start_col, value="Indicador").style = "st_kpi_header"
    ws.cell(row=start_row, column=start_col + 1, value="Valor").style = "st_kpi_header"
    ws.row_dimensions[start_row].height = 26

    for i, kpi in enumerate(kpis, start=1):
        row = start_row + i
        ws.cell(row=row, column=start_col, value=kpi.label).style = "st_kpi_label"
        cell = ws.cell(row=row, column=start_col + 1, value=float(kpi.valor))
        if kpi.kind == "pct" and kpi.semaforo:
            cell.style = f"st_kpi_value_pct_{_semaforo_tier(float(kpi.valor))}"
        elif kpi.kind == "pct":
            cell.style = "st_body_pct"
        else:
            cell.style = "st_kpi_value_int"
        ws.row_dimensions[row].height = 22

    return start_row + len(kpis) + 1


# ---------------------------------------------------------------------------
# Logo corporativo
# ---------------------------------------------------------------------------
def insert_logo(
    ws: Worksheet,
    logo_path: Path | str,
    anchor: str = "A1",
    max_height_px: int = 60,
) -> OpenpyxlImage | None:
    """Inserta el logo NutriAvicola en la celda `anchor` escalado a `max_height_px`.

    Retorna el objeto Image o None si el archivo no existe. Silencioso: en
    entornos sin Pillow o sin el logo, la portada sigue funcionando sin
    imagen (solo aparece el titulo).
    """
    p = Path(logo_path)
    if not p.exists():
        return None
    try:
        img = OpenpyxlImage(str(p))
    except Exception:
        return None
    if img.height and img.height > max_height_px:
        ratio = max_height_px / img.height
        img.height = int(img.height * ratio)
        img.width = int(img.width * ratio)
    img.anchor = anchor
    ws.add_image(img)
    return img


# ---------------------------------------------------------------------------
# Autofit de columnas
# ---------------------------------------------------------------------------
def autofit_columns(
    ws: Worksheet,
    columns: Iterable[int] | None = None,
    max_width: int = 40,
    min_width: int = 10,
    padding: int = 2,
) -> None:
    """Ajusta ancho de columnas al maximo contenido, capado a `max_width`.

    Ignora celdas mergeadas (openpyxl `MergedCell` no tiene `.value` unico
    accesible desde la columna). Solo mira celdas normales.
    """
    from openpyxl.cell import MergedCell

    if columns is None:
        columns = range(1, ws.max_column + 1)

    for col_idx in columns:
        letter = get_column_letter(col_idx)
        max_len = min_width
        for cell in ws[letter]:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is None:
                continue
            if isinstance(cell.value, float):
                text = f"{cell.value:,.2f}"
            else:
                text = str(cell.value)
            for line in text.splitlines() or [""]:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[letter].width = min(max_len + padding, max_width)
