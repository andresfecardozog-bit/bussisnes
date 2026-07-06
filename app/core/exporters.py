"""Exportadores CSV/XLSX del historico y de los runs individuales.

Escenarios cubiertos:

- `export_run_summary`: escribe `data/runs/{run_id}/summary.csv` con una fila
  por material del cruce. Consumido por Power Automate cuando quiere ver
  "un solo run" sin abrir la DB.
- `export_monthly_view`: consulta la tabla `cruce` y arma la vista mensual
  (una fila por dia). Sale como .csv y .xlsx en `data/onedrive_export/`.
- `export_cumplimiento_xlsx`: **entregable oficial de la Fase 6.** Genera
  un `.xlsx` bien formateado (paleta corporativa, Excel Tables, semaforo)
  con 5 hojas: Portada, Resumen, Por_Categoria, Detalle_Material,
  No_Cruzados. Es lo que el equipo de BI descarga y usa como fuente para
  armar sus dashboards en Power BI.

Regla: este modulo **no** puede importar `Font`, `PatternFill`, `Border`,
`Alignment` ni `Side` de openpyxl. Toda la estilizacion se delega en
`app.core.excel_style`. Un test verifica que este contrato se cumpla.
"""
from __future__ import annotations

import sqlite3
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app.config import DB_PATH, ONEDRIVE_EXPORT_DIR, RUNS_DIR
from app.core import excel_style as xs
from app.core.excel_style import ColumnSpec, KpiRow

# Orden fijo de categorias que aparecen en RESUMEN (filas 39-43 del xlsx).
CATEGORIAS_TIPO = ("A", "AA", "AAA", "AAAA", "B", "C")

# Logo corporativo NutriAvicola en `resources/`.
LOGO_PATH = Path(__file__).resolve().parents[2] / "resources" / "image_720508810_0.jpg"


# ---------------------------------------------------------------------------
# Exports historicos previos (usados por Power Automate y consultas puntuales)
# ---------------------------------------------------------------------------
def export_run_summary(
    matched_df: pd.DataFrame,
    run_id: str,
    fecha_produccion: date,
    output_dir: Path | None = None,
) -> Path:
    """Escribe summary CSV del cruce para un run especifico."""
    output_dir = output_dir or (RUNS_DIR / run_id)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"summary_{fecha_produccion:%Y%m%d}.csv"
    matched_df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def export_monthly_view(
    year: int,
    month: int,
    output_dir: Path | None = None,
    db_path: str | Path = DB_PATH,
) -> tuple[Path, Path]:
    """Exporta vista mensual del cruce a CSV + XLSX en OneDrive folder."""
    output_dir = output_dir or ONEDRIVE_EXPORT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        query = """
            SELECT
                fecha_produccion, material, referencia, nomb_material_flash,
                notificado_unidades, producir_unidades, real_unidades_flash,
                delta_unidades, cumplimiento_pct, match_bool
            FROM cruce
            WHERE strftime('%Y', fecha_produccion) = ?
              AND strftime('%m', fecha_produccion) = ?
            ORDER BY fecha_produccion, material
        """
        df = pd.read_sql_query(
            query,
            conn,
            params=(str(year), f"{month:02d}"),
        )

    stem = f"cruce_mensual_{year}{month:02d}"
    csv_path = output_dir / f"{stem}.csv"
    xlsx_path = output_dir / f"{stem}.xlsx"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_excel(xlsx_path, index=False, sheet_name=f"{year}-{month:02d}")
    return csv_path, xlsx_path


# ---------------------------------------------------------------------------
# Fase 6: cumplimiento_xxx.xlsx (entregable para el equipo de BI)
# ---------------------------------------------------------------------------
def _fecha_iso(d: date | str) -> str:
    if isinstance(d, str):
        return d
    return d.isoformat()


def _fetch_cruce(
    conn: sqlite3.Connection, desde: date, hasta: date
) -> pd.DataFrame:
    """Trae cruce enriquecido con `tipo`, `formato`, `unidades_por_empaque`.

    El join a `sku_catalog` usa `material` -> `material_sap`. Un material
    puede aparecer varias veces en el catalogo si se aprendio por distintas
    rutas; usamos MAX() para deduplicar sin cambiar el material.
    """
    query = """
        SELECT
            c.fecha_produccion, c.material,
            COALESCE(c.referencia, cat.referencia)                AS referencia,
            cat.tipo                                              AS tipo,
            cat.formato                                           AS formato,
            cat.unidades_por_empaque                              AS unidades_por_empaque,
            c.notificado_unidades                                 AS plan,
            c.real_unidades_flash                                 AS real,
            c.delta_unidades                                      AS delta,
            c.cumplimiento_pct                                    AS cumplimiento_pct
        FROM cruce c
        LEFT JOIN (
            SELECT material_sap, MAX(referencia) AS referencia,
                   MAX(tipo) AS tipo, MAX(formato) AS formato,
                   MAX(unidades_por_empaque) AS unidades_por_empaque
            FROM sku_catalog GROUP BY material_sap
        ) cat ON cat.material_sap = c.material
        WHERE c.fecha_produccion BETWEEN ? AND ?
        ORDER BY c.fecha_produccion, c.material
    """
    df = pd.read_sql_query(query, conn, params=(_fecha_iso(desde), _fecha_iso(hasta)))
    if not df.empty:
        df["fecha_produccion"] = pd.to_datetime(df["fecha_produccion"]).dt.date
        df["cumplimiento_pct"] = df["cumplimiento_pct"] / 100.0
    return df


def _fetch_no_cruzados(
    conn: sqlite3.Connection, desde: date, hasta: date
) -> pd.DataFrame:
    query = """
        SELECT
            nc.fecha_produccion, nc.origen, nc.material,
            nc.referencia_o_nombre AS referencia, nc.valor AS unidades,
            nc.motivo
        FROM no_cruzados nc
        WHERE nc.fecha_produccion BETWEEN ? AND ?
        ORDER BY nc.fecha_produccion, nc.origen, nc.material
    """
    df = pd.read_sql_query(query, conn, params=(_fecha_iso(desde), _fecha_iso(hasta)))
    if not df.empty:
        df["fecha_produccion"] = pd.to_datetime(df["fecha_produccion"]).dt.date
    return df


def _resumen_totales(cruce: pd.DataFrame) -> pd.DataFrame:
    """Agrupa el cruce por fecha_produccion para armar la hoja Resumen.

    Retorna columnas fecha_produccion, plan_total, real_total, delta_total,
    cumplimiento_pct.
    """
    if cruce.empty:
        return pd.DataFrame(columns=[
            "fecha_produccion", "plan_total", "real_total",
            "delta_total", "cumplimiento_pct",
        ])
    grouped = (
        cruce.groupby("fecha_produccion", dropna=False)
        .agg(plan_total=("plan", "sum"), real_total=("real", "sum"))
        .reset_index()
    )
    grouped["delta_total"] = grouped["real_total"] - grouped["plan_total"]
    grouped["cumplimiento_pct"] = grouped.apply(
        lambda r: (r["real_total"] / r["plan_total"]) if r["plan_total"] else 0.0,
        axis=1,
    )
    return grouped


def _resumen_por_categoria(cruce: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por (fecha_produccion, tipo). Rellena tipos faltantes con 0.

    Garantiza que las 6 categorias aparezcan para cada fecha (con plan=0 y
    real=0 si no hubo actividad). Esto ayuda al equipo de BI a hacer pivots
    consistentes sin tener que rellenar valores faltantes.
    """
    if cruce.empty:
        return pd.DataFrame(columns=[
            "fecha_produccion", "tipo", "plan_categoria",
            "real_categoria", "delta_categoria", "cumplimiento_pct",
        ])
    df = cruce.copy()
    df["tipo"] = df["tipo"].fillna("SIN_TIPO")
    grouped = (
        df.groupby(["fecha_produccion", "tipo"], dropna=False)
        .agg(plan_categoria=("plan", "sum"), real_categoria=("real", "sum"))
        .reset_index()
    )

    fechas = sorted(grouped["fecha_produccion"].unique())
    tipos_reales = sorted(set(grouped["tipo"].unique()) | set(CATEGORIAS_TIPO))
    idx = pd.MultiIndex.from_product([fechas, tipos_reales],
                                     names=["fecha_produccion", "tipo"])
    grouped = grouped.set_index(["fecha_produccion", "tipo"]).reindex(idx, fill_value=0.0).reset_index()

    grouped["delta_categoria"] = grouped["real_categoria"] - grouped["plan_categoria"]
    grouped["cumplimiento_pct"] = grouped.apply(
        lambda r: (r["real_categoria"] / r["plan_categoria"])
        if r["plan_categoria"] else 0.0,
        axis=1,
    )
    return grouped.sort_values(["fecha_produccion", "tipo"]).reset_index(drop=True)


def _resumen_por_semana(cruce: pd.DataFrame) -> pd.DataFrame:
    """Agrupa por semana ISO. Retorna una fila por semana con dias habiles y KPIs.

    - `semana_label`: "S06 2026" para ordenamiento visual estable.
    - `fecha_inicio` / `fecha_fin`: primer y ultimo dia laboral observado.
    - `dias_habiles`: cuantas fechas distintas de la semana tuvieron cruce.
    - `plan_total`, `real_total`, `delta_total`, `cumplimiento_pct`.
    """
    cols = [
        "semana_label", "fecha_inicio", "fecha_fin", "dias_habiles",
        "plan_total", "real_total", "delta_total", "cumplimiento_pct",
    ]
    if cruce.empty:
        return pd.DataFrame(columns=cols)

    df = cruce.copy()
    fechas_dt = pd.to_datetime(df["fecha_produccion"])
    iso = fechas_dt.dt.isocalendar()
    df["_year"] = iso.year.astype(int)
    df["_week"] = iso.week.astype(int)

    grouped = (
        df.groupby(["_year", "_week"])
        .agg(
            fecha_inicio=("fecha_produccion", "min"),
            fecha_fin=("fecha_produccion", "max"),
            dias_habiles=("fecha_produccion", "nunique"),
            plan_total=("plan", "sum"),
            real_total=("real", "sum"),
        )
        .reset_index()
    )
    grouped["semana_label"] = grouped.apply(
        lambda r: f"S{int(r['_week']):02d} {int(r['_year'])}", axis=1
    )
    grouped["delta_total"] = grouped["real_total"] - grouped["plan_total"]
    grouped["cumplimiento_pct"] = grouped.apply(
        lambda r: (r["real_total"] / r["plan_total"]) if r["plan_total"] else 0.0,
        axis=1,
    )
    grouped = grouped.sort_values(["_year", "_week"]).reset_index(drop=True)
    return grouped[cols]


def _detalle_material(cruce: pd.DataFrame) -> pd.DataFrame:
    if cruce.empty:
        return pd.DataFrame(columns=[
            "fecha_produccion", "material", "referencia", "tipo", "formato",
            "unidades_por_empaque", "plan", "real", "delta", "cumplimiento_pct",
        ])
    df = cruce[[
        "fecha_produccion", "material", "referencia", "tipo", "formato",
        "unidades_por_empaque", "plan", "real", "delta", "cumplimiento_pct",
    ]].copy()
    return df.sort_values(["fecha_produccion", "material"]).reset_index(drop=True)


def _no_cruzados_sheet(no_cruz: pd.DataFrame) -> pd.DataFrame:
    if no_cruz.empty:
        return pd.DataFrame(columns=[
            "fecha_produccion", "origen", "material", "referencia",
            "unidades", "motivo",
        ])
    df = no_cruz[[
        "fecha_produccion", "origen", "material", "referencia",
        "unidades", "motivo",
    ]].copy()
    return df.sort_values(["fecha_produccion", "origen", "material"]).reset_index(drop=True)


def _tiene_multiples_fechas(df: pd.DataFrame) -> bool:
    return "fecha_produccion" in df.columns and df["fecha_produccion"].nunique() > 1


# ---------- construccion de hojas ----------

def _write_portada(
    wb: Workbook,
    desde: date,
    hasta: date,
    resumen: pd.DataFrame,
    detalle: pd.DataFrame,
    no_cruz_count: int,
) -> None:
    """Portada corporativa: logo + titulo + tabla de indicadores. Sin literatura."""
    ws = wb.create_sheet("Portada")

    plan_total = float(resumen["plan_total"].sum()) if not resumen.empty else 0.0
    real_total = float(resumen["real_total"].sum()) if not resumen.empty else 0.0
    cumplimiento_global = (real_total / plan_total) if plan_total else 0.0

    # Logo en A1 + titulo/subtitulo desde la fila 4 para no chocar con el logo.
    xs.insert_logo(ws, LOGO_PATH, anchor="A1", max_height_px=60)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    xs.add_title_block(
        ws,
        "Cumplimiento PRE CORTE vs FLASH",
        f"Rango: {desde:%d/%m/%Y}  a  {hasta:%d/%m/%Y}"
        f"      Generado: {datetime.now():%d/%m/%Y %H:%M}",
        span_cols=4,
        start_row=4,
    )

    delta_total = real_total - plan_total
    dias = len(resumen)
    materiales = len(detalle)

    kpis = (
        KpiRow("Cumplimiento global (%)", cumplimiento_global, "pct", semaforo=True),
        KpiRow("Plan total (unidades)", plan_total, "int"),
        KpiRow("Real total (unidades)", real_total, "int"),
        KpiRow("Delta (real - plan)", delta_total, "int"),
        KpiRow("Dias en el rango", dias, "int"),
        KpiRow("Materiales cruzados (SKUs)", materiales, "int"),
        KpiRow("Filas no cruzadas", no_cruz_count, "int"),
    )
    xs.write_kpi_table(ws, kpis, start_row=8, start_col=1)

    xs.set_page_setup(ws, freeze_from_row=8)


def _write_resumen(
    wb: Workbook, desde: date, hasta: date, df: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Resumen")
    columns: Sequence[ColumnSpec] = (
        ColumnSpec("Fecha produccion", "fecha_produccion", "date"),
        ColumnSpec("Plan total (unid.)", "plan_total", "int"),
        ColumnSpec("Real total (unid.)", "real_total", "int"),
        ColumnSpec("Delta (real - plan)", "delta_total", "int"),
        ColumnSpec("Cumplimiento %", "cumplimiento_pct", "pct"),
    )
    xs.add_title_block(
        ws,
        "Resumen diario",
        f"Cumplimiento total de huevos por dia  -  {desde:%d/%m/%Y} a {hasta:%d/%m/%Y}",
        span_cols=len(columns),
    )
    start_row = 4
    xs.write_dataframe_as_table(ws, df, columns, "Resumen",
                                start_row=start_row)

    n_rows = max(len(df), 1)
    cumpl_col_idx = len(columns)
    cumpl_col_letter = get_column_letter(cumpl_col_idx)
    xs.apply_traffic_light(
        ws, cumpl_col_letter,
        from_row=start_row + 1,
        to_row=start_row + n_rows,
    )

    if not df.empty:
        total_row = start_row + n_rows + 1
        plan = float(df["plan_total"].sum())
        real = float(df["real_total"].sum())
        cumpl = (real / plan) if plan else 0.0
        xs.add_total_row(
            ws, columns,
            values={
                "fecha_produccion": None,
                "plan_total": plan,
                "real_total": real,
                "delta_total": real - plan,
                "cumplimiento_pct": cumpl,
            },
            row_num=total_row,
            label_col_key="fecha_produccion",
            label_text="TOTAL",
        )

    xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def _write_por_categoria(
    wb: Workbook, desde: date, hasta: date, df: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Por_Categoria")
    columns: Sequence[ColumnSpec] = (
        ColumnSpec("Fecha produccion", "fecha_produccion", "date"),
        ColumnSpec("Tipo (categoria)", "tipo", "text"),
        ColumnSpec("Plan categoria (unid.)", "plan_categoria", "int"),
        ColumnSpec("Real categoria (unid.)", "real_categoria", "int"),
        ColumnSpec("Delta (real - plan)", "delta_categoria", "int"),
        ColumnSpec("Cumplimiento %", "cumplimiento_pct", "pct"),
    )
    xs.add_title_block(
        ws,
        "Cumplimiento por categoria de huevo",
        f"Filas por fecha x tipo (A, AA, AAA, AAAA, B, C)"
        f"  -  {desde:%d/%m/%Y} a {hasta:%d/%m/%Y}",
        span_cols=len(columns),
    )
    start_row = 4
    group_key = "fecha_produccion" if _tiene_multiples_fechas(df) else None
    xs.write_dataframe_as_table(ws, df, columns, "Por_Categoria",
                                start_row=start_row, group_by_key=group_key)

    n_rows = max(len(df), 1)
    cumpl_col_idx = len(columns)
    cumpl_col_letter = get_column_letter(cumpl_col_idx)
    xs.apply_traffic_light(
        ws, cumpl_col_letter,
        from_row=start_row + 1,
        to_row=start_row + n_rows,
    )

    xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def _write_detalle_material(
    wb: Workbook, desde: date, hasta: date, df: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Detalle_Material")
    columns: Sequence[ColumnSpec] = (
        ColumnSpec("Fecha produccion", "fecha_produccion", "date"),
        ColumnSpec("Material SAP", "material", "int"),
        ColumnSpec("Referencia", "referencia", "text"),
        ColumnSpec("Tipo", "tipo", "text"),
        ColumnSpec("Formato", "formato", "text"),
        ColumnSpec("Unidades por empaque", "unidades_por_empaque", "int"),
        ColumnSpec("Plan (unid.)", "plan", "int"),
        ColumnSpec("Real (unid.)", "real", "int"),
        ColumnSpec("Delta (real - plan)", "delta", "int"),
        ColumnSpec("Cumplimiento %", "cumplimiento_pct", "pct"),
    )
    xs.add_title_block(
        ws,
        "Detalle por material",
        f"Una fila por SAP y fecha  -  {desde:%d/%m/%Y} a {hasta:%d/%m/%Y}",
        span_cols=len(columns),
    )
    start_row = 4
    group_key = "fecha_produccion" if _tiene_multiples_fechas(df) else None
    xs.write_dataframe_as_table(ws, df, columns, "Detalle_Material",
                                start_row=start_row, group_by_key=group_key)

    n_rows = max(len(df), 1)
    cumpl_col_idx = len(columns)
    cumpl_col_letter = get_column_letter(cumpl_col_idx)
    xs.apply_traffic_light(
        ws, cumpl_col_letter,
        from_row=start_row + 1,
        to_row=start_row + n_rows,
    )

    xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def _write_no_cruzados(
    wb: Workbook, desde: date, hasta: date, df: pd.DataFrame
) -> None:
    ws = wb.create_sheet("No_Cruzados")
    columns: Sequence[ColumnSpec] = (
        ColumnSpec("Fecha produccion", "fecha_produccion", "date"),
        ColumnSpec("Origen", "origen", "text"),
        ColumnSpec("Material SAP", "material", "int"),
        ColumnSpec("Referencia / Nombre", "referencia", "text"),
        ColumnSpec("Unidades", "unidades", "int"),
        ColumnSpec("Motivo", "motivo", "text"),
    )
    xs.add_title_block(
        ws,
        "Filas sin cruce (fugas y solo-plan)",
        f"origen='flash' = vendido sin plan  |  origen='pre_corte' = planeado sin venta"
        f"  -  {desde:%d/%m/%Y} a {hasta:%d/%m/%Y}",
        span_cols=len(columns),
    )
    start_row = 4
    group_key = "fecha_produccion" if _tiene_multiples_fechas(df) else None
    xs.write_dataframe_as_table(ws, df, columns, "No_Cruzados",
                                start_row=start_row, group_by_key=group_key)
    xs.autofit_columns(ws, columns=range(1, len(columns) + 1), max_width=60)
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def _write_por_semana(
    wb: Workbook, desde: date, hasta: date, df: pd.DataFrame
) -> None:
    """Hoja Por_Semana: KPI de cumplimiento agrupado por semana ISO.

    Solo se agrega al consolidado (no a los dailies). Requiere al menos 1
    fecha; con una sola fecha muestra 1 fila (la semana que contiene).
    """
    ws = wb.create_sheet("Por_Semana")
    columns: Sequence[ColumnSpec] = (
        ColumnSpec("Semana", "semana_label", "text"),
        ColumnSpec("Desde", "fecha_inicio", "date"),
        ColumnSpec("Hasta", "fecha_fin", "date"),
        ColumnSpec("Dias habiles", "dias_habiles", "int"),
        ColumnSpec("Plan total (unid.)", "plan_total", "int"),
        ColumnSpec("Real total (unid.)", "real_total", "int"),
        ColumnSpec("Delta (real - plan)", "delta_total", "int"),
        ColumnSpec("Cumplimiento %", "cumplimiento_pct", "pct"),
    )
    xs.add_title_block(
        ws,
        "Resumen semanal",
        f"Cumplimiento agregado por semana ISO  -  {desde:%d/%m/%Y} a {hasta:%d/%m/%Y}",
        span_cols=len(columns),
    )
    start_row = 4
    xs.write_dataframe_as_table(ws, df, columns, "Por_Semana",
                                start_row=start_row)
    n_rows = max(len(df), 1)
    cumpl_col_letter = get_column_letter(len(columns))
    xs.apply_traffic_light(ws, cumpl_col_letter,
                           from_row=start_row + 1,
                           to_row=start_row + n_rows)

    if not df.empty:
        total_row = start_row + n_rows + 1
        plan = float(df["plan_total"].sum())
        real = float(df["real_total"].sum())
        cumpl = (real / plan) if plan else 0.0
        xs.add_total_row(
            ws, columns,
            values={
                "semana_label": None,
                "fecha_inicio": None,
                "fecha_fin": None,
                "dias_habiles": int(df["dias_habiles"].sum()),
                "plan_total": plan,
                "real_total": real,
                "delta_total": real - plan,
                "cumplimiento_pct": cumpl,
            },
            row_num=total_row,
            label_col_key="semana_label",
            label_text="TOTAL",
        )

    xs.autofit_columns(ws, columns=range(1, len(columns) + 1))
    xs.set_page_setup(ws, freeze_from_row=start_row + 1)


def export_cumplimiento_xlsx(
    desde: date,
    hasta: date,
    dest: Path | str,
    conn: sqlite3.Connection | None = None,
    db_path: str | Path = DB_PATH,
) -> Path:
    """Entregable oficial de la Fase 6: `.xlsx` bien formateado para BI.

    Estructura del archivo:
    - Portada:          card con cumplimiento global + informacion del rango.
    - Resumen:          una fila por fecha_produccion (KPI total).
    - Por_Semana:       (solo si hay > 1 fecha) KPI agrupado por semana ISO.
    - Por_Categoria:    filas fecha x tipo (A/AA/AAA/AAAA/B/C), tipos faltantes
                        se rellenan con plan=0/real=0. Con multiples fechas,
                        la columna fecha se agrupa visualmente.
    - Detalle_Material: dump legible del `cruce`, listo para tabla dinamica.
                        Fecha agrupada visualmente si hay multiples.
    - No_Cruzados:      fugas y solo-plan con motivo. Idem agrupacion.

    KPI unico: `cumplimiento_pct = real_flash / plan_resumen`. Se guarda
    como fraccion (0.9821 con formato "0.00%").
    """
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    close_after = False
    if conn is None:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        close_after = True

    try:
        cruce = _fetch_cruce(conn, desde, hasta)
        no_cruz = _fetch_no_cruzados(conn, desde, hasta)

        resumen = _resumen_totales(cruce)
        por_cat = _resumen_por_categoria(cruce)
        detalle = _detalle_material(cruce)
        no_cruz_sheet = _no_cruzados_sheet(no_cruz)
        semanal = _resumen_por_semana(cruce)

        wb = Workbook()
        wb.remove(wb.active)
        xs.register_styles(wb)

        _write_portada(wb, desde, hasta, resumen, detalle, len(no_cruz_sheet))
        _write_resumen(wb, desde, hasta, resumen)
        if len(resumen) > 1:
            _write_por_semana(wb, desde, hasta, semanal)
        _write_por_categoria(wb, desde, hasta, por_cat)
        _write_detalle_material(wb, desde, hasta, detalle)
        _write_no_cruzados(wb, desde, hasta, no_cruz_sheet)

        wb.save(dest)
    finally:
        if close_after:
            conn.close()

    return dest


def export_cumplimiento_diario(
    fecha: date,
    dest: Path | str,
    conn: sqlite3.Connection | None = None,
    db_path: str | Path = DB_PATH,
) -> Path:
    """Genera el `.xlsx` de un unico dia (Fase 6.5).

    Es la version por-dia del consolidado, usada por `export_batch_completo`
    para producir N archivos individuales cuando el usuario procesa un mes
    entero. No lleva hoja Por_Semana (no aporta).
    """
    return export_cumplimiento_xlsx(fecha, fecha, dest, conn=conn, db_path=db_path)


def export_batch_completo(
    desde: date,
    hasta: date,
    output_dir: Path | str,
    conn: sqlite3.Connection | None = None,
    db_path: str | Path = DB_PATH,
    zip_name: str | None = None,
) -> dict[str, object]:
    """Genera todo el paquete de entregables para un batch multi-fecha.

    Produce:
    - N `cumplimiento_YYYYMMDD.xlsx`: uno por cada fecha con datos en el
      rango [desde, hasta]. Fechas sin datos se saltan (no se generan
      archivos vacios).
    - 1 `cumplimiento_consolidado_YYYYMMDD_YYYYMMDD.xlsx`: rango completo
      con la hoja Por_Semana + agrupacion por fecha en Detalle/Categoria/
      NoCruzados.
    - 1 `cumplimiento_batch_YYYYMMDD_YYYYMMDD.zip`: todos los `.xlsx`
      anteriores comprimidos con nombres relativos (sin path del sistema).

    Retorna:
        {
            "consolidado": Path,
            "dailies": [Path, ...],
            "zip": Path,
            "fechas_procesadas": [date, ...],
            "fechas_sin_datos_en_rango": [date, ...],
        }
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    close_after = False
    if conn is None:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        close_after = True

    try:
        cruce = _fetch_cruce(conn, desde, hasta)
        fechas_con_datos = sorted(cruce["fecha_produccion"].unique()) if not cruce.empty else []

        dailies: list[Path] = []
        for f in fechas_con_datos:
            daily_dest = output_dir / suggested_daily_filename(f)
            export_cumplimiento_diario(f, daily_dest, conn=conn)
            dailies.append(daily_dest)

        consolidado_dest = output_dir / suggested_consolidado_filename(desde, hasta)
        export_cumplimiento_xlsx(desde, hasta, consolidado_dest, conn=conn)

        zip_dest = output_dir / (zip_name or suggested_zip_filename(desde, hasta))
        with zipfile.ZipFile(zip_dest, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(consolidado_dest, arcname=consolidado_dest.name)
            for d in dailies:
                zf.write(d, arcname=d.name)
    finally:
        if close_after:
            conn.close()

    dias_rango = _dias_en_rango(desde, hasta)
    sin_datos = [d for d in dias_rango if d not in fechas_con_datos]

    return {
        "consolidado": consolidado_dest,
        "dailies": dailies,
        "zip": zip_dest,
        "fechas_procesadas": list(fechas_con_datos),
        "fechas_sin_datos_en_rango": sin_datos,
    }


def _dias_en_rango(desde: date, hasta: date) -> list[date]:
    """Todos los dias del rango [desde, hasta], sin filtrar por laboral."""
    from datetime import timedelta

    if hasta < desde:
        return []
    out: list[date] = []
    cur = desde
    while cur <= hasta:
        out.append(cur)
        cur = cur + timedelta(days=1)
    return out


def suggested_export_filename(desde: date, hasta: date) -> str:
    """Nombre del archivo consolidado o de rango puntual."""
    if desde == hasta:
        return suggested_daily_filename(desde)
    return suggested_consolidado_filename(desde, hasta)


def suggested_daily_filename(fecha: date) -> str:
    return f"cumplimiento_{fecha:%Y%m%d}.xlsx"


def suggested_consolidado_filename(desde: date, hasta: date) -> str:
    return f"cumplimiento_consolidado_{desde:%Y%m%d}_{hasta:%Y%m%d}.xlsx"


def suggested_zip_filename(desde: date, hasta: date) -> str:
    return f"cumplimiento_batch_{desde:%Y%m%d}_{hasta:%Y%m%d}.zip"
