"""Parser de la hoja RESUMEN del PRE CORTE.

La hoja RESUMEN es una matriz pivoteada disenada para lectura humana:

    - Fila 7: FORMATO (ESTUCHE, VITAFILM, AMARRADO, SUELTO) - celdas fusionadas
      horizontalmente sobre sus sub-columnas.
    - Fila 8: UNIDADES por empaque (6, 12, 15, 20, 30, 45, 60, 180, 1).
    - Filas 9-36: datos. Columna A tiene REFERENCIA fusionada verticalmente
      (por ejemplo A9:A14 = MARCA ORO abarca 6 filas de TIPO A, AA, AAA, AAAA,
      B, C). Columna B tiene TIPO. Cols C-Q tienen las bandejas (no unidades).
    - Fila 37+: totales - se descartan del parseo de detalle.

Este modulo convierte esa matriz en un DataFrame largo con una fila por
(referencia, tipo, formato, unidades_por_empaque) donde bandejas > 0.

El CSV export destruye la fusion de celdas -> **exigimos .xlsx**.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


HEADER_ROW_FORMATO = 7
HEADER_ROW_UNIDADES = 8
DATA_START_ROW = 9
COL_REFERENCIA = 1
COL_TIPO = 2
COL_DATA_START = 3
COL_DATA_END = 17
STOP_MARKERS_COL_A = {"TOTALES", "TOTAL"}


@dataclass
class ResumenColumn:
    """Metadatos de una columna de datos del RESUMEN."""

    excel_col: int
    formato: str
    unidades_por_empaque: int


def _normalize_formato(raw: str) -> str:
    if raw is None:
        return ""
    return str(raw).strip().upper()


def _normalize_tipo(raw: str) -> str:
    if raw is None:
        return ""
    return str(raw).strip().upper().replace(" ", "").replace("/", "-")


def _normalize_referencia(raw: str) -> str:
    if raw is None:
        return ""
    return str(raw).strip().upper()


def _resolve_formato_por_columna(ws: Worksheet) -> dict[int, str]:
    """Para cada columna de datos, retorna el FORMATO efectivo.

    El FORMATO en fila 7 se guarda en la primera columna de su rango y las
    demas quedan None por fusion. Se propaga hacia la derecha hasta encontrar
    el siguiente valor no vacio (o cambio de bloque).
    """
    formatos: dict[int, str] = {}
    ultimo_formato: str | None = None
    for c in range(COL_DATA_START, COL_DATA_END + 1):
        val = ws.cell(HEADER_ROW_FORMATO, c).value
        if val:
            ultimo_formato = _normalize_formato(val)
        if ultimo_formato:
            formatos[c] = ultimo_formato
    return formatos


def _detectar_columnas(ws: Worksheet) -> list[ResumenColumn]:
    """Detecta las columnas de datos combinando fila 7 (FORMATO) + fila 8 (UNIDADES)."""
    formato_por_col = _resolve_formato_por_columna(ws)
    columnas: list[ResumenColumn] = []
    for c in range(COL_DATA_START, COL_DATA_END + 1):
        unidades_raw = ws.cell(HEADER_ROW_UNIDADES, c).value
        if unidades_raw is None:
            continue
        try:
            unidades = int(unidades_raw)
        except (TypeError, ValueError):
            continue
        formato = formato_por_col.get(c)
        if not formato:
            continue
        columnas.append(
            ResumenColumn(
                excel_col=c, formato=formato, unidades_por_empaque=unidades
            )
        )
    return columnas


def _detectar_fila_fin(ws: Worksheet) -> int:
    """Retorna la ultima fila de datos (inclusive) antes de TOTALES."""
    for r in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(r, COL_REFERENCIA).value
        if val and _normalize_referencia(val) in STOP_MARKERS_COL_A:
            return r - 1
    return ws.max_row


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ("-", ".", "#VALUE!", "#N/A", "#DIV/0!", "#REF!"):
        return 0.0
    s = s.replace(",", "").replace("$", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_resumen_worksheet(ws: Worksheet) -> pd.DataFrame:
    """Parsea la hoja RESUMEN y retorna un DataFrame largo.

    Columnas:
        referencia, tipo, formato, unidades_por_empaque, bandejas, unidades_totales

    Solo se emiten filas con bandejas > 0.
    """
    columnas = _detectar_columnas(ws)
    if not columnas:
        raise ValueError("No se detectaron columnas de datos en la hoja RESUMEN")

    fila_fin = _detectar_fila_fin(ws)

    filas: list[dict[str, Any]] = []
    referencia_actual: str | None = None
    for r in range(DATA_START_ROW, fila_fin + 1):
        ref_raw = ws.cell(r, COL_REFERENCIA).value
        tipo_raw = ws.cell(r, COL_TIPO).value

        if ref_raw:
            referencia_actual = _normalize_referencia(ref_raw)

        tipo = _normalize_tipo(tipo_raw)
        if not tipo or referencia_actual is None:
            continue

        for col in columnas:
            valor = ws.cell(r, col.excel_col).value
            bandejas = _to_float(valor)
            if bandejas <= 0:
                continue
            filas.append(
                {
                    "referencia": referencia_actual,
                    "tipo": tipo,
                    "formato": col.formato,
                    "unidades_por_empaque": col.unidades_por_empaque,
                    "bandejas": bandejas,
                    "unidades_totales": bandejas * col.unidades_por_empaque,
                }
            )

    df = pd.DataFrame(
        filas,
        columns=[
            "referencia",
            "tipo",
            "formato",
            "unidades_por_empaque",
            "bandejas",
            "unidades_totales",
        ],
    )
    return df.reset_index(drop=True)


def load_resumen(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Carga la hoja RESUMEN de un .xlsx PRE CORTE.

    Retorna `(df_long, meta)` donde meta incluye numero de filas emitidas,
    total de unidades sumadas (para reconciliar con NOTIFICACION), y el
    listado de referencias distintas.
    """
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"RESUMEN requiere .xlsx; recibido '{path.suffix}'. "
            "CSV/TSV pierden los merged cells de REFERENCIA y FORMATO."
        )
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if "RESUMEN" not in wb.sheetnames:
        raise ValueError(
            f"El archivo '{path.name}' no contiene la hoja RESUMEN. "
            f"Hojas presentes: {wb.sheetnames}"
        )
    ws = wb["RESUMEN"]
    df = parse_resumen_worksheet(ws)
    meta = {
        "filename": path.name,
        "hoja": "RESUMEN",
        "num_filas_emitidas": len(df),
        "total_bandejas": float(df["bandejas"].sum()) if not df.empty else 0.0,
        "total_unidades": float(df["unidades_totales"].sum()) if not df.empty else 0.0,
        "referencias_distintas": sorted(df["referencia"].unique().tolist()) if not df.empty else [],
        "formatos_distintos": sorted(df["formato"].unique().tolist()) if not df.empty else [],
    }
    wb.close()
    return df, meta
