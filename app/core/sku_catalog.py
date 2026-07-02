"""Catalogo persistente SKU: puente entre RESUMEN y codigo SAP.

Contenido:

- **import_from_homologacion(path)**: carga el archivo maestro
  `homologacion materiales Nuevo.xlsx` (hojas Hoja2(2) y Ata). Solo importa
  tuplas (referencia, tipo, formato, unidades) no ambiguas.
- **update_catalog_from_pair(conn, resumen_df, notif_df)**: cuando el mismo
  archivo trae RESUMEN + NOTIFICACION, aparea filas por `bandejas ==
  necesidad_bandeja` y aprende SAPs autoritativos.
- **resolve_sap(conn, referencia, tipo, formato, unidades)**: lookup directo
  con normalizacion + aliases de FORMATO (ESTUCHE<->ESTUCHERIA,
  VITAFILM<->TERMOENCOGIDO). Prioriza `aprendido_pair`/`manual` sobre
  `homologacion` cuando hay entradas duplicadas.
- **attach_sap_to_resumen(conn, resumen_df)**: enriquece el DataFrame con
  columna `material_sap`; separa filas irresolubles.
- **CLI**: `python -m app.core.sku_catalog --import-homologacion <xlsx>` y
  `--backfill data/uploads/*.xlsx`.

Prioridad de fuentes al haber conflicto:
    manual (3) > aprendido_pair (2) > homologacion (1)
"""
from __future__ import annotations

import argparse
import glob
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from app.config import DB_PATH
from app.core.db import get_conn, init_db


FORMATO_ALIASES: dict[str, str] = {
    "ESTUCHERIA": "ESTUCHE",
    "ESTUCHE": "ESTUCHE",
    "TERMOENCOGIDO": "VITAFILM",
    "VITAFILM": "VITAFILM",
    "AMARRADO": "AMARRADO",
    "SUELTO": "SUELTO",
    "PASTER": "PASTER",
    "OTRAS VENTAS": "OTRAS",
}

FUENTE_PRIORITY: dict[str, int] = {
    "homologacion": 1,
    "aprendido_pair": 2,
    "manual": 3,
}

CLASE_TOKENS: set[str] = {"A", "AA", "AAA", "AAAA", "B", "C", "AA-A", "AAA-A"}

BRAND_KEYWORDS: tuple[str, ...] = (
    "CAMPESINO", "SELENIO", "JUNIOR", "KOSHER", "DHA", "PLUS",
    "TAEQ", "OXXO", "OLIMPICA", "TAT", "MAKRO", "ATA", "CEREALES",
    "SURTIMAX", "COLSUBSIDIO", "MAXIMA", "FRESCAMPO", "EKONO",
)


def _normalize_formato(raw: str | None) -> str:
    if not raw:
        return ""
    key = str(raw).strip().upper()
    return FORMATO_ALIASES.get(key, key)


def _normalize_tipo(raw: str | None) -> str:
    if not raw:
        return ""
    return str(raw).strip().upper().replace(" ", "").replace("/", "-")


def _normalize_referencia(raw: str | None) -> str:
    if not raw:
        return ""
    return str(raw).strip().upper()


def _int_or_none(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _referencia_from_homologacion(tipo1: Any, flash: Any, material: Any) -> str:
    """Extrae REFERENCIA que coincide con la convencion del RESUMEN.

    Prioridad:
    1. Palabra clave de marca (CAMPESINO, PLUS, KOSHER, ...) encontrada en el
       nombre del material.
    2. FLASH si no es solo una clase (A/AA/AAA/AAAA/B/C).
    3. Tipo 1 como fallback (MARCA ORO, MARCAS PROPIAS, MARCA ATA).
    """
    material_upper = str(material or "").upper()
    for brand in BRAND_KEYWORDS:
        if brand in material_upper:
            return brand

    if flash:
        flash_norm = str(flash).strip().upper()
        if flash_norm and flash_norm not in CLASE_TOKENS:
            return flash_norm

    return _normalize_referencia(tipo1)


def upsert_entry(
    conn: sqlite3.Connection,
    *,
    referencia: str,
    tipo: str,
    formato: str,
    unidades_por_empaque: int,
    material_sap: int,
    nombre_notificacion: str | None = None,
    fuente: str = "aprendido_pair",
) -> tuple[int, bool]:
    """Inserta o actualiza una entrada respetando prioridad por `fuente`.

    Reglas:
    - Si ya existe con mismo SAP: incrementa `veces_visto`.
    - Si existe con SAP distinto y `fuente` nueva tiene mayor prioridad:
      sobrescribe SAP y fuente.
    - Si existe con SAP distinto y prioridad menor o igual: NO sobrescribe.
    Retorna (id, es_nueva).
    """
    ref = _normalize_referencia(referencia)
    tp = _normalize_tipo(tipo)
    fmt = _normalize_formato(formato)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    existing = conn.execute(
        """
        SELECT id, material_sap, fuente FROM sku_catalog
        WHERE referencia = ? AND tipo = ? AND formato = ?
          AND unidades_por_empaque = ?
        """,
        (ref, tp, fmt, int(unidades_por_empaque)),
    ).fetchone()

    if existing:
        if int(existing["material_sap"]) == int(material_sap):
            conn.execute(
                "UPDATE sku_catalog SET ultima_vez_visto = ?, veces_visto = veces_visto + 1 WHERE id = ?",
                (now, existing["id"]),
            )
            return int(existing["id"]), False
        old_prio = FUENTE_PRIORITY.get(str(existing["fuente"]), 0)
        new_prio = FUENTE_PRIORITY.get(fuente, 0)
        if new_prio > old_prio:
            conn.execute(
                """
                UPDATE sku_catalog SET material_sap = ?, fuente = ?, nombre_notificacion = ?,
                       ultima_vez_visto = ?, veces_visto = veces_visto + 1
                WHERE id = ?
                """,
                (int(material_sap), fuente, nombre_notificacion, now, existing["id"]),
            )
        return int(existing["id"]), False

    cur = conn.execute(
        """
        INSERT INTO sku_catalog (
            referencia, tipo, formato, unidades_por_empaque, material_sap,
            nombre_notificacion, fuente, primera_vez_visto, ultima_vez_visto
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            ref, tp, fmt, int(unidades_por_empaque), int(material_sap),
            nombre_notificacion, fuente, now, now,
        ),
    )
    return int(cur.lastrowid), True


def resolve_sap(
    conn: sqlite3.Connection,
    referencia: str,
    tipo: str,
    formato: str,
    unidades: int,
) -> int | None:
    ref = _normalize_referencia(referencia)
    tp = _normalize_tipo(tipo)
    fmt = _normalize_formato(formato)
    row = conn.execute(
        """
        SELECT material_sap FROM sku_catalog
        WHERE referencia = ? AND tipo = ? AND formato = ?
          AND unidades_por_empaque = ?
        """,
        (ref, tp, fmt, int(unidades)),
    ).fetchone()
    return int(row["material_sap"]) if row else None


def attach_sap_to_resumen(
    conn: sqlite3.Connection, resumen_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Divide `resumen_df` en (con_sap, sin_sap) segun cobertura del catalogo."""
    if resumen_df.empty:
        cols = list(resumen_df.columns) + ["material_sap"]
        return pd.DataFrame(columns=cols), pd.DataFrame(columns=cols)

    df = resumen_df.copy()
    df["material_sap"] = df.apply(
        lambda r: resolve_sap(
            conn, r["referencia"], r["tipo"], r["formato"], r["unidades_por_empaque"]
        ),
        axis=1,
    )
    con_sap = df[df["material_sap"].notna()].copy()
    con_sap["material_sap"] = con_sap["material_sap"].astype(int)
    sin_sap = df[df["material_sap"].isna()].copy()
    return (
        con_sap.reset_index(drop=True),
        sin_sap.reset_index(drop=True),
    )


# ---------- Homologacion (catalogo maestro externo) ----------

_HOMOLOG_SHEETS = ("Hoja2 (2)", "Ata")


def _iter_homologacion_rows(path: Path):
    """Emite (sap, material, und, tipo, clase, tipo1, flash) desde las hojas relevantes."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet_name in _HOMOLOG_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            codigo = row[0] if len(row) > 0 else None
            material = row[1] if len(row) > 1 else None
            und = row[2] if len(row) > 2 else None
            tipo = row[3] if len(row) > 3 else None
            clase = row[4] if len(row) > 4 else None
            tipo1 = row[5] if len(row) > 5 else None
            flash = row[6] if len(row) > 6 else None

            if sheet_name == "Ata" and codigo is None:
                codigo = row[1] if len(row) > 1 else None
                material = row[2] if len(row) > 2 else None
                und = row[3] if len(row) > 3 else None
                tipo = row[4] if len(row) > 4 else None
                clase = row[5] if len(row) > 5 else None
                tipo1 = row[6] if len(row) > 6 else None
                flash = None

            sap = _int_or_none(codigo)
            und_int = _int_or_none(und)
            if not sap or not und_int or not tipo or not clase:
                continue
            yield (
                sap,
                str(material or ""),
                und_int,
                str(tipo),
                str(clase),
                str(tipo1) if tipo1 else "",
                str(flash) if flash else "",
            )
    wb.close()


def import_from_homologacion(
    path: str | Path,
    conn: sqlite3.Connection | None = None,
) -> dict[str, int]:
    """Carga la homologacion externa como catalogo maestro.

    Solo importa tuplas (referencia, tipo, formato, unidades) que tengan
    UN SOLO SAP asociado. Las tuplas ambiguas (multiples SAPs) se dejan
    para que pair-learn las resuelva con la NOTIFICACION del run actual.
    """
    path = Path(path)
    stats = {
        "leidas": 0,
        "insertadas": 0,
        "ya_existian": 0,
        "ambiguas_no_insertadas": 0,
    }

    # Primera pasada: agrupar por tupla RESUMEN
    grouped: dict[tuple, set[tuple[int, str]]] = defaultdict(set)
    for sap, material, und, tipo, clase, tipo1, flash in _iter_homologacion_rows(path):
        stats["leidas"] += 1
        ref = _referencia_from_homologacion(tipo1, flash, material)
        key = (
            _normalize_referencia(ref),
            _normalize_tipo(clase),
            _normalize_formato(tipo),
            und,
        )
        grouped[key].add((sap, material))

    def _run(conn_: sqlite3.Connection) -> None:
        for key, candidatos in grouped.items():
            saps = {c[0] for c in candidatos}
            if len(saps) > 1:
                stats["ambiguas_no_insertadas"] += 1
                continue
            sap, material = next(iter(candidatos))
            ref, tp, fmt, und = key
            _, es_nueva = upsert_entry(
                conn_,
                referencia=ref,
                tipo=tp,
                formato=fmt,
                unidades_por_empaque=und,
                material_sap=sap,
                nombre_notificacion=material,
                fuente="homologacion",
            )
            if es_nueva:
                stats["insertadas"] += 1
            else:
                stats["ya_existian"] += 1
        conn_.commit()

    if conn is None:
        with get_conn() as c:
            _run(c)
    else:
        _run(conn)
    return stats


# ---------- Aprendizaje desde el par RESUMEN + NOTIFICACION ----------

def update_catalog_from_pair(
    conn: sqlite3.Connection,
    resumen_df: pd.DataFrame,
    notif_df: pd.DataFrame,
) -> dict[str, int]:
    """Aparea RESUMEN <-> NOTIFICACION por bandejas para aprender SAPs autoritativos.

    Cada celda no-cero del RESUMEN tiene bandejas y unidades_totales.
    NOTIFICACION tiene `necesidad_bandeja` y `necesidad_unidades` por SAP.
    Si (bandejas, unidades) del RESUMEN coincide con UNA UNICA fila de
    NOTIFICACION, se aprende el mapping.
    """
    stats = {"aprendidas": 0, "sobrescritas": 0, "ambiguas": 0, "sin_match": 0, "ya_correctas": 0}
    if resumen_df.empty or notif_df.empty:
        return stats

    notif_valid = notif_df[
        notif_df["necesidad_bandeja"].notna()
        & (notif_df["necesidad_bandeja"] > 0)
    ].copy()

    for _, row in resumen_df.iterrows():
        bandejas = float(row["bandejas"])
        if bandejas <= 0:
            continue
        unidades = float(row["unidades_totales"])

        candidatos = notif_valid[
            (notif_valid["necesidad_bandeja"] == bandejas)
            & ((notif_valid["necesidad_unidades"] - unidades).abs() < 0.5)
        ]
        if len(candidatos) == 0:
            stats["sin_match"] += 1
            continue
        if len(candidatos) > 1:
            stats["ambiguas"] += 1
            continue

        cand = candidatos.iloc[0]
        sap_correcto = int(cand["material"])

        existing_sap = resolve_sap(
            conn,
            row["referencia"],
            row["tipo"],
            row["formato"],
            int(row["unidades_por_empaque"]),
        )
        if existing_sap == sap_correcto:
            stats["ya_correctas"] += 1
            continue

        upsert_entry(
            conn,
            referencia=row["referencia"],
            tipo=row["tipo"],
            formato=row["formato"],
            unidades_por_empaque=int(row["unidades_por_empaque"]),
            material_sap=sap_correcto,
            nombre_notificacion=str(cand.get("referencia", "")),
            fuente="aprendido_pair",
        )
        if existing_sap is None:
            stats["aprendidas"] += 1
        else:
            stats["sobrescritas"] += 1

    conn.commit()
    return stats


# ---------- Consultas para API/UI ----------

def list_catalog(
    conn: sqlite3.Connection,
    referencia: str | None = None,
    tipo: str | None = None,
    formato: str | None = None,
    limit: int = 500,
) -> list[dict[str, Any]]:
    filtros: list[str] = []
    params: list[Any] = []
    if referencia:
        filtros.append("referencia = ?")
        params.append(_normalize_referencia(referencia))
    if tipo:
        filtros.append("tipo = ?")
        params.append(_normalize_tipo(tipo))
    if formato:
        filtros.append("formato = ?")
        params.append(_normalize_formato(formato))
    where = "WHERE " + " AND ".join(filtros) if filtros else ""
    params.append(limit)
    rows = conn.execute(
        f"""
        SELECT * FROM sku_catalog {where}
        ORDER BY referencia, tipo, formato, unidades_por_empaque
        LIMIT ?
        """,
        tuple(params),
    ).fetchall()
    return [dict(r) for r in rows]


def catalog_stats(conn: sqlite3.Connection) -> dict[str, Any]:
    total = conn.execute("SELECT COUNT(*) AS n FROM sku_catalog").fetchone()["n"]
    por_fuente_rows = conn.execute(
        "SELECT fuente, COUNT(*) AS n FROM sku_catalog GROUP BY fuente"
    ).fetchall()
    return {
        "total_entradas": int(total),
        "por_fuente": {r["fuente"]: int(r["n"]) for r in por_fuente_rows},
    }


# ---------- CLI ----------

def _cli(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Utilidades del catalogo SKU.")
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument(
        "--import-homologacion",
        metavar="PATH",
        help="Importa el archivo homologacion materiales Nuevo.xlsx",
    )
    parser.add_argument(
        "--backfill",
        nargs="+",
        metavar="PRE_CORTE_XLSX",
        help="Rutas o glob de PRE CORTE .xlsx para aprender de sus pares RESUMEN+NOTIFICACION",
    )
    parser.add_argument("--stats", action="store_true", help="Imprime stats del catalogo")
    parser.add_argument(
        "--list", action="store_true", help="Lista contenido del catalogo (hasta 500)"
    )
    args = parser.parse_args(argv)

    init_db(args.db)

    if args.import_homologacion:
        with get_conn(args.db) as conn:
            stats = import_from_homologacion(args.import_homologacion, conn)
        print(f"import_homologacion: {stats}")

    if args.backfill:
        from app.core.loaders import load_notificacion
        from app.core.resumen_parser import load_resumen

        paths: list[str] = []
        for pat in args.backfill:
            paths.extend(sorted(glob.glob(pat)))
        with get_conn(args.db) as conn:
            for p in paths:
                try:
                    r_df, _ = load_resumen(p)
                    n_df, _ = load_notificacion(p)
                    s = update_catalog_from_pair(conn, r_df, n_df)
                    print(f"  {p}: {s}")
                except Exception as exc:
                    print(f"  {p}: ERROR {exc}")

    if args.stats:
        with get_conn(args.db) as conn:
            print(catalog_stats(conn))

    if args.list:
        with get_conn(args.db) as conn:
            for r in list_catalog(conn):
                print(
                    f"  [{r['fuente']:16s}] {r['referencia']:15s} {r['tipo']:6s} "
                    f"{r['formato']:12s} x{r['unidades_por_empaque']:<3d} -> SAP {r['material_sap']}"
                )

    return 0


if __name__ == "__main__":
    sys.exit(_cli())
