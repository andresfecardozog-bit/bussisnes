"""Carga y normalizacion de los archivos PRE CORTE y FLASH.

Cambio importante 2026-07-02: la fuente autoritativa del PRE CORTE es ahora
la hoja RESUMEN (orden de comercializacion para el dia siguiente), no la
hoja NOTIFICACION (notificacion de bodega). NOTIFICACION queda como fuente
opcional para aprender el mapping (referencia, tipo, formato, unidades) ->
SAP MATERIAL y validar el total.

Reglas de intake:
- PRE CORTE: solo .xlsx / .xlsm (CSV destruye los merged cells del RESUMEN).
- FLASH: xlsx o csv como antes.

Objetivos originales que se mantienen:
- Retornar siempre `(DataFrame, meta_dict)` con hash SHA256, num_filas, etc.
- Sanidad numerica (comas de miles, simbolos de moneda) via
  `_clean_numeric_series`.
- Validacion independiente (`validators.py`) usando el archivo original.
"""
from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from app.config import (
    DATE_FORMAT_FLASH,
    FLASH_COLUMNS,
    FLASH_DATE_COLUMN,
    FLASH_KEY_MATERIAL,
    FLASH_NUMERIC_COLUMNS,
)
from app.core.db import get_conn
from app.core.resumen_parser import load_resumen
from app.core.sku_catalog import (
    attach_sap_to_resumen,
    update_catalog_from_pair,
)


_CURRENCY_STRIP_RE = re.compile(r"[\s$,]")
_EXCEL_ERRORS = {"#VALUE!", "#N/A", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!"}


def hash_file(path: str | Path) -> str:
    path = Path(path)
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65_536), b""):
            h.update(chunk)
    return h.hexdigest()


def _clean_numeric_series(series: pd.Series) -> pd.Series:
    """Convierte strings con separador de miles/moneda/errores Excel a float."""
    if pd.api.types.is_numeric_dtype(series):
        return series.fillna(0).astype(float)

    def _clean(v: Any) -> Any:
        if pd.isna(v):
            return 0
        s = str(v).strip()
        if not s or s in ("-", "-.", "."):
            return 0
        if s in _EXCEL_ERRORS:
            return 0
        s = _CURRENCY_STRIP_RE.sub("", s)
        if not s or s == "-":
            return 0
        return s

    cleaned = series.apply(_clean)
    numeric = pd.to_numeric(cleaned, errors="coerce")
    return numeric.fillna(0).astype(float)


def _rename_and_keep(df: pd.DataFrame, mapping: dict[str, str]) -> pd.DataFrame:
    present = {orig: new for orig, new in mapping.items() if orig in df.columns}
    missing = [orig for orig in mapping if orig not in df.columns]
    if missing:
        raise ValueError(
            f"Faltan columnas obligatorias en el archivo: {missing}. "
            f"Columnas presentes: {df.columns.tolist()}"
        )
    return df[list(present.keys())].rename(columns=present)


def _read_any(path: Path) -> pd.DataFrame:
    """Lee xlsx o csv segun la extension. Para xlsx usa la primera hoja."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, dtype=object)
    if suffix == ".csv":
        return pd.read_csv(path, dtype=object)
    raise ValueError(f"Extension no soportada: '{suffix}' ({path.name})")


# ---------- NOTIFICACION (fuente secundaria del PRE CORTE) ----------

_NOTIFICACION_COLUMNS = {
    "MATERIAL": "material",
    "REFERENCIA": "referencia",
    "NECESIDAD BANDEJA": "necesidad_bandeja",
    "NECESIDAD UNIDADES": "necesidad_unidades",
    "FISICO BANDEJAS": "fisico_bandejas",
    "FISICO UNIDADES": "fisico_unidades",
    "PRODUCIR BANDEJAS": "producir_bandeja",
    "PRODUCIR UNIDADES": "producir_unidades",
    "NOTIFICADO": "notificado",
}
_NOTIFICACION_NUMERIC = [
    "necesidad_bandeja", "necesidad_unidades",
    "fisico_bandejas", "fisico_unidades",
    "producir_bandeja", "producir_unidades",
    "notificado",
]


def _find_notificacion_header_row(rows: list[list[Any]]) -> int:
    """Localiza la fila del header en la hoja NOTIFICACION (busca 'MATERIAL')."""
    for idx, row in enumerate(rows):
        vals = [str(c).strip().upper() if c is not None else "" for c in row]
        if "MATERIAL" in vals and "REFERENCIA" in vals:
            return idx
    raise ValueError("No se encontro fila de header en la hoja NOTIFICACION")


def _notificacion_from_worksheet(path: Path) -> pd.DataFrame:
    """Lee la hoja NOTIFICACION del .xlsx con openpyxl, tolera #VALUE!."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if "NOTIFICACION" not in wb.sheetnames:
            raise ValueError(
                f"{path.name} no tiene hoja NOTIFICACION. Hojas: {wb.sheetnames}"
            )
        ws = wb["NOTIFICACION"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return pd.DataFrame(columns=list(_NOTIFICACION_COLUMNS.values()))

    header_row_idx = _find_notificacion_header_row(rows)
    headers_raw = rows[header_row_idx]
    headers = [
        str(c).strip() if c is not None else f"_col_{i}"
        for i, c in enumerate(headers_raw)
    ]
    data_rows = rows[header_row_idx + 1 :]
    df = pd.DataFrame(data_rows, columns=headers)

    for expected in _NOTIFICACION_COLUMNS:
        if expected not in df.columns:
            df[expected] = pd.NA

    df = _rename_and_keep(df, _NOTIFICACION_COLUMNS)

    df["material"] = pd.to_numeric(df["material"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["material"]).copy()
    df["material"] = df["material"].astype(int)

    for col in _NOTIFICACION_NUMERIC:
        df[col] = _clean_numeric_series(df[col])

    df["referencia"] = df["referencia"].astype(str).str.strip()
    return df.reset_index(drop=True)


def load_notificacion(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Carga la hoja NOTIFICACION del .xlsx. Retorna `(df, meta)`."""
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"NOTIFICACION requiere .xlsx; recibido '{path.suffix}'"
        )
    df = _notificacion_from_worksheet(path)
    meta = {
        "filename": path.name,
        "hoja": "NOTIFICACION",
        "num_filas": len(df),
        "total_necesidad_bandeja": float(df["necesidad_bandeja"].sum()) if not df.empty else 0.0,
        "total_necesidad_unidades": float(df["necesidad_unidades"].sum()) if not df.empty else 0.0,
    }
    return df, meta


# ---------- PRE CORTE (orquestador RESUMEN + NOTIFICACION + catalogo) ----------

def load_pre_corte(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Carga PRE CORTE desde .xlsx usando RESUMEN como fuente autoritativa.

    Pipeline interno:
    1. Parsea RESUMEN (obligatorio).
    2. Si el .xlsx trae NOTIFICACION -> lee, aprende SAPs (pair-learn) y usa
       como validacion cruzada.
    3. Resuelve SAP para cada fila de RESUMEN desde `sku_catalog`.
    4. Filas sin SAP quedan reportadas en `meta['sin_sap']` (df aparte).

    Retorna un DataFrame con las columnas que el resto del pipeline consume:
        material, referencia, notificado, producir_unidades,
        necesidad_bandeja, necesidad_unidades, fisico_bandejas,
        fisico_unidades, producir_bandeja
    """
    path = Path(path)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"PRE CORTE requiere .xlsx; recibido '{path.suffix}'. "
            "El CSV destruye los merged cells del RESUMEN."
        )

    resumen_df, resumen_meta = load_resumen(path)

    notif_df: pd.DataFrame | None = None
    notif_meta: dict[str, Any] | None = None
    pair_learn_stats: dict[str, int] | None = None
    wb_probe = openpyxl.load_workbook(path, data_only=True, read_only=True)
    has_notificacion = "NOTIFICACION" in wb_probe.sheetnames
    wb_probe.close()

    if has_notificacion:
        notif_df, notif_meta = load_notificacion(path)
        with get_conn() as conn:
            pair_learn_stats = update_catalog_from_pair(conn, resumen_df, notif_df)

    with get_conn() as conn:
        con_sap, sin_sap = attach_sap_to_resumen(conn, resumen_df)

    df_out = pd.DataFrame(
        {
            "material": con_sap["material_sap"].astype(int) if not con_sap.empty else pd.Series(dtype=int),
            "referencia": con_sap["referencia"] if not con_sap.empty else pd.Series(dtype=str),
            "notificado": con_sap["unidades_totales"] if not con_sap.empty else pd.Series(dtype=float),
            "producir_unidades": con_sap["unidades_totales"] if not con_sap.empty else pd.Series(dtype=float),
            "necesidad_bandeja": con_sap["bandejas"] if not con_sap.empty else pd.Series(dtype=float),
            "necesidad_unidades": con_sap["unidades_totales"] if not con_sap.empty else pd.Series(dtype=float),
            "fisico_bandejas": [0.0] * len(con_sap) if not con_sap.empty else pd.Series(dtype=float),
            "fisico_unidades": [0.0] * len(con_sap) if not con_sap.empty else pd.Series(dtype=float),
            "producir_bandeja": con_sap["bandejas"] if not con_sap.empty else pd.Series(dtype=float),
            "formato": con_sap["formato"] if not con_sap.empty else pd.Series(dtype=str),
            "tipo": con_sap["tipo"] if not con_sap.empty else pd.Series(dtype=str),
            "unidades_por_empaque": (
                con_sap["unidades_por_empaque"].astype(int) if not con_sap.empty else pd.Series(dtype=int)
            ),
        }
    )
    if df_out.empty:
        df_out = pd.DataFrame(columns=[
            "material", "referencia", "notificado", "producir_unidades",
            "necesidad_bandeja", "necesidad_unidades",
            "fisico_bandejas", "fisico_unidades", "producir_bandeja",
            "formato", "tipo", "unidades_por_empaque",
        ])

    meta: dict[str, Any] = {
        "filename": path.name,
        "path": str(path),
        "hash_sha256": hash_file(path),
        "num_filas_original": resumen_meta["num_filas_emitidas"],
        "num_filas_procesadas": len(df_out),
        "num_filas_sin_sap": len(sin_sap),
        "tipo": "pre_corte",
        "fuente_primaria": "RESUMEN",
        "notificacion_presente": has_notificacion,
        "resumen_total_unidades": resumen_meta["total_unidades"],
        "resumen_total_bandejas": resumen_meta["total_bandejas"],
        "sin_sap_detalle": sin_sap.to_dict(orient="records") if not sin_sap.empty else [],
    }
    if notif_meta is not None:
        meta["notificacion_total_unidades"] = notif_meta["total_necesidad_unidades"]
        meta["notificacion_total_bandejas"] = notif_meta["total_necesidad_bandeja"]
    if pair_learn_stats is not None:
        meta["pair_learn_stats"] = pair_learn_stats

    return df_out.reset_index(drop=True), meta


# ---------- FLASH (sin cambios) ----------

def load_flash(path: str | Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    path = Path(path)
    raw = _read_any(path)
    num_filas_original = len(raw)

    df = _rename_and_keep(raw, FLASH_COLUMNS)

    df[FLASH_KEY_MATERIAL] = (
        df[FLASH_KEY_MATERIAL].astype(str).str.strip().str.lstrip("0").replace("", "0")
    )
    df[FLASH_KEY_MATERIAL] = pd.to_numeric(df[FLASH_KEY_MATERIAL], errors="coerce")
    df = df.dropna(subset=[FLASH_KEY_MATERIAL]).copy()
    df[FLASH_KEY_MATERIAL] = df[FLASH_KEY_MATERIAL].astype("int64")

    df[FLASH_DATE_COLUMN] = pd.to_datetime(
        df[FLASH_DATE_COLUMN], format=DATE_FORMAT_FLASH, errors="coerce"
    ).dt.date

    for col in FLASH_NUMERIC_COLUMNS:
        df[col] = _clean_numeric_series(df[col])

    if "nomb_material" in df.columns:
        df["nomb_material"] = df["nomb_material"].astype(str).str.strip()

    meta = {
        "filename": path.name,
        "path": str(path),
        "hash_sha256": hash_file(path),
        "num_filas_original": num_filas_original,
        "num_filas_procesadas": len(df),
        "tipo": "flash",
    }
    return df.reset_index(drop=True), meta
